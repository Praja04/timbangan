"""
Aplikasi GUI Pembaca Timbangan AND GX-4000
Fitur: Pilih Mesin & Variant, Start/Stop, Simpan ke Excel

Requirements:
pip install pyserial openpyxl

Author: [Your Name]
Date: 2025-12-18
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import serial
import serial.tools.list_ports
import threading
import queue
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re


class ScaleGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Timbangan AND GX-4000 - Data Logger with Machine & Variant")
        self.root.geometry("1000x750")
        
        # Variables
        self.serial_conn = None
        self.is_reading = False
        self.data_queue = queue.Queue()
        self.current_data = None
        self.saved_data = []  # Data yang sudah di-klik simpan
        
        # Serial settings untuk AND GX-4000
        self.baudrate = 9600
        self.databits = 7
        self.parity = 'E'  # Even
        self.stopbits = 1
        
        # Machine and Variant lists
        self.machines = [
            "A (F2)", "AE (D12)", "AF (D13)", "AG (D14)", "AH (D15)", "AI (D16)",
            "AJ (D17)", "AK (D18)", "B (D11/E5)", "C (D9)", "D (D1)", "E (D2)",
            "F (D3)", "G (D4)", "H (D5)", "I (D6)", "J (D7)", "K (D8)",
            "L (LD10)", "O (C1)", "P (C2)", "Q (A2)", "R (C3)", "U (F3)",
            "V (F1)", "W (C7)", "X (C8)", "Y (B6)", "Z (B3)"
        ]
        
        self.variants = [
            "VARIANT YB P1000G",
            "VARIANT YB P700G PIRING",
            "VARIANT BB P700G NP HARGA",
            "VARIANT BB P270G",
            "VARIANT BB P725G",
            "VARIANT YB P700G",
            "VARIANT BB P77G HARGA BDKT",
            "VARIANT BB P77G HARGA",
            "VARIANT YB P77G B5G1",
            "VARIANT YB P77G B5G1 BDKT",
            "VARIANT YB P250G"
        ]
        
        # Setup GUI
        self.setup_ui()
        
        # Load COM ports
        self.refresh_ports()
        
        # Start queue checker
        self.check_queue()
    
    def setup_ui(self):
        """Setup UI components"""
        
        # ===== TOP FRAME: Connection Settings =====
        top_frame = ttk.LabelFrame(self.root, text="Koneksi Serial", padding=10)
        top_frame.pack(fill="x", padx=10, pady=5)
        
        # COM Port selection
        ttk.Label(top_frame, text="COM Port:").grid(row=0, column=0, sticky="w", padx=5)
        self.port_var = tk.StringVar()
        self.port_combo = ttk.Combobox(top_frame, textvariable=self.port_var, width=15, state="readonly")
        self.port_combo.grid(row=0, column=1, padx=5)
        
        ttk.Button(top_frame, text="Refresh", command=self.refresh_ports, width=10).grid(row=0, column=2, padx=5)
        
        # Settings display
        settings_text = f"Setting: {self.baudrate} bps, {self.databits} bit, Parity Even"
        ttk.Label(top_frame, text=settings_text, foreground="gray").grid(row=0, column=3, padx=20)
        
        # Connection status
        self.status_label = ttk.Label(top_frame, text="● Disconnected", foreground="red", font=("Arial", 10, "bold"))
        self.status_label.grid(row=0, column=4, padx=10)
        
        # ===== MACHINE & VARIANT FRAME =====
        selection_frame = ttk.LabelFrame(self.root, text="⚙️ Pilih Mesin & Variant", padding=10)
        selection_frame.pack(fill="x", padx=10, pady=5)
        
        # Machine selection
        ttk.Label(selection_frame, text="Mesin:", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.machine_var = tk.StringVar()
        self.machine_combo = ttk.Combobox(
            selection_frame, 
            textvariable=self.machine_var, 
            values=self.machines,
            width=20,
            state="readonly",
            font=("Arial", 10)
        )
        self.machine_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.machine_combo.set("-- Pilih Mesin --")
        
        # Variant selection
        ttk.Label(selection_frame, text="Variant:", font=("Arial", 10, "bold")).grid(row=0, column=2, sticky="w", padx=20, pady=5)
        self.variant_var = tk.StringVar()
        self.variant_combo = ttk.Combobox(
            selection_frame,
            textvariable=self.variant_var,
            values=self.variants,
            width=30,
            state="readonly",
            font=("Arial", 10)
        )
        self.variant_combo.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.variant_combo.set("-- Pilih Variant --")
        
        # Selection indicator
        self.selection_label = tk.Label(
            selection_frame,
            text="⚠ Pilih Mesin & Variant terlebih dahulu",
            font=("Arial", 9),
            fg="red"
        )
        self.selection_label.grid(row=1, column=0, columnspan=4, pady=5)
        
        # Bind selection change
        self.machine_combo.bind("<<ComboboxSelected>>", self.update_selection_status)
        self.variant_combo.bind("<<ComboboxSelected>>", self.update_selection_status)
        
        # ===== MIDDLE FRAME: Current Reading =====
        current_frame = ttk.LabelFrame(self.root, text="📊 Pembacaan Saat Ini", padding=15)
        current_frame.pack(fill="x", padx=10, pady=5)
        
        # Large weight display
        self.weight_label = tk.Label(
            current_frame, 
            text="0.00", 
            font=("Arial", 48, "bold"),
            fg="#2E7D32"
        )
        self.weight_label.pack()
        
        # Unit and status
        info_frame = tk.Frame(current_frame)
        info_frame.pack()
        
        self.unit_label = tk.Label(info_frame, text="gram", font=("Arial", 14))
        self.unit_label.pack(side="left", padx=10)
        
        self.status_indicator = tk.Label(
            info_frame, 
            text="●", 
            font=("Arial", 20),
            fg="gray"
        )
        self.status_indicator.pack(side="left", padx=5)
        
        self.status_text = tk.Label(info_frame, text="Waiting...", font=("Arial", 12))
        self.status_text.pack(side="left")
        
        # ===== CONTROL FRAME: Buttons =====
        control_frame = tk.Frame(self.root)
        control_frame.pack(fill="x", padx=10, pady=10)
        
        # Start button
        self.start_btn = tk.Button(
            control_frame,
            text="▶ START",
            command=self.start_reading,
            bg="#4CAF50",
            fg="white",
            font=("Arial", 12, "bold"),
            height=2,
            width=12
        )
        self.start_btn.pack(side="left", padx=5)
        
        # Stop button
        # self.stop_btn = tk.Button(
        #     control_frame,
        #     text="■ STOP",
        #     command=self.stop_reading,
        #     bg="#F44336",
        #     fg="white",
        #     font=("Arial", 12, "bold"),
        #     height=2,
        #     width=12,
        #     state="disabled"
        # )
        # self.stop_btn.pack(side="left", padx=5)
        
        # Save button
        self.save_btn = tk.Button(
            control_frame,
            text="💾 SIMPAN DATA INI",
            command=self.save_current_data,
            bg="#2196F3",
            fg="white",
            font=("Arial", 12, "bold"),
            height=2,
            width=18,
            state="disabled"
        )
        self.save_btn.pack(side="left", padx=5)
        
        # Export button
        self.export_btn = tk.Button(
            control_frame,
            text="📊 EXPORT EXCEL",
            command=self.export_to_excel,
            bg="#FF9800",
            fg="white",
            font=("Arial", 12, "bold"),
            height=2,
            width=15,
            state="disabled"
        )
        self.export_btn.pack(side="left", padx=5)
        
        # ===== DATA TABLE: Saved Data =====
        table_frame = ttk.LabelFrame(self.root, text="📋 Data Tersimpan", padding=10)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # Counter label
        self.counter_label = ttk.Label(table_frame, text="Total data: 0", font=("Arial", 10, "bold"))
        self.counter_label.pack(anchor="w")
        
        # Scrollbar
        scroll_y = ttk.Scrollbar(table_frame)
        scroll_y.pack(side="right", fill="y")
        
        scroll_x = ttk.Scrollbar(table_frame, orient="horizontal")
        scroll_x.pack(side="bottom", fill="x")
        
        # Treeview
        self.tree = ttk.Treeview(
            table_frame,
            columns=("No", "Mesin", "Variant", "Timestamp", "Status", "Weight", "Unit"),
            show="headings",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set,
            height=12
        )
        
        scroll_y.config(command=self.tree.yview)
        scroll_x.config(command=self.tree.xview)
        
        # Define columns
        self.tree.heading("No", text="No")
        self.tree.heading("Mesin", text="Mesin")
        self.tree.heading("Variant", text="Variant")
        self.tree.heading("Timestamp", text="Waktu")
        self.tree.heading("Status", text="Status")
        self.tree.heading("Weight", text="Berat")
        self.tree.heading("Unit", text="Unit")
        
        self.tree.column("No", width=50, anchor="center")
        self.tree.column("Mesin", width=100, anchor="center")
        self.tree.column("Variant", width=250, anchor="w")
        self.tree.column("Timestamp", width=150, anchor="center")
        self.tree.column("Status", width=70, anchor="center")
        self.tree.column("Weight", width=100, anchor="center")
        self.tree.column("Unit", width=60, anchor="center")
        
        self.tree.pack(fill="both", expand=True)
        
        # Delete button
        delete_frame = tk.Frame(table_frame)
        delete_frame.pack(fill="x", pady=5)
        
        ttk.Button(delete_frame, text="🗑️ Hapus Selected", command=self.delete_selected).pack(side="left", padx=5)
        ttk.Button(delete_frame, text="🗑️ Hapus Semua", command=self.clear_all_data).pack(side="left", padx=5)
        
        # ===== BOTTOM: Status Bar =====
        status_bar = tk.Frame(self.root, relief="sunken", borderwidth=1)
        status_bar.pack(fill="x", side="bottom")
        
        self.statusbar_label = tk.Label(status_bar, text="Ready - Pilih Mesin & Variant terlebih dahulu", anchor="w")
        self.statusbar_label.pack(fill="x", padx=5)
    
    def update_selection_status(self, event=None):
        """Update status when machine or variant is selected"""
        machine = self.machine_var.get()
        variant = self.variant_var.get()
        
        if machine != "-- Pilih Mesin --" and variant != "-- Pilih Variant --":
            self.selection_label.config(
                text=f"✓ Dipilih: {machine} | {variant}",
                fg="green"
            )
            self.statusbar_label.config(text=f"Ready - Mesin: {machine} | Variant: {variant}")
        else:
            self.selection_label.config(
                text="⚠ Pilih Mesin & Variant terlebih dahulu",
                fg="red"
            )
    
    def refresh_ports(self):
        """Refresh COM port list"""
        ports = serial.tools.list_ports.comports()
        port_list = [p.device for p in ports]
        
        self.port_combo['values'] = port_list
        
        if port_list:
            self.port_combo.current(0)
            self.statusbar_label.config(text=f"Found {len(port_list)} COM port(s) - Pilih Mesin & Variant terlebih dahulu")
        else:
            self.statusbar_label.config(text="No COM ports found")
    
    def start_reading(self):
        """Start reading from scale"""
        port = self.port_var.get()
        
        if not port:
            messagebox.showerror("Error", "Pilih COM port terlebih dahulu!")
            return
        
        try:
            # Open serial connection
            parity_map = {'N': serial.PARITY_NONE, 'E': serial.PARITY_EVEN, 'O': serial.PARITY_ODD}
            
            self.serial_conn = serial.Serial(
                port=port,
                baudrate=self.baudrate,
                bytesize=self.databits,
                parity=parity_map[self.parity],
                stopbits=self.stopbits,
                timeout=1
            )
            
            self.serial_conn.reset_input_buffer()
            self.is_reading = True
            
            # Update UI
            self.start_btn.config(state="disabled")
            self.stop_btn.config(state="normal")
            self.save_btn.config(state="normal")
            self.port_combo.config(state="disabled")
            
            self.status_label.config(text="● Connected", foreground="green")
            
            # Start reading thread
            thread = threading.Thread(target=self.read_thread, daemon=True)
            thread.start()
            
        except Exception as e:
            messagebox.showerror("Connection Error", f"Failed to connect:\n{str(e)}")
            self.stop_reading()
    
    def stop_reading(self):
        """Stop reading from scale"""
        self.is_reading = False
        
        if self.serial_conn and self.serial_conn.is_open:
            self.serial_conn.close()
        
        # Update UI
        self.start_btn.config(state="normal")
        self.stop_btn.config(state="disabled")
        self.save_btn.config(state="disabled")
        self.port_combo.config(state="readonly")
        
        self.status_label.config(text="● Disconnected", foreground="red")
        self.statusbar_label.config(text="Stopped")
        
        # Reset display
        self.weight_label.config(text="0.00", fg="gray")
        self.status_indicator.config(fg="gray")
        self.status_text.config(text="Stopped")
    
    def read_thread(self):
        """Background thread for reading serial data"""
        while self.is_reading:
            try:
                if self.serial_conn and self.serial_conn.in_waiting > 0:
                    raw_data = self.serial_conn.readline().decode('ascii', errors='ignore').strip()
                    
                    if raw_data:
                        parsed = self.parse_data(raw_data)
                        if parsed:
                            self.data_queue.put(parsed)
                
            except Exception as e:
                print(f"Read error: {e}")
                break
    
    def parse_data(self, raw_data):
        """Parse scale data - Format: ST,+00029.84  g"""
        try:
            # Pattern: ST,+00029.84  g
            pattern = r'^([A-Z]{2}),([+-]?\d+\.?\d*)\s*([a-zA-Z]+)$'
            match = re.match(pattern, raw_data)
            
            if match:
                status = match.group(1)
                weight = float(match.group(2))
                unit = match.group(3)
                
                return {
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'status': status,
                    'weight': weight,
                    'unit': unit,
                    'raw': raw_data
                }
        except Exception as e:
            print(f"Parse error: {e}")
        
        return None
    
    def check_queue(self):
        """Check data queue and update UI"""
        try:
            while True:
                data = self.data_queue.get_nowait()
                self.update_display(data)
        except queue.Empty:
            pass
        
        # Schedule next check
        self.root.after(50, self.check_queue)
    
    def update_display(self, data):
        """Update current reading display"""
        self.current_data = data
        
        # Update weight
        self.weight_label.config(text=f"{data['weight']:.2f}")
        
        # Update unit
        self.unit_label.config(text=data['unit'])
        
        # Update status
        if data['status'] == 'ST':
            self.weight_label.config(fg="#2E7D32")  # Green
            self.status_indicator.config(fg="#4CAF50")
            self.status_text.config(text="Stable")
        else:
            self.weight_label.config(fg="#FF9800")  # Orange
            self.status_indicator.config(fg="#FF9800")
            self.status_text.config(text="Unstable")
        
        # Update status bar
        machine = self.machine_var.get()
        variant = self.variant_var.get()
        self.statusbar_label.config(
            text=f"Last: {data['timestamp']} | {data['status']} | {data['weight']:.2f} {data['unit']} | {machine} | {variant}"
        )
    
    def save_current_data(self):
        """Save current displayed data to list"""
        if not self.current_data:
            messagebox.showwarning("No Data", "Tidak ada data untuk disimpan!")
            return
        
        # Check if machine and variant are selected
        machine = self.machine_var.get()
        variant = self.variant_var.get()
        
        if machine == "-- Pilih Mesin --" or variant == "-- Pilih Variant --":
            messagebox.showerror(
                "Selection Required",
                "Pilih Mesin dan Variant terlebih dahulu sebelum menyimpan data!"
            )
            return
        
        # Add machine and variant to data
        save_data = self.current_data.copy()
        save_data['machine'] = machine
        save_data['variant'] = variant
        
        # Add to saved data
        self.saved_data.append(save_data)
        
        # Add to treeview
        no = len(self.saved_data)
        self.tree.insert("", "end", values=(
            no,
            machine,
            variant,
            self.current_data['timestamp'],
            self.current_data['status'],
            f"{self.current_data['weight']:.2f}",
            self.current_data['unit']
        ))
        
        # Update counter
        self.counter_label.config(text=f"Total data: {len(self.saved_data)}")
        
        # Enable export button
        self.export_btn.config(state="normal")
        
        # Flash save button
        self.save_btn.config(bg="#4CAF50")
        self.root.after(200, lambda: self.save_btn.config(bg="#2196F3"))
        
        # Update status
        self.statusbar_label.config(text=f"✓ Data disimpan! Total: {len(self.saved_data)} | {machine} | {variant}")
        
        # Auto scroll to bottom
        self.tree.see(self.tree.get_children()[-1])
    
    def delete_selected(self):
        """Delete selected row"""
        selected = self.tree.selection()
        
        if not selected:
            messagebox.showwarning("No Selection", "Pilih data yang ingin dihapus!")
            return
        
        if messagebox.askyesno("Confirm", "Hapus data terpilih?"):
            for item in selected:
                # Get index
                idx = self.tree.index(item)
                # Delete from list
                del self.saved_data[idx]
                # Delete from tree
                self.tree.delete(item)
            
            # Reindex tree
            self.refresh_tree_numbers()
            
            # Update counter
            self.counter_label.config(text=f"Total data: {len(self.saved_data)}")
            
            if len(self.saved_data) == 0:
                self.export_btn.config(state="disabled")
    
    def clear_all_data(self):
        """Clear all saved data"""
        if not self.saved_data:
            return
        
        if messagebox.askyesno("Confirm", f"Hapus semua {len(self.saved_data)} data?"):
            self.saved_data.clear()
            
            # Clear treeview
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Update counter
            self.counter_label.config(text="Total data: 0")
            self.export_btn.config(state="disabled")
            self.statusbar_label.config(text="All data cleared")
    
    def refresh_tree_numbers(self):
        """Refresh row numbers in tree"""
        for i, item in enumerate(self.tree.get_children(), 1):
            self.tree.set(item, "No", i)
    
    def export_to_excel(self):
        """Export saved data to Excel"""
        if not self.saved_data:
            messagebox.showwarning("No Data", "Tidak ada data untuk di-export!")
            return
        
        # Ask for filename
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"timbangan_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Timbangan"
            
            # Header style
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            headers = ["No", "Mesin", "Variant", "Tanggal", "Waktu", "Status", "Berat", "Unit"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # Write data
            for i, data in enumerate(self.saved_data, 1):
                dt = datetime.strptime(data['timestamp'], '%Y-%m-%d %H:%M:%S')
                
                row = i + 1
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=data['machine'])
                ws.cell(row=row, column=3, value=data['variant'])
                ws.cell(row=row, column=4, value=dt.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=5, value=dt.strftime('%H:%M:%S'))
                ws.cell(row=row, column=6, value=data['status'])
                ws.cell(row=row, column=7, value=data['weight'])
                ws.cell(row=row, column=8, value=data['unit'])
                
                # Alignment and borders
                for col in range(1, 9):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="center" if col != 3 else "left", vertical="center")
                    cell.border = thin_border
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 8
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 6
            
            # Add summary
            summary_row = len(self.saved_data) + 3
            
            # Summary header
            summary_cell = ws.cell(row=summary_row, column=1, value="RINGKASAN DATA")
            summary_cell.font = Font(bold=True, size=12)
            summary_cell.fill = PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid")
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
            
            weights = [d['weight'] for d in self.saved_data]
            
            # Summary data
            summary_data = [
                ("Total Data:", len(weights)),
                ("Berat Minimum:", f"{min(weights):.2f} g"),
                ("Berat Maximum:", f"{max(weights):.2f} g"),
                ("Berat Rata-rata:", f"{sum(weights)/len(weights):.2f} g"),
            ]
            
            for idx, (label, value) in enumerate(summary_data, 1):
                row = summary_row + idx
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
            
            # Group by machine
            machine_summary_row = summary_row + len(summary_data) + 2
            ws.cell(row=machine_summary_row, column=1, value="RINGKASAN PER MESIN").font = Font(bold=True, size=12)
            
            from collections import defaultdict
            machine_count = defaultdict(int)
            for data in self.saved_data:
                machine_count[data['machine']] += 1
            
            for idx, (machine, count) in enumerate(sorted(machine_count.items()), 1):
                row = machine_summary_row + idx
                ws.cell(row=row, column=1, value=machine)
                ws.cell(row=row, column=2, value=count)
            
            # Save
            wb.save(filename)
            
            messagebox.showinfo(
                "Success",
                f"Data berhasil di-export!\n\n"
                f"File: {filename}\n"
                f"Total: {len(self.saved_data)} baris\n"
                f"Mesin: {len(machine_count)} mesin berbeda"
            )
            self.statusbar_label.config(text=f"✓ Exported to {filename}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Gagal export ke Excel:\n{str(e)}")


def main():
    root = tk.Tk()
    app = ScaleGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()