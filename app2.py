"""
Aplikasi GUI Modern Pembaca Timbangan AND GX-4000
Fitur: Multi-Page, Animated, Modern UI

Requirements:
pip install pyserial openpyxl

Author: Enhanced Version
Date: 2025-12-19
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import serial
import serial.tools.list_ports
import threading
import queue
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re
from collections import defaultdict


class ModernScaleGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Modern Scale Logger - AND GX-4000")
        self.root.geometry("1200x800")
        self.root.configure(bg='#1a1a2e')
        
        # Variables
        self.serial_conn = None
        self.is_reading = False
        self.data_queue = queue.Queue()
        self.current_data = None
        self.saved_data = []
        self.current_page = 0
        self.animation_id = None
        
        # Serial settings
        self.baudrate = 9600
        self.databits = 7
        self.parity = 'E'
        self.stopbits = 1
        
        # Machine and Variant lists
        self.machines = [
            "A (F2)", "AE (D12)", "AF (D13)", "AG (D14)", "AH (D15)", "AI (D16)",
            "AJ (D17)", "AK (D18)", "B (D11/E5)", "C (D9)", "D (D1)", "E (D2)",
            "F (D3)", "G (D4)", "H (D5)", "I (D6)", "J (D7)", "K (D8)",
            "L (LD10)", "O (C1)", "P (C2)", "Q (A2)", "R (C3)", "U (F3)",
            "V (F1)", "W (C7)", "X (C8)", "Y (B6)", "Z (B3)"
        ]
        
        self.variants = [
            "VARIANT YB P1000G", "VARIANT YB P700G PIRING",
            "VARIANT BB P700G NP HARGA", "VARIANT BB P270G",
            "VARIANT BB P725G", "VARIANT YB P700G",
            "VARIANT BB P77G HARGA BDKT", "VARIANT BB P77G HARGA",
            "VARIANT YB P77G B5G1", "VARIANT YB P77G B5G1 BDKT",
            "VARIANT YB P250G"
        ]
        
        # Colors
        self.colors = {
            'bg_dark': '#1a1a2e',
            'bg_light': '#16213e',
            'accent': '#0f3460',
            'primary': '#00d4ff',
            'success': '#00ff88',
            'warning': '#ffaa00',
            'danger': '#ff3366',
            'text_light': '#ffffff',
            'text_dark': '#a0a0a0'
        }
        
        # Style configuration
        self.setup_styles()
        
        # Create main container
        self.main_container = tk.Frame(self.root, bg=self.colors['bg_dark'])
        self.main_container.pack(fill='both', expand=True)
        
        # Create pages
        self.pages = {}
        self.create_pages()
        
        # Create navigation
        self.create_navigation()
        
        # Show first page
        self.show_page(0)
        
        # Start animations
        self.start_animations()
        
        # Load COM ports
        self.refresh_ports()
        
        # Start queue checker
        self.check_queue()
    
    def setup_styles(self):
        """Setup custom styles"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure styles
        style.configure('Modern.TFrame', background=self.colors['bg_light'])
        style.configure('Dark.TFrame', background=self.colors['bg_dark'])
        
        style.configure('Modern.TLabel',
                       background=self.colors['bg_light'],
                       foreground=self.colors['text_light'],
                       font=('Segoe UI', 10))
        
        style.configure('Title.TLabel',
                       background=self.colors['bg_dark'],
                       foreground=self.colors['primary'],
                       font=('Segoe UI', 24, 'bold'))
        
        style.configure('Subtitle.TLabel',
                       background=self.colors['bg_light'],
                       foreground=self.colors['text_dark'],
                       font=('Segoe UI', 12))
        
        style.configure('Modern.TCombobox',
                       fieldbackground=self.colors['accent'],
                       background=self.colors['accent'],
                       foreground=self.colors['text_light'])
        
        style.map('Modern.TCombobox',
                 fieldbackground=[('readonly', self.colors['accent'])],
                 selectbackground=[('readonly', self.colors['primary'])])
    
    def create_pages(self):
        """Create all pages"""
        # Page 1: Setup & Configuration
        self.pages[0] = self.create_setup_page()
        
        # Page 2: Live Monitoring
        self.pages[1] = self.create_monitoring_page()
        
        # Page 3: Data Management
        self.pages[2] = self.create_data_page()
    
    def create_navigation(self):
        """Create navigation bar"""
        nav_frame = tk.Frame(self.root, bg=self.colors['accent'], height=80)
        nav_frame.pack(side='top', fill='x')
        nav_frame.pack_propagate(False)
        
        # App title
        title_frame = tk.Frame(nav_frame, bg=self.colors['accent'])
        title_frame.pack(side='left', padx=20)
        
        tk.Label(title_frame, text="⚡", font=('Segoe UI', 28),
                bg=self.colors['accent'], fg=self.colors['primary']).pack(side='left')
        
        tk.Label(title_frame, text="SCALE LOGGER", font=('Segoe UI', 18, 'bold'),
                bg=self.colors['accent'], fg=self.colors['text_light']).pack(side='left', padx=10)
        
        # Navigation buttons
        btn_frame = tk.Frame(nav_frame, bg=self.colors['accent'])
        btn_frame.pack(side='right', padx=20)
        
        pages_info = [
            ("1️⃣ SETUP", 0),
            ("2️⃣ MONITOR", 1),
            ("3️⃣ DATA", 2)
        ]
        
        self.nav_buttons = []
        for text, page_num in pages_info:
            btn = tk.Button(btn_frame, text=text,
                          command=lambda p=page_num: self.show_page(p),
                          font=('Segoe UI', 11, 'bold'),
                          bg=self.colors['bg_light'],
                          fg=self.colors['text_light'],
                          activebackground=self.colors['primary'],
                          activeforeground=self.colors['bg_dark'],
                          bd=0, padx=20, pady=10,
                          cursor='hand2')
            btn.pack(side='left', padx=5)
            self.nav_buttons.append(btn)
            
            # Hover effects
            btn.bind('<Enter>', lambda e, b=btn: b.config(bg=self.colors['primary'], 
                                                          fg=self.colors['bg_dark']))
            btn.bind('<Leave>', lambda e, b=btn, p=page_num: b.config(
                bg=self.colors['primary'] if p == self.current_page else self.colors['bg_light'],
                fg=self.colors['bg_dark'] if p == self.current_page else self.colors['text_light']))
    
    def create_setup_page(self):
        """Create setup and configuration page"""
        page = tk.Frame(self.main_container, bg=self.colors['bg_dark'])
        
        # Center content
        content = tk.Frame(page, bg=self.colors['bg_dark'])
        content.place(relx=0.5, rely=0.5, anchor='center')
        
        # Title with animation
        title = tk.Label(content, text="⚙️ SETUP & KONFIGURASI",
                        font=('Segoe UI', 32, 'bold'),
                        bg=self.colors['bg_dark'],
                        fg=self.colors['primary'])
        title.pack(pady=30)
        
        # Card-style frames
        card_style = {
            'bg': self.colors['bg_light'],
            'relief': 'flat',
            'bd': 0
        }
        
        # Connection Card
        conn_card = tk.Frame(content, **card_style, padx=40, pady=30)
        conn_card.pack(pady=10, fill='x')
        
        tk.Label(conn_card, text="🔌 KONEKSI SERIAL",
                font=('Segoe UI', 16, 'bold'),
                bg=self.colors['bg_light'],
                fg=self.colors['success']).pack(anchor='w', pady=(0, 15))
        
        # COM Port
        port_frame = tk.Frame(conn_card, bg=self.colors['bg_light'])
        port_frame.pack(fill='x', pady=8)
        
        tk.Label(port_frame, text="COM Port:",
                font=('Segoe UI', 12),
                bg=self.colors['bg_light'],
                fg=self.colors['text_light'],
                width=15, anchor='w').pack(side='left')
        
        self.port_var = tk.StringVar()
        self.port_combo = ttk.Combobox(port_frame, textvariable=self.port_var,
                                      font=('Segoe UI', 11),
                                      width=20, state='readonly')
        self.port_combo.pack(side='left', padx=10)
        
        refresh_btn = tk.Button(port_frame, text="🔄 Refresh",
                               command=self.refresh_ports,
                               font=('Segoe UI', 10),
                               bg=self.colors['accent'],
                               fg=self.colors['text_light'],
                               bd=0, padx=15, pady=5,
                               cursor='hand2')
        refresh_btn.pack(side='left', padx=5)
        
        # Settings info
        settings_text = f"⚡ {self.baudrate} bps | {self.databits} bit | Parity: Even"
        tk.Label(conn_card, text=settings_text,
                font=('Segoe UI', 10, 'italic'),
                bg=self.colors['bg_light'],
                fg=self.colors['text_dark']).pack(anchor='w', pady=(10, 0))
        
        # Machine Card
        machine_card = tk.Frame(content, **card_style, padx=40, pady=30)
        machine_card.pack(pady=10, fill='x')
        
        tk.Label(machine_card, text="🏭 PILIH MESIN",
                font=('Segoe UI', 16, 'bold'),
                bg=self.colors['bg_light'],
                fg=self.colors['warning']).pack(anchor='w', pady=(0, 15))
        
        machine_frame = tk.Frame(machine_card, bg=self.colors['bg_light'])
        machine_frame.pack(fill='x', pady=8)
        
        tk.Label(machine_frame, text="Mesin:",
                font=('Segoe UI', 12),
                bg=self.colors['bg_light'],
                fg=self.colors['text_light'],
                width=15, anchor='w').pack(side='left')
        
        self.machine_var = tk.StringVar()
        self.machine_combo = ttk.Combobox(machine_frame, textvariable=self.machine_var,
                                         values=self.machines,
                                         font=('Segoe UI', 11),
                                         width=30, state='readonly')
        self.machine_combo.pack(side='left', padx=10)
        self.machine_combo.set("-- Pilih Mesin --")
        self.machine_combo.bind("<<ComboboxSelected>>", self.update_setup_status)
        
        # Variant Card
        variant_card = tk.Frame(content, **card_style, padx=40, pady=30)
        variant_card.pack(pady=10, fill='x')
        
        tk.Label(variant_card, text="📦 PILIH VARIANT",
                font=('Segoe UI', 16, 'bold'),
                bg=self.colors['bg_light'],
                fg=self.colors['danger']).pack(anchor='w', pady=(0, 15))
        
        variant_frame = tk.Frame(variant_card, bg=self.colors['bg_light'])
        variant_frame.pack(fill='x', pady=8)
        
        tk.Label(variant_frame, text="Variant:",
                font=('Segoe UI', 12),
                bg=self.colors['bg_light'],
                fg=self.colors['text_light'],
                width=15, anchor='w').pack(side='left')
        
        self.variant_var = tk.StringVar()
        self.variant_combo = ttk.Combobox(variant_frame, textvariable=self.variant_var,
                                         values=self.variants,
                                         font=('Segoe UI', 11),
                                         width=40, state='readonly')
        self.variant_combo.pack(side='left', padx=10)
        self.variant_combo.set("-- Pilih Variant --")
        self.variant_combo.bind("<<ComboboxSelected>>", self.update_setup_status)
        
        # Status indicator
        self.setup_status = tk.Label(content, text="⚠️ Lengkapi setup untuk melanjutkan",
                                    font=('Segoe UI', 13, 'bold'),
                                    bg=self.colors['bg_dark'],
                                    fg=self.colors['warning'])
        self.setup_status.pack(pady=20)
        
        # Next button
        self.next_btn = tk.Button(content, text="▶️ LANJUT KE MONITORING",
                                 command=lambda: self.show_page(1),
                                 font=('Segoe UI', 14, 'bold'),
                                 bg=self.colors['primary'],
                                 fg=self.colors['bg_dark'],
                                 bd=0, padx=40, pady=15,
                                 cursor='hand2',
                                 state='disabled')
        self.next_btn.pack(pady=20)
        
        return page
    
    def create_monitoring_page(self):
        """Create live monitoring page"""
        page = tk.Frame(self.main_container, bg=self.colors['bg_dark'])
        
        # Top section: Title and controls
        top_section = tk.Frame(page, bg=self.colors['bg_dark'])
        top_section.pack(fill='x', padx=30, pady=20)
        
        tk.Label(top_section, text="📊 LIVE MONITORING",
                font=('Segoe UI', 28, 'bold'),
                bg=self.colors['bg_dark'],
                fg=self.colors['primary']).pack()
        
        # Selection display
        self.selection_display = tk.Label(top_section, text="",
                                         font=('Segoe UI', 11),
                                         bg=self.colors['bg_dark'],
                                         fg=self.colors['text_dark'])
        self.selection_display.pack(pady=10)
        
        # Center section: Weight display
        center_frame = tk.Frame(page, bg=self.colors['bg_light'],
                               relief='flat', bd=0)
        center_frame.pack(fill='both', expand=True, padx=30, pady=10)
        
        # Animated border effect
        self.weight_container = tk.Frame(center_frame, bg=self.colors['primary'],
                                        relief='flat', bd=3)
        self.weight_container.place(relx=0.5, rely=0.4, anchor='center')
        
        weight_display = tk.Frame(self.weight_container, bg=self.colors['bg_light'],
                                 padx=100, pady=60)
        weight_display.pack(padx=3, pady=3)
        
        self.weight_label = tk.Label(weight_display, text="0.00",
                                     font=('Segoe UI', 72, 'bold'),
                                     bg=self.colors['bg_light'],
                                     fg=self.colors['success'])
        self.weight_label.pack()
        
        unit_frame = tk.Frame(weight_display, bg=self.colors['bg_light'])
        unit_frame.pack()
        
        self.unit_label = tk.Label(unit_frame, text="gram",
                                   font=('Segoe UI', 20),
                                   bg=self.colors['bg_light'],
                                   fg=self.colors['text_dark'])
        self.unit_label.pack(side='left', padx=10)
        
        self.status_indicator = tk.Label(unit_frame, text="●",
                                        font=('Segoe UI', 32),
                                        bg=self.colors['bg_light'],
                                        fg=self.colors['text_dark'])
        self.status_indicator.pack(side='left', padx=10)
        
        self.status_text = tk.Label(unit_frame, text="Ready",
                                    font=('Segoe UI', 16),
                                    bg=self.colors['bg_light'],
                                    fg=self.colors['text_dark'])
        self.status_text.pack(side='left')
        
        # Control buttons
        btn_container = tk.Frame(page, bg=self.colors['bg_dark'])
        btn_container.pack(pady=20)
        
        buttons = [
            ("▶️ START", self.start_reading, self.colors['success'], 'start'),
            ("⏸️ STOP", self.stop_reading, self.colors['danger'], 'stop'),
            ("💾 SIMPAN", self.save_current_data, self.colors['primary'], 'save')
        ]
        
        self.control_buttons = {}
        for text, command, color, key in buttons:
            btn = tk.Button(btn_container, text=text,
                          command=command,
                          font=('Segoe UI', 14, 'bold'),
                          bg=color, fg=self.colors['bg_dark'],
                          bd=0, padx=40, pady=15,
                          cursor='hand2')
            btn.pack(side='left', padx=10)
            self.control_buttons[key] = btn
            
            if key in ['stop', 'save']:
                btn.config(state='disabled')
        
        # Connection status
        self.conn_status = tk.Label(page, text="● Disconnected",
                                   font=('Segoe UI', 12, 'bold'),
                                   bg=self.colors['bg_dark'],
                                   fg=self.colors['danger'])
        self.conn_status.pack(pady=10)
        
        return page
    
    def create_data_page(self):
        """Create data management page"""
        page = tk.Frame(self.main_container, bg=self.colors['bg_dark'])
        
        # Title
        title_frame = tk.Frame(page, bg=self.colors['bg_dark'])
        title_frame.pack(fill='x', padx=30, pady=20)
        
        tk.Label(title_frame, text="📋 DATA MANAGEMENT",
                font=('Segoe UI', 28, 'bold'),
                bg=self.colors['bg_dark'],
                fg=self.colors['primary']).pack()
        
        # Stats cards
        stats_frame = tk.Frame(page, bg=self.colors['bg_dark'])
        stats_frame.pack(fill='x', padx=30, pady=10)
        
        self.stat_cards = {}
        stats_info = [
            ("📊 Total Data", "total", self.colors['primary']),
            ("🏭 Machines", "machines", self.colors['success']),
            ("📦 Variants", "variants", self.colors['warning'])
        ]
        
        for label, key, color in stats_info:
            card = tk.Frame(stats_frame, bg=self.colors['bg_light'],
                          relief='flat', bd=0, padx=30, pady=20)
            card.pack(side='left', expand=True, fill='both', padx=5)
            
            tk.Label(card, text=label,
                    font=('Segoe UI', 12),
                    bg=self.colors['bg_light'],
                    fg=self.colors['text_dark']).pack()
            
            value_label = tk.Label(card, text="0",
                                  font=('Segoe UI', 32, 'bold'),
                                  bg=self.colors['bg_light'],
                                  fg=color)
            value_label.pack()
            self.stat_cards[key] = value_label
        
        # Table frame
        table_frame = tk.Frame(page, bg=self.colors['bg_light'],
                              relief='flat', bd=0)
        table_frame.pack(fill='both', expand=True, padx=30, pady=10)
        
        # Scrollbars
        scroll_y = ttk.Scrollbar(table_frame)
        scroll_y.pack(side='right', fill='y')
        
        scroll_x = ttk.Scrollbar(table_frame, orient='horizontal')
        scroll_x.pack(side='bottom', fill='x')
        
        # Treeview
        style = ttk.Style()
        style.configure('Modern.Treeview',
                       background=self.colors['bg_light'],
                       foreground=self.colors['text_light'],
                       fieldbackground=self.colors['bg_light'],
                       font=('Segoe UI', 10))
        style.map('Modern.Treeview',
                 background=[('selected', self.colors['primary'])])
        
        self.tree = ttk.Treeview(table_frame,
                                columns=("No", "Mesin", "Variant", "Timestamp", 
                                        "Status", "Weight", "Unit"),
                                show='headings',
                                yscrollcommand=scroll_y.set,
                                xscrollcommand=scroll_x.set,
                                style='Modern.Treeview')
        
        scroll_y.config(command=self.tree.yview)
        scroll_x.config(command=self.tree.xview)
        
        # Configure columns
        columns_config = {
            "No": (50, 'center'),
            "Mesin": (100, 'center'),
            "Variant": (250, 'w'),
            "Timestamp": (150, 'center'),
            "Status": (70, 'center'),
            "Weight": (100, 'center'),
            "Unit": (60, 'center')
        }
        
        for col, (width, anchor) in columns_config.items():
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor=anchor)
        
        self.tree.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Action buttons
        action_frame = tk.Frame(page, bg=self.colors['bg_dark'])
        action_frame.pack(fill='x', padx=30, pady=20)
        
        actions = [
            ("🗑️ Hapus Selected", self.delete_selected, self.colors['danger']),
            ("🗑️ Hapus Semua", self.clear_all_data, self.colors['danger']),
            ("📊 Export Excel", self.export_to_excel, self.colors['success'])
        ]
        
        self.export_btn = None
        for text, command, color in actions:
            btn = tk.Button(action_frame, text=text,
                          command=command,
                          font=('Segoe UI', 12, 'bold'),
                          bg=color, fg=self.colors['text_light'],
                          bd=0, padx=30, pady=12,
                          cursor='hand2')
            btn.pack(side='left', padx=5)
            
            if "Export" in text:
                self.export_btn = btn
                btn.config(state='disabled')
        
        return page
    
    def show_page(self, page_num):
        """Show specific page with animation"""
        if page_num == self.current_page:
            return
        
        # Validation for page 1
        if page_num == 1:
            if (self.machine_var.get() == "-- Pilih Mesin --" or 
                self.variant_var.get() == "-- Pilih Variant --"):
                messagebox.showwarning("Setup Incomplete",
                                      "Lengkapi setup di Page 1 terlebih dahulu!")
                return
        
        # Hide current page
        if self.current_page in self.pages:
            self.pages[self.current_page].pack_forget()
        
        # Show new page
        self.current_page = page_num
        self.pages[page_num].pack(fill='both', expand=True)
        
        # Update navigation buttons
        for i, btn in enumerate(self.nav_buttons):
            if i == page_num:
                btn.config(bg=self.colors['primary'], fg=self.colors['bg_dark'])
            else:
                btn.config(bg=self.colors['bg_light'], fg=self.colors['text_light'])
        
        # Update monitoring page display
        if page_num == 1:
            machine = self.machine_var.get()
            variant = self.variant_var.get()
            self.selection_display.config(
                text=f"🏭 {machine}  |  📦 {variant}"
            )
    
    def start_animations(self):
        """Start UI animations"""
        self.animate_pulse()
    
    def animate_pulse(self):
        """Pulse animation for weight display border"""
        if hasattr(self, 'weight_container') and self.is_reading:
            colors = [self.colors['primary'], self.colors['success'], 
                     self.colors['warning']]
            color = colors[datetime.now().second % len(colors)]
            self.weight_container.config(bg=color)
        
        self.animation_id = self.root.after(1000, self.animate_pulse)
    
    def update_setup_status(self, event=None):
        """Update setup status"""
        machine = self.machine_var.get()
        variant = self.variant_var.get()
        
        if machine != "-- Pilih Mesin --" and variant != "-- Pilih Variant --":
            self.setup_status.config(
                text="✅ Setup lengkap! Siap monitoring",
                fg=self.colors['success']
            )
            self.next_btn.config(state='normal', bg=self.colors['success'])
        else:
            self.setup_status.config(
                text="⚠️ Lengkapi setup untuk melanjutkan",
                fg=self.colors['warning']
            )
            self.next_btn.config(state='disabled', bg=self.colors['text_dark'])
    
    def refresh_ports(self):
        """Refresh COM ports"""
        ports = serial.tools.list_ports.comports()
        port_list = [p.device for p in ports]
        
        self.port_combo['values'] = port_list
        if port_list:
            self.port_combo.current(0)
    
    def start_reading(self):
        """Start reading from scale"""
        port = self.port_var.get()
        
        if not port:
            messagebox.showerror("Error", "Pilih COM port terlebih dahulu!")
            return
        
        try:
            parity_map = {'N': serial.PARITY_NONE, 'E': serial.PARITY_EVEN, 
                         'O': serial.PARITY_ODD}
            
            self.serial_conn = serial.Serial(
                port=port,
                baudrate=self.baudrate,
                bytesize=self.databits,
                parity=parity_map[self.parity],
                stopbits=self.stopbits,
                timeout=1
            )
            
            self.serial_conn.reset_input_buffer()
            self.is_reading = True
            
            # Update UI
            self.control_buttons['start'].config(state='disabled')
            self.control_buttons['stop'].config(state='normal')
            self.control_buttons['save'].config(state='normal')
            
            self.conn_status.config(text="● Connected", fg=self.colors['success'])
            
            # Start thread
            thread = threading.Thread(target=self.read_thread, daemon=True)
            thread.start()
            
        except Exception as e:
            messagebox.showerror("Connection Error", f"Failed:\n{str(e)}")
    
    def stop_reading(self):
        """Stop reading"""
        self.is_reading = False
        
        if self.serial_conn and self.serial_conn.is_open:
            self.serial_conn.close()
        
        self.control_buttons['start'].config(state='normal')
        self.control_buttons['stop'].config(state='disabled')
        self.control_buttons['save'].config(state='disabled')
        
        self.conn_status.config(text="● Disconnected", fg=self.colors['danger'])
        self.weight_label.config(text="0.00", fg=self.colors['text_dark'])
        self.status_text.config(text="Stopped")
    
    def read_thread(self):
        """Background reading thread"""
        while self.is_reading:
            try:
                if self.serial_conn and self.serial_conn.in_waiting > 0:
                    raw_data = self.serial_conn.readline().decode('ascii', errors='ignore').strip()
                    
                    if raw_data:
                        parsed = self.parse_data(raw_data)
                        if parsed:
                            self.data_queue.put(parsed)
            except Exception as e:
                print(f"Read error: {e}")
                break
    
    def parse_data(self, raw_data):
        """Parse scale data"""
        try:
            pattern = r'^([A-Z]{2}),([+-]?\d+\.?\d*)\s*([a-zA-Z]+)$'
            match = re.match(pattern, raw_data)
            
            if match:
                status = match.group(1)
                weight = float(match.group(2))
                unit = match.group(3)
                
                return {
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'status': status,
                    'weight': weight,
                    'unit': unit,
                    'raw': raw_data
                }
        except Exception as e:
            print(f"Parse error: {e}")
        
        return None
    
    def check_queue(self):
        """Check data queue and update UI"""
        try:
            while True:
                data = self.data_queue.get_nowait()
                self.update_display(data)
        except queue.Empty:
            pass
        
        self.root.after(50, self.check_queue)
    
    def update_display(self, data):
        """Update current reading display"""
        self.current_data = data
        
        # Update weight
        self.weight_label.config(text=f"{data['weight']:.2f}")
        
        # Update unit
        self.unit_label.config(text=data['unit'])
        
        # Update status
        if data['status'] == 'ST':
            self.weight_label.config(fg=self.colors['success'])
            self.status_indicator.config(fg=self.colors['success'])
            self.status_text.config(text="Stable", fg=self.colors['success'])
        else:
            self.weight_label.config(fg=self.colors['warning'])
            self.status_indicator.config(fg=self.colors['warning'])
            self.status_text.config(text="Unstable", fg=self.colors['warning'])
    
    def save_current_data(self):
        """Save current displayed data to list"""
        if not self.current_data:
            messagebox.showwarning("No Data", "Tidak ada data untuk disimpan!")
            return
        
        machine = self.machine_var.get()
        variant = self.variant_var.get()
        
        # Add machine and variant to data
        save_data = self.current_data.copy()
        save_data['machine'] = machine
        save_data['variant'] = variant
        
        # Add to saved data
        self.saved_data.append(save_data)
        
        # Add to treeview
        no = len(self.saved_data)
        self.tree.insert("", "end", values=(
            no,
            machine,
            variant,
            self.current_data['timestamp'],
            self.current_data['status'],
            f"{self.current_data['weight']:.2f}",
            self.current_data['unit']
        ))
        
        # Update stats
        self.update_stats()
        
        # Enable export button
        self.export_btn.config(state='normal')
        
        # Flash save button
        original_bg = self.control_buttons['save']['bg']
        self.control_buttons['save'].config(bg=self.colors['success'])
        self.root.after(200, lambda: self.control_buttons['save'].config(bg=original_bg))
        
        # Auto scroll to bottom
        if self.tree.get_children():
            self.tree.see(self.tree.get_children()[-1])
    
    def update_stats(self):
        """Update statistics cards"""
        total = len(self.saved_data)
        machines = len(set(d['machine'] for d in self.saved_data))
        variants = len(set(d['variant'] for d in self.saved_data))
        
        self.stat_cards['total'].config(text=str(total))
        self.stat_cards['machines'].config(text=str(machines))
        self.stat_cards['variants'].config(text=str(variants))
    
    def delete_selected(self):
        """Delete selected row"""
        selected = self.tree.selection()
        
        if not selected:
            messagebox.showwarning("No Selection", "Pilih data yang ingin dihapus!")
            return
        
        if messagebox.askyesno("Confirm", "Hapus data terpilih?"):
            for item in selected:
                idx = self.tree.index(item)
                del self.saved_data[idx]
                self.tree.delete(item)
            
            self.refresh_tree_numbers()
            self.update_stats()
            
            if len(self.saved_data) == 0:
                self.export_btn.config(state='disabled')
    
    def clear_all_data(self):
        """Clear all saved data"""
        if not self.saved_data:
            return
        
        if messagebox.askyesno("Confirm", f"Hapus semua {len(self.saved_data)} data?"):
            self.saved_data.clear()
            
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            self.update_stats()
            self.export_btn.config(state='disabled')
    
    def refresh_tree_numbers(self):
        """Refresh row numbers in tree"""
        for i, item in enumerate(self.tree.get_children(), 1):
            self.tree.set(item, "No", i)
    
    def export_to_excel(self):
        """Export saved data to Excel"""
        if not self.saved_data:
            messagebox.showwarning("No Data", "Tidak ada data untuk di-export!")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"timbangan_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Timbangan"
            
            # Header style
            header_fill = PatternFill(start_color="00d4ff", end_color="00d4ff", fill_type="solid")
            header_font = Font(bold=True, color="000000", size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            headers = ["No", "Mesin", "Variant", "Tanggal", "Waktu", "Status", "Berat", "Unit"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # Write data
            for i, data in enumerate(self.saved_data, 1):
                dt = datetime.strptime(data['timestamp'], '%Y-%m-%d %H:%M:%S')
                
                row = i + 1
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=data['machine'])
                ws.cell(row=row, column=3, value=data['variant'])
                ws.cell(row=row, column=4, value=dt.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=5, value=dt.strftime('%H:%M:%S'))
                ws.cell(row=row, column=6, value=data['status'])
                ws.cell(row=row, column=7, value=data['weight'])
                ws.cell(row=row, column=8, value=data['unit'])
                
                for col in range(1, 9):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="center" if col != 3 else "left", vertical="center")
                    cell.border = thin_border
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 8
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 6
            
            # Add summary
            summary_row = len(self.saved_data) + 3
            
            summary_cell = ws.cell(row=summary_row, column=1, value="RINGKASAN DATA")
            summary_cell.font = Font(bold=True, size=12)
            summary_cell.fill = PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid")
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
            
            weights = [d['weight'] for d in self.saved_data]
            
            summary_data = [
                ("Total Data:", len(weights)),
                ("Berat Minimum:", f"{min(weights):.2f} g"),
                ("Berat Maximum:", f"{max(weights):.2f} g"),
                ("Berat Rata-rata:", f"{sum(weights)/len(weights):.2f} g"),
            ]
            
            for idx, (label, value) in enumerate(summary_data, 1):
                row = summary_row + idx
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
            
            # Group by machine
            machine_summary_row = summary_row + len(summary_data) + 2
            ws.cell(row=machine_summary_row, column=1, value="RINGKASAN PER MESIN").font = Font(bold=True, size=12)
            
            machine_count = defaultdict(int)
            for data in self.saved_data:
                machine_count[data['machine']] += 1
            
            for idx, (machine, count) in enumerate(sorted(machine_count.items()), 1):
                row = machine_summary_row + idx
                ws.cell(row=row, column=1, value=machine)
                ws.cell(row=row, column=2, value=count)
            
            wb.save(filename)
            
            messagebox.showinfo(
                "Success",
                f"Data berhasil di-export!\n\n"
                f"File: {filename}\n"
                f"Total: {len(self.saved_data)} baris\n"
                f"Mesin: {len(machine_count)} mesin berbeda"
            )
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Gagal export ke Excel:\n{str(e)}")


def main():
    root = tk.Tk()
    app = ModernScaleGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()