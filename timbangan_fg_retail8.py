"""
Timbangan AND GX-4000 — Quality Control Data Logger
UI: Merah Terang & Putih, 3-column layout
Flow: NIK → Variant → Mesin → Filler → Simpan

Revisi v11:
  - FIX PORT: hanya konek ke USB-SERIAL CH340
    → _is_ch340_port() filter keras: port lain (Prolific, Bluetooth,
      virtual, COM sistem) langsung dilewati tanpa dicoba sama sekali
    → Jika CH340 tidak terpasang → status "USB-SERIAL CH340 tidak ditemukan"
      dan retry otomatis tiap 5 detik
  - Revisi v10 (tetap):
    → data WAJIB terkirim ke API sebelum dianggap tersimpan
    → loading overlay blokir UI sampai POST selesai / gagal
    → retry otomatis 3x dengan jeda 2 detik
  - Revisi v9 (tetap):
    → frozen data tidak hilang saat banner disembunyikan

Requirements:
    pip install pyserial openpyxl requests
"""

import tkinter as tk
from tkinter import messagebox, filedialog
import serial
import serial.tools.list_ports
import threading
import queue
import time
import re
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ── PALETTE ────────────────────────────────────────────────────
RED           = "#E53E3E"
RED_DARK      = "#C53030"
RED_LIGHT     = "#FC8181"
RED_PALE      = "#FFF5F5"
RED_SOFT      = "#FED7D7"
WHITE         = "#FFFFFF"
GRAY_50       = "#FAFAFA"
GRAY_100      = "#F4F4F5"
GRAY_200      = "#E4E4E7"
GRAY_400      = "#A1A1AA"
GRAY_600      = "#52525B"
GRAY_800      = "#27272A"
GREEN         = "#16A34A"
GREEN_LIGHT   = "#DCFCE7"
RED_ALERT     = "#DC2626"
BG_ROOT       = "#F8ECEC"
AMBER         = "#D97706"
CONFIRM_BG    = "#FEF9C3"
CONFIRM_BORDER= "#FDE68A"
CONFIRM_FG    = "#92400E"
CONFIRM_BTN   = "#FDE68A"
CONFIRM_OK    = "#15803D"
ABNORMAL_BG   = "#FEF2F2"
ABNORMAL_BORDER="#FECACA"
ABNORMAL_FG   = "#991B1B"

# ── FONTS ──────────────────────────────────────────────────────
F_TITLE  = ("Segoe UI", 13, "bold")
F_SUB    = ("Segoe UI", 9)
F_LABEL  = ("Segoe UI", 8, "bold")
F_BODY_B = ("Segoe UI", 10, "bold")
F_SMALL  = ("Segoe UI", 8)
F_CHIP   = ("Segoe UI", 9, "bold")
F_CHIP_S = ("Segoe UI", 8)
F_MESIN  = ("Segoe UI", 12, "bold")
F_MESIN_S= ("Segoe UI", 8)
F_WEIGHT = ("Courier New", 52, "bold")
F_UNIT   = ("Segoe UI", 13)
F_STATUS = ("Segoe UI", 10, "bold")
F_NIK    = ("Courier New", 13, "bold")
F_BTN    = ("Segoe UI", 11, "bold")
F_BTN_SM = ("Segoe UI", 9, "bold")
F_TABLE  = ("Segoe UI", 9)
F_TABLE_H= ("Segoe UI", 8, "bold")
F_MONO   = ("Courier New", 9)
F_BADGE  = ("Segoe UI", 8, "bold")

# ── SERIAL CONFIG ───────────────────────────────────────────────
BAUDRATE = 9600
DATABITS = 7
PARITY   = "E"
STOPBITS = 1

# ── API ─────────────────────────────────────────────────────────
API_URL = "http://10.11.10.130:8081/api/mesin2"

# ── NIK WHITELIST ───────────────────────────────────────────────
VALID_NIKS = {
    "24000522", "24000551", "220000247", "17000047", "23000475",
    "220000265", "120000064", "17000123", "24000539", "20000336",
    "23000480", "220000264", "17000113", "24000529", "18000252",
    "220000230", "220000262",
}

# ── DATA VARIAN + STANDAR ───────────────────────────────────────
VARIANT_STANDARDS = {
    "Sachet YB 12,5gr PCS":    {"rev_lo":   8.55, "tu2":  10.80, "tu1":  11.93, "min":  12.05, "std":  13.05, "max":  14.05, "rev_hi":  15.05, "code": "S12.5G-P"},
    "Sachet YB 12,5gr RENCENG":{"rev_lo": 102.60, "tu2": 129.60, "tu1": 143.10, "min": 154.60, "std": 156.60, "max": 168.60, "rev_hi": 180.60, "code": "S12.5G-R"},
    "Sachet YB 20gr PCS":      {"rev_lo":  13.44, "tu2":  17.04, "tu1":  18.84, "min":  19.14, "std":  20.64, "max":  21.64, "rev_hi":  22.64, "code": "S20G-P"},
    "Sachet YB 20gr RENCENG":  {"rev_lo": 161.28, "tu2": 204.48, "tu1": 226.08, "min": 244.68, "std": 247.68, "max": 259.68, "rev_hi": 271.68, "code": "S20G-R"},
    "Sachet BB 40gr PCS":      {"rev_lo":  26.70, "tu2":  33.90, "tu1":  37.50, "min":  39.10, "std":  41.10, "max":  42.10, "rev_hi":  43.10, "code": "S40G-P"},
    "Sachet BB 40gr RENCENG":  {"rev_lo": 320.40, "tu2": 406.80, "tu1": 450.00, "min": 489.20, "std": 493.20, "max": 505.20, "rev_hi": 517.20, "code": "S40G-R"},
    "Pouch YB 77gr":           {"rev_lo":  61.20, "tu2":  70.20, "tu1":  74.70, "min":  78.70, "std":  79.20, "max":  82.70, "rev_hi":  86.20, "code": "P77G-YB"},
    "Pouch BB 77gr":           {"rev_lo":  61.20, "tu2":  70.20, "tu1":  74.70, "min":  78.70, "std":  79.20, "max":  82.70, "rev_hi":  86.20, "code": "P77G-BB"},
    "Pouch YB 250gr":          {"rev_lo": 219.00, "tu2": 237.00, "tu1": 246.00, "min": 254.00, "std": 255.00, "max": 257.00, "rev_hi": 259.00, "code": "P250G"},
    "Pouch BB 270gr":          {"rev_lo": 239.00, "tu2": 257.00, "tu1": 266.00, "min": 274.00, "std": 275.00, "max": 277.00, "rev_hi": 279.00, "code": "P270G"},
    "Pouch YB 550gr":          {"rev_lo": 500.60, "tu2": 530.80, "tu1": 545.80, "min": 558.00, "std": 561.00, "max": 566.00, "rev_hi": 571.00, "code": "P550G"},
    "Pouch YB 700gr":          {"rev_lo": 651.00, "tu2": 681.00, "tu1": 696.00, "min": 708.00, "std": 711.00, "max": 716.00, "rev_hi": 721.00, "code": "P700G"},
    "Pouch BB 725gr":          {"rev_lo": 675.00, "tu2": 705.00, "tu1": 720.00, "min": 732.00, "std": 735.00, "max": 740.00, "rev_hi": 745.00, "code": "P725G"},
    "Pouch YB 1000gr":         {"rev_lo": 952.50, "tu2": 982.50, "tu1": 997.50, "min":1009.50, "std":1012.50, "max":1017.50, "rev_hi":1022.50, "code": "P1000G"},
}

# ── RELASI VARIAN → MESIN ────────────────────────────────────────
VARIANT_MESIN = {
    "Sachet YB 12,5gr PCS":    {"Y", "Z"},
    "Sachet YB 12,5gr RENCENG":{"Y", "Z"},
    "Sachet YB 20gr PCS":      {"O", "P", "W", "X"},
    "Sachet YB 20gr RENCENG":  {"O", "P", "W", "X"},
    "Sachet BB 40gr PCS":      {"Q", "R"},
    "Sachet BB 40gr RENCENG":  {"Q", "R"},
    "Pouch YB 77gr":           {"F","G","H","I","D","E","J","K","C","L","AE","AG"},
    "Pouch BB 77gr":           {"C","L","AE","AG","B","AF","AI","AJ"},
    "Pouch YB 250gr":          {"AH"},
    "Pouch BB 270gr":          {"AH"},
    "Pouch YB 550gr":          {"A", "U", "V"},
    "Pouch YB 700gr":          {"A", "U", "V"},
    "Pouch BB 725gr":          {"A", "U", "V"},
    "Pouch YB 1000gr":         {"A", "U", "V"},
}

# ── RELASI VARIAN → JUMLAH FILLER ───────────────────────────────
VARIANT_FILLER_MAX = {
    "Sachet YB 12,5gr PCS":    8,
    "Sachet YB 12,5gr RENCENG":8,
    "Sachet YB 20gr PCS":      6,
    "Sachet YB 20gr RENCENG":  6,
    "Sachet BB 40gr PCS":      6,
    "Sachet BB 40gr RENCENG":  6,
    "Pouch YB 77gr":           2,
    "Pouch BB 77gr":           2,
    "Pouch YB 250gr":          2,
    "Pouch BB 270gr":          2,
    "Pouch YB 550gr":          2,
    "Pouch YB 700gr":          2,
    "Pouch BB 725gr":          2,
    "Pouch YB 1000gr":         2,
}

# ── SEMUA MESIN ─────────────────────────────────────────────────
MACHINES = [
    ("A","F2"),("AE","D12"),("AF","D13"),("AG","D14"),("AH","D15"),
    ("AI","D16"),("AJ","D17"),("AK","D18"),("B","D11/E5"),("C","D9"),
    ("D","D1"),("E","D2"),("F","D3"),("G","D4"),("H","D5"),("I","D6"),
    ("J","D7"),("K","D8"),("L","LD10"),("O","C1"),("P","C2"),("Q","A2"),
    ("R","C3"),("U","F3"),("V","F1"),("W","C7"),("X","C8"),("Y","B6"),("Z","B3"),
]

ALL_FILLER_VALUES = [str(i) for i in range(1, 9)]   # 1–8


def _is_abnormal(weight, std_data):
    if std_data is None:
        return False
    lo = std_data.get("rev_lo", std_data["std"] / 2.0)
    hi = std_data.get("rev_hi", std_data["std"] * 1.5)
    return weight < lo or weight > hi


# ── v11: FILTER KERAS — hanya USB-SERIAL CH340 ──────────────────
def _is_ch340_port(port_info):
    """
    Kembalikan True HANYA jika port ini adalah USB-SERIAL CH340.
    Semua port lain (Prolific, Bluetooth, virtual, COM sistem)
    langsung dilewati tanpa dicoba sama sekali.
    """
    desc  = (port_info.description  or "").lower()
    mfr   = (port_info.manufacturer or "").lower()
    hwid  = (port_info.hwid         or "").lower()
    combined = desc + " " + mfr + " " + hwid
    return "ch340" in combined


# ════════════════════════════════════════════════════════════════
class FillerChip(tk.Frame):
    SIZE   = 26
    RADIUS = 4

    def __init__(self, parent, value, on_select, **kwargs):
        super().__init__(parent, bg=WHITE, **kwargs)
        self.value     = value
        self._on_sel   = on_select
        self._selected = False
        self._disabled = False

        self._canvas = tk.Canvas(
            self,
            width=self.SIZE, height=self.SIZE,
            bg=WHITE, highlightthickness=0, cursor="hand2")
        self._canvas.pack()
        self._draw("normal")

        self._canvas.bind("<Button-1>", self._click)
        self._canvas.bind("<Enter>",    self._hover_on)
        self._canvas.bind("<Leave>",    self._hover_off)

    def _draw(self, state):
        c = self._canvas
        c.delete("all")
        S = self.SIZE

        if self._disabled:
            bg_col, fg_col, bd_col = GRAY_100, GRAY_400, GRAY_200
        elif state == "selected":
            bg_col, fg_col, bd_col = RED, WHITE, RED_DARK
        elif state == "hover":
            bg_col, fg_col, bd_col = RED_PALE, RED, RED_LIGHT
        else:
            bg_col, fg_col, bd_col = WHITE, GRAY_600, GRAY_200

        r = self.RADIUS
        c.create_rectangle(r, 0, S-r, S, fill=bg_col, outline=bg_col)
        c.create_rectangle(0, r, S, S-r, fill=bg_col, outline=bg_col)
        c.create_oval(0,   0,   2*r, 2*r, fill=bg_col, outline=bg_col)
        c.create_oval(S-2*r, 0, S, 2*r,   fill=bg_col, outline=bg_col)
        c.create_oval(0, S-2*r, 2*r, S,   fill=bg_col, outline=bg_col)
        c.create_oval(S-2*r, S-2*r, S, S, fill=bg_col, outline=bg_col)

        c.create_rectangle(r, 0, S-r, S,   fill="", outline=bd_col)
        c.create_rectangle(0, r, S, S-r,   fill="", outline=bd_col)
        c.create_oval(0,   0,   2*r, 2*r,   fill="", outline=bd_col)
        c.create_oval(S-2*r, 0, S,   2*r,   fill="", outline=bd_col)
        c.create_oval(0,   S-2*r, 2*r, S,   fill="", outline=bd_col)
        c.create_oval(S-2*r, S-2*r, S, S,   fill="", outline=bd_col)

        c.create_text(S//2, S//2, text=self.value,
                      font=("Segoe UI", 8, "bold"), fill=fg_col)

    def _click(self, e=None):
        if not self._disabled:
            self._on_sel(self.value)

    def _hover_on(self, e=None):
        if not self._selected and not self._disabled:
            self._draw("hover")

    def _hover_off(self, e=None):
        if not self._selected and not self._disabled:
            self._draw("normal")

    def select(self):
        self._selected = True
        self._draw("selected")

    def deselect(self):
        self._selected = False
        if self._disabled:
            self._draw("disabled")
        else:
            self._draw("normal")

    def enable(self):
        self._disabled = False
        self._canvas.config(cursor="hand2")
        if self._selected:
            self._draw("selected")
        else:
            self._draw("normal")

    def disable(self):
        self._disabled = True
        self._selected = False
        self._canvas.config(cursor="arrow")
        self._draw("disabled")


# ════════════════════════════════════════════════════════════════
class ChipButton(tk.Frame):
    def __init__(self, parent, text, subtext="", command=None, **kwargs):
        super().__init__(parent, bg=parent["bg"], **kwargs)
        self._cmd = command
        self._sel = False

        self._card = tk.Frame(self, bg=WHITE,
                              highlightbackground=GRAY_200,
                              highlightthickness=1, cursor="hand2")
        self._card.pack(fill="both", expand=True)

        self._lbl = tk.Label(self._card, text=text, font=F_CHIP,
                             bg=WHITE, fg=GRAY_800, anchor="w", padx=7, pady=3)
        self._lbl.pack(side="top", fill="x")

        if subtext:
            self._sub = tk.Label(self._card, text=subtext, font=F_CHIP_S,
                                 bg=WHITE, fg=GRAY_400, anchor="w", padx=7, pady=0)
            self._sub.pack(side="top", fill="x")
        else:
            self._sub = None

        for w in (self._card, self._lbl) + ((self._sub,) if self._sub else ()):
            w.bind("<Button-1>", self._click)
            w.bind("<Enter>",    self._hover)
            w.bind("<Leave>",    self._leave)

    def _click(self, e=None): self._cmd and self._cmd()
    def _hover(self, e=None):
        if not self._sel:
            self._card.config(bg=RED_PALE, highlightbackground=RED_LIGHT)
            self._lbl.config(bg=RED_PALE)
            if self._sub: self._sub.config(bg=RED_PALE)
    def _leave(self, e=None):
        if not self._sel:
            self._card.config(bg=WHITE, highlightbackground=GRAY_200)
            self._lbl.config(bg=WHITE)
            if self._sub: self._sub.config(bg=WHITE)
    def select(self):
        self._sel = True
        self._card.config(bg=RED, highlightbackground=RED)
        self._lbl.config(bg=RED, fg=WHITE)
        if self._sub: self._sub.config(bg=RED, fg=WHITE)
    def deselect(self):
        self._sel = False
        self._card.config(bg=WHITE, highlightbackground=GRAY_200)
        self._lbl.config(bg=WHITE, fg=GRAY_800)
        if self._sub: self._sub.config(bg=WHITE, fg=GRAY_400)


# ════════════════════════════════════════════════════════════════
class MesinButton(tk.Frame):
    def __init__(self, parent, letter, pos, command=None, **kwargs):
        super().__init__(parent, bg=parent["bg"], **kwargs)
        self._cmd = command
        self._sel = False
        self._disabled = False

        self._card = tk.Frame(self, bg=WHITE,
                              highlightbackground=GRAY_200,
                              highlightthickness=1, cursor="hand2")
        self._card.pack(fill="both", expand=True)

        self._let = tk.Label(self._card, text=letter, font=F_MESIN, bg=WHITE, fg=RED, pady=4)
        self._let.pack()
        self._pos = tk.Label(self._card, text=pos, font=F_MESIN_S, bg=WHITE, fg=GRAY_400, pady=0)
        self._pos.pack()

        for w in (self._card, self._let, self._pos):
            w.bind("<Button-1>", self._click)
            w.bind("<Enter>",    self._hover)
            w.bind("<Leave>",    self._leave)

    def _click(self, e=None):
        if not self._disabled and self._cmd: self._cmd()
    def _hover(self, e=None):
        if not self._sel and not self._disabled:
            for w in (self._card, self._let, self._pos): w.config(bg=RED_PALE)
            self._card.config(highlightbackground=RED_LIGHT)
    def _leave(self, e=None):
        if not self._sel and not self._disabled:
            for w in (self._card, self._let, self._pos): w.config(bg=WHITE)
            self._card.config(highlightbackground=GRAY_200)
    def select(self):
        self._sel = True; self._disabled = False
        self._card.config(cursor="hand2")
        for w in (self._card, self._let, self._pos): w.config(bg=RED)
        self._let.config(fg=WHITE); self._pos.config(fg=WHITE)
        self._card.config(highlightbackground=RED_DARK)
    def deselect(self):
        self._sel = False
        for w in (self._card, self._let, self._pos): w.config(bg=WHITE)
        self._let.config(fg=RED); self._pos.config(fg=GRAY_400)
        self._card.config(highlightbackground=GRAY_200)
    def disable(self):
        self._disabled = True; self._sel = False
        self._card.config(cursor="arrow")
        for w in (self._card, self._let, self._pos): w.config(bg=GRAY_100)
        self._let.config(fg=GRAY_400); self._pos.config(fg=GRAY_200)
        self._card.config(highlightbackground=GRAY_200)
    def enable(self):
        self._disabled = False
        self._card.config(cursor="hand2")
        for w in (self._card, self._let, self._pos): w.config(bg=WHITE)
        self._let.config(fg=RED); self._pos.config(fg=GRAY_400)
        self._card.config(highlightbackground=GRAY_200)


# ════════════════════════════════════════════════════════════════
class ScaleApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Timbangan AND GX-4000 — Quality Control")
        self.root.configure(bg=BG_ROOT)
        self.root.resizable(True, True)

        # State
        self.serial_conn      = None
        self.thread           = None
        self.is_reading       = False
        self.data_queue       = queue.Queue()
        self.current_data     = None
        self.saved_data       = []
        self.sel_variant      = None
        self.sel_machine      = None
        self.sel_filler_val   = ""
        self.nik_confirmed    = False
        self.live_weight      = 0.0
        self.live_unit        = "g"
        self.auto_port        = None
        self.filtered_variant = None
        self._active_config   = "7E1"
        self._reconnecting    = False
        self._confirm_pending = False
        self._frozen_data     = None

        self.variant_btns  = {}
        self.mesin_btns    = {}
        self._row_widgets  = []
        self._filler_chips = {}

        self._build_ui()
        self._bind_keys()
        self._auto_detect_port()
        self._check_queue()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── BUILD UI ────────────────────────────────────────────────
    def _build_ui(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(2, weight=1)
        self._build_header()
        self._build_stepbar()
        self._build_body()
        self._build_footer()

    def _bind_keys(self):
        self.root.bind("<space>",  self._space_save)
        self.root.bind("<Escape>", self._space_cancel)

    # ─────────────────────────────────────────────────────────────
    # SPACE / SAVE LOGIC
    # ─────────────────────────────────────────────────────────────
    def _space_save(self, event=None):
        if isinstance(self.root.focus_get(), tk.Entry):
            return
        if not self._is_ready():
            return

        if self._confirm_pending:
            frozen = self._frozen_data
            if frozen is None:
                return
            if _is_abnormal(frozen["weight"], VARIANT_STANDARDS.get(self.sel_variant)):
                self._alert_abnormal(frozen["weight"], VARIANT_STANDARDS[self.sel_variant])
                self._confirm_pending = False
                self._hide_confirm_banner()
                return
            self._confirm_pending = False
            self._hide_confirm_banner()
            self._save_data(source_override=frozen)
        else:
            if not self.current_data:
                messagebox.showwarning("Tidak Ada Data",
                    "Timbangan belum terbaca. Pastikan kabel terhubung.")
                return
            import copy
            frozen = copy.deepcopy(self.current_data)
            if _is_abnormal(frozen["weight"], VARIANT_STANDARDS.get(self.sel_variant)):
                self._alert_abnormal(frozen["weight"], VARIANT_STANDARDS[self.sel_variant])
                return
            self._frozen_data     = frozen
            self._confirm_pending = True
            self._show_confirm_banner()

    def _space_cancel(self, event=None):
        if self._confirm_pending:
            self._confirm_pending = False
            self._hide_confirm_banner()

    def _is_ready(self):
        return (self.nik_confirmed and self.sel_variant
                and self.sel_machine and self.sel_filler_val)

    def _alert_abnormal(self, weight, std_data):
        rev_lo = std_data.get("rev_lo", std_data["std"] / 2.0)
        rev_hi = std_data.get("rev_hi", std_data["std"] * 1.5)
        msg = (
            f"DATA ABNORMAL — tidak dapat disimpan!\n\n"
            f"Berat terbaca        : {weight:.2f} g\n"
            f"Standar variant      : {std_data['std']:.2f} g\n\n"
            f"Batas penerimaan sistem:\n"
            f"  Bawah (rev) : {rev_lo:.2f} g\n"
            f"  Atas  (rev) : {rev_hi:.2f} g\n\n"
            f"Kemungkinan penyebab:\n"
            f"  • Timbangan kosong / berat 0 atau minus\n"
            f"  • Produk salah variant\n"
            f"  • Gangguan pembacaan sensor\n\n"
            f"Letakkan produk dengan benar lalu coba lagi."
        )
        messagebox.showerror("⚠ Data Abnormal", msg)

    # ─────────────────────────────────────────────────────────────
    # HEADER
    # ─────────────────────────────────────────────────────────────
    def _build_header(self):
        hdr = tk.Frame(self.root, bg=RED, height=60)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.columnconfigure(1, weight=1)
        hdr.grid_propagate(False)

        tk.Label(hdr, text="⚖", font=("Segoe UI", 18),
                 bg=RED, fg=WHITE, padx=16).grid(row=0, column=0, sticky="ns", pady=10)

        tf = tk.Frame(hdr, bg=RED)
        tf.grid(row=0, column=1, sticky="w", pady=10)
        tk.Label(tf, text="Timbangan AND GX-4000", font=F_TITLE, bg=RED, fg=WHITE).pack(anchor="w")
        tk.Label(tf, text="Quality Control · Data Logger", font=F_SUB, bg=RED, fg=WHITE).pack(anchor="w")

        badge = tk.Frame(hdr, bg=RED, highlightbackground=WHITE, highlightthickness=1)
        badge.grid(row=0, column=2, sticky="e", padx=(0, 8), pady=15)
        self.conn_dot = tk.Label(badge, text="●", font=("Segoe UI", 10), bg=RED, fg=GRAY_400)
        self.conn_dot.pack(side="left", padx=(10, 4))
        self.conn_lbl = tk.Label(badge, text="Mendeteksi CH340...", font=F_SMALL, bg=RED, fg=WHITE)
        self.conn_lbl.pack(side="left", padx=(0, 6))

        tk.Button(hdr, text="↺ Reconnect",
                  font=F_BTN_SM, bg=RED_DARK, fg=WHITE, relief="flat",
                  cursor="hand2", activebackground=RED_DARK, activeforeground=WHITE,
                  padx=10, pady=4, command=self._manual_reconnect).grid(
            row=0, column=3, sticky="e", padx=12, pady=15)

    # ─────────────────────────────────────────────────────────────
    # STEP BAR
    # ─────────────────────────────────────────────────────────────
    def _build_stepbar(self):
        bar = tk.Frame(self.root, bg=RED_PALE,
                       highlightbackground=RED_SOFT, highlightthickness=1)
        bar.grid(row=1, column=0, sticky="ew")
        inner = tk.Frame(bar, bg=RED_PALE)
        inner.pack(side="left", padx=24, pady=10)

        steps = [("1","NIK Operator"), ("2","Variant & Mesin"), ("3","Catat Timbangan")]
        self.step_nums = []; self.step_lbls = []
        for i, (num, lbl) in enumerate(steps):
            if i > 0:
                tk.Label(inner, text="──", font=("Segoe UI", 9),
                         bg=RED_PALE, fg=RED_SOFT).pack(side="left", padx=4)
            sf = tk.Frame(inner, bg=RED_PALE)
            sf.pack(side="left")
            nf = tk.Label(sf, text=num, font=("Segoe UI", 9, "bold"),
                          bg=RED_SOFT, fg=RED, width=2, relief="flat")
            nf.pack(side="left", padx=(0, 6))
            ll = tk.Label(sf, text=lbl, font=("Segoe UI", 9, "bold"),
                          bg=RED_PALE, fg=GRAY_400)
            ll.pack(side="left")
            self.step_nums.append(nf); self.step_lbls.append(ll)
        self._set_step(0)

    def _set_step(self, active):
        for i, (nf, ll) in enumerate(zip(self.step_nums, self.step_lbls)):
            if i < active:
                nf.config(bg=GREEN, fg=WHITE, text="✓"); ll.config(fg=GREEN)
            elif i == active:
                nf.config(bg=RED, fg=WHITE, text=str(i+1)); ll.config(fg=RED)
            else:
                nf.config(bg=RED_SOFT, fg=RED, text=str(i+1)); ll.config(fg=GRAY_400)

    # ─────────────────────────────────────────────────────────────
    # BODY
    # ─────────────────────────────────────────────────────────────
    def _build_body(self):
        body = tk.Frame(self.root, bg=WHITE)
        body.grid(row=2, column=0, sticky="nsew")
        body.columnconfigure(2, weight=1)
        body.rowconfigure(0, weight=1)
        self._build_col1(body)
        self._build_col2(body)
        self._build_col3(body)

    def _build_col1(self, parent):
        col = tk.Frame(parent, bg=WHITE,
                       highlightbackground=GRAY_100, highlightthickness=1)
        col.grid(row=0, column=0, sticky="nsew")
        parent.columnconfigure(0, minsize=320)

        inner = tk.Frame(col, bg=WHITE)
        inner.pack(fill="both", expand=True, padx=18, pady=16)

        tk.Label(inner, text="NIK OPERATOR", font=F_LABEL,
                 bg=WHITE, fg=RED).pack(anchor="w", pady=(0, 6))
        nik_row = tk.Frame(inner, bg=WHITE)
        nik_row.pack(fill="x")
        self.nik_entry = tk.Entry(nik_row, font=F_NIK, bg=GRAY_50, fg=GRAY_800,
                                  insertbackground=RED, relief="flat",
                                  highlightbackground=GRAY_200, highlightthickness=1, width=12)
        self.nik_entry.pack(side="left", fill="y", ipady=6, padx=(0, 6))
        self.nik_entry.bind("<Return>", lambda e: self._confirm_nik())
        self.nik_entry.bind("<Key>",    lambda e: self._nik_reset_style())
        tk.Button(nik_row, text="Masuk", font=F_BTN_SM,
                  bg=RED, fg=WHITE, relief="flat",
                  activebackground=RED_DARK, activeforeground=WHITE,
                  cursor="hand2", padx=10, command=self._confirm_nik).pack(side="left", fill="y")
        self.nik_ok_lbl = tk.Label(inner, text="", font=("Segoe UI", 9), bg=WHITE, fg=GREEN)
        self.nik_ok_lbl.pack(anchor="w", pady=(4, 0))

        tk.Frame(inner, bg=GRAY_100, height=1).pack(fill="x", pady=12)

        tk.Label(inner, text="VARIANT PRODUK", font=F_LABEL,
                 bg=WHITE, fg=RED).pack(anchor="w", pady=(0, 8))
        vf = tk.Frame(inner, bg=WHITE)
        vf.pack(fill="x")
        for i, (name, data) in enumerate(VARIANT_STANDARDS.items()):
            row_i = i // 2; ci = i % 2
            chip = ChipButton(vf, text=name, subtext=data["code"],
                              command=lambda n=name: self._pick_variant(n))
            chip.grid(row=row_i, column=ci, sticky="ew",
                      padx=(0 if ci == 0 else 3, 3 if ci == 0 else 0), pady=2)
            self.variant_btns[name] = chip
        vf.columnconfigure(0, weight=1); vf.columnconfigure(1, weight=1)

    def _build_col2(self, parent):
        col = tk.Frame(parent, bg=WHITE,
                       highlightbackground=GRAY_100, highlightthickness=1)
        col.grid(row=0, column=1, sticky="nsew")
        parent.columnconfigure(1, minsize=300)

        inner = tk.Frame(col, bg=WHITE)
        inner.pack(fill="both", expand=True, padx=18, pady=16)

        tk.Label(inner, text="PILIH MESIN", font=F_LABEL,
                 bg=WHITE, fg=RED).pack(anchor="w", pady=(0, 4))
        self.mesin_hint_lbl = tk.Label(
            inner,
            text="Pilih variant terlebih dahulu untuk melihat mesin yang tersedia",
            font=F_SMALL, bg=WHITE, fg=GRAY_400, wraplength=260, justify="left")
        self.mesin_hint_lbl.pack(anchor="w", pady=(0, 8))

        gf = tk.Frame(inner, bg=WHITE)
        gf.pack(fill="x")
        COLS = 4
        for i, (letter, pos) in enumerate(MACHINES):
            row_i = i // COLS; ci = i % COLS
            btn = MesinButton(gf, letter=letter, pos=pos,
                              command=lambda l=letter: self._pick_mesin(l))
            btn.grid(row=row_i, column=ci, sticky="ew", padx=2, pady=2, ipady=2)
            self.mesin_btns[letter] = btn
            btn.disable()
        for c in range(COLS):
            gf.columnconfigure(c, weight=1)

    def _build_col3(self, parent):
        col = tk.Frame(parent, bg=WHITE)
        col.grid(row=0, column=2, sticky="nsew")

        inner = tk.Frame(col, bg=WHITE)
        inner.pack(fill="both", expand=True, padx=18, pady=16)
        inner.rowconfigure(7, weight=1)
        inner.columnconfigure(0, weight=1)

        wcard = tk.Frame(inner, bg=RED_PALE,
                         highlightbackground=RED_SOFT, highlightthickness=1)
        wcard.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        tk.Label(wcard, text="BERAT TERBACA", font=F_LABEL,
                 bg=RED_PALE, fg=RED, pady=10).pack()
        self.weight_lbl = tk.Label(wcard, text="–––.––",
                                   font=F_WEIGHT, bg=RED_PALE, fg=RED)
        self.weight_lbl.pack()
        self.unit_lbl = tk.Label(wcard, text="gram",
                                 font=F_UNIT, bg=RED_PALE, fg=GRAY_400)
        self.unit_lbl.pack(pady=(0, 4))

        status_row = tk.Frame(wcard, bg=RED_PALE)
        status_row.pack(pady=(0, 6))
        self.status_dot = tk.Label(status_row, text="●", font=("Segoe UI", 10),
                                   bg=RED_PALE, fg=GRAY_400)
        self.status_dot.pack(side="left", padx=(0, 4))
        self.status_txt = tk.Label(status_row, text="Pilih variant untuk validasi",
                                   font=F_STATUS, bg=RED_PALE, fg=GRAY_600)
        self.status_txt.pack(side="left")

        self.abnormal_banner = tk.Frame(wcard, bg=ABNORMAL_BG,
                                        highlightbackground=ABNORMAL_BORDER,
                                        highlightthickness=1)
        self.abnormal_lbl = tk.Label(self.abnormal_banner, text="",
                                     font=("Segoe UI", 8, "bold"),
                                     bg=ABNORMAL_BG, fg=ABNORMAL_FG,
                                     wraplength=340, justify="center", pady=4)
        self.abnormal_lbl.pack(padx=8, pady=4)

        std_row = tk.Frame(inner, bg=WHITE)
        std_row.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        for c in range(5): std_row.columnconfigure(c, weight=1)
        self.std_tu2 = self._std_pill(std_row, "TU2",     "–", GRAY_400,  0)
        self.std_tu1 = self._std_pill(std_row, "TU1",     "–", GRAY_600,  1)
        self.std_min = self._std_pill(std_row, "Min",     "–", RED_ALERT, 2)
        self.std_std = self._std_pill(std_row, "Standar", "–", GRAY_800,  3)
        self.std_max = self._std_pill(std_row, "Max",     "–", RED,       4)

        self.valid_lbl = tk.Label(inner, text="", font=("Segoe UI", 9),
                                  bg=WHITE, fg=GRAY_400)
        self.valid_lbl.grid(row=2, column=0, sticky="w", pady=(0, 6))

        self._build_filler_row(inner, row=3)

        btn_row = tk.Frame(inner, bg=WHITE)
        btn_row.grid(row=4, column=0, sticky="ew", pady=(0, 14))
        btn_row.columnconfigure(0, weight=1)

        self.confirm_banner = tk.Frame(btn_row, bg=CONFIRM_BG,
                                       highlightbackground=CONFIRM_BORDER,
                                       highlightthickness=1)
        bi = tk.Frame(self.confirm_banner, bg=CONFIRM_BG)
        bi.pack(fill="x", padx=10, pady=8)
        cl = tk.Frame(bi, bg=CONFIRM_BG)
        cl.pack(side="left", fill="x", expand=True)
        tk.Label(cl, text="⚠  Apakah data sudah sesuai?",
                 font=("Segoe UI", 10, "bold"),
                 bg=CONFIRM_BG, fg=CONFIRM_FG).pack(anchor="w")
        tk.Label(cl,
                 text="Tekan SPACE sekali lagi atau klik Konfirmasi untuk menyimpan",
                 font=F_SMALL, bg=CONFIRM_BG, fg=CONFIRM_FG).pack(anchor="w", pady=(2, 0))
        bb = tk.Frame(bi, bg=CONFIRM_BG)
        bb.pack(side="right", padx=(10, 0))
        tk.Button(bb, text="✕ Batal  [ESC]",
                  font=F_BTN_SM, bg=CONFIRM_BTN, fg=CONFIRM_FG, relief="flat",
                  cursor="hand2", activebackground="#FCD34D", activeforeground=CONFIRM_FG,
                  padx=8, pady=4, command=self._space_cancel).pack(side="left", padx=(0, 6))
        tk.Button(bb, text="✓ Konfirmasi  [SPACE]",
                  font=F_BTN_SM, bg=CONFIRM_OK, fg=WHITE, relief="flat",
                  cursor="hand2", activebackground="#166534", activeforeground=WHITE,
                  padx=8, pady=4, command=self._confirm_and_save).pack(side="left")

        self.save_btn = tk.Button(
            btn_row, text="Simpan Data Timbangan",
            font=F_BTN, bg=GRAY_200, fg=GRAY_400,
            relief="flat", cursor="arrow",
            activebackground=RED_DARK, activeforeground=WHITE,
            command=self._space_save, state="disabled", pady=10)
        self.save_btn.grid(row=0, column=0, sticky="ew")

        self.space_hint = tk.Label(btn_row, text="",
                                   font=F_SMALL, bg=WHITE, fg=GRAY_400)
        self.space_hint.grid(row=1, column=0, sticky="e", pady=(2, 0))

        tbl_hdr = tk.Frame(inner, bg=WHITE)
        tbl_hdr.grid(row=6, column=0, sticky="ew", pady=(0, 6))
        tbl_hdr.columnconfigure(0, weight=1)
        tk.Label(tbl_hdr, text="DATA TERSIMPAN", font=F_LABEL,
                 bg=WHITE, fg=RED).grid(row=0, column=0, sticky="w")

        ri = tk.Frame(tbl_hdr, bg=WHITE)
        ri.grid(row=0, column=1, sticky="e")
        self.filter_lbl = tk.Label(ri, text="", font=F_SMALL, bg=WHITE, fg=RED)
        self.filter_lbl.pack(side="left", padx=(0, 6))
        self.count_lbl = tk.Label(ri, text="0 data", font=F_SMALL, bg=WHITE, fg=GRAY_400)
        self.count_lbl.pack(side="left")
        self.reset_filter_btn = tk.Button(
            ri, text="Tampilkan Semua",
            font=F_SMALL, bg=WHITE, fg=GRAY_600, relief="flat",
            highlightbackground=GRAY_200, highlightthickness=1,
            cursor="hand2", padx=6, pady=1, command=self._reset_filter)
        self.reset_filter_btn.pack(side="left", padx=(6, 0))
        self.reset_filter_btn.pack_forget()

        self._build_table(inner, row=7)

    def _build_filler_row(self, parent, row):
        outer = tk.Frame(parent, bg=WHITE,
                         highlightbackground=GRAY_200, highlightthickness=1)
        outer.grid(row=row, column=0, sticky="ew", pady=(0, 10))

        lbl_f = tk.Frame(outer, bg=RED, width=80)
        lbl_f.pack(side="left", fill="y")
        lbl_f.pack_propagate(False)
        tk.Label(lbl_f, text="FILLER", font=F_LABEL,
                 bg=RED, fg=WHITE).pack(expand=True)

        chips_area = tk.Frame(outer, bg=WHITE)
        chips_area.pack(side="left", fill="both", expand=True, padx=10, pady=6)

        self.filler_hint_lbl = tk.Label(
            chips_area, text="Pilih variant dulu",
            font=F_SMALL, bg=WHITE, fg=GRAY_400)
        self.filler_hint_lbl.pack(side="left", padx=(0, 8))

        self._chips_frame = tk.Frame(chips_area, bg=WHITE)
        self._chips_frame.pack(side="left")

        for val in ALL_FILLER_VALUES:
            chip = FillerChip(self._chips_frame, value=val,
                              on_select=self._pick_filler)
            chip.pack(side="left", padx=2)
            chip.disable()
            self._filler_chips[val] = chip

        self.filler_selected_lbl = tk.Label(
            chips_area, text="",
            font=("Segoe UI", 8, "bold"), bg=WHITE, fg=GREEN)
        self.filler_selected_lbl.pack(side="left", padx=(10, 0))

    def _std_pill(self, parent, label, val, color, col):
        f = tk.Frame(parent, bg=WHITE, highlightbackground=GRAY_200, highlightthickness=1)
        f.grid(row=0, column=col, sticky="ew", padx=(0, 3) if col < 4 else 0)
        tk.Label(f, text=label, font=F_SMALL, bg=WHITE, fg=GRAY_400, pady=4).pack()
        lbl = tk.Label(f, text=val, font=F_BODY_B, bg=WHITE, fg=color, pady=4)
        lbl.pack()
        return lbl

    def _build_table(self, parent, row):
        tbl_frame = tk.Frame(parent, bg=WHITE)
        tbl_frame.grid(row=row, column=0, sticky="nsew")
        parent.rowconfigure(row, weight=1)
        tbl_frame.columnconfigure(0, weight=1)
        tbl_frame.rowconfigure(1, weight=1)

        cols   = ["#", "Waktu", "Variant", "Mesin", "Filler", "Berat", "Status"]
        widths = [30, 110, 160, 60, 55, 90, 70]

        hdr = tk.Frame(tbl_frame, bg=GRAY_50)
        hdr.grid(row=0, column=0, sticky="ew")
        for c, w in zip(cols, widths):
            tk.Label(hdr, text=c, font=F_TABLE_H, bg=GRAY_50,
                     fg=GRAY_400, width=w//8, anchor="w",
                     padx=6, pady=5).pack(side="left")

        canvas = tk.Canvas(tbl_frame, bg=WHITE, highlightthickness=0)
        sb = tk.Scrollbar(tbl_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        canvas.grid(row=1, column=0, sticky="nsew")
        sb.grid(row=1, column=1, sticky="ns")

        self.table_inner  = tk.Frame(canvas, bg=WHITE)
        self.table_window = canvas.create_window(0, 0, window=self.table_inner, anchor="nw")
        self._table_canvas = canvas

        def _on_fc(e): canvas.configure(scrollregion=canvas.bbox("all"))
        def _on_cc(e): canvas.itemconfig(self.table_window, width=e.width)
        self.table_inner.bind("<Configure>", _on_fc)
        canvas.bind("<Configure>", _on_cc)
        self._table_row_count = 0

    def _make_row_widget(self, parent, values, ok=True, row_index=0):
        bg     = WHITE if row_index % 2 == 0 else GRAY_50
        widths = [30, 110, 160, 60, 55, 90, 70]
        row_f = tk.Frame(parent, bg=bg)
        tk.Frame(row_f, bg=GRAY_100, height=1).pack(fill="x")
        cells = tk.Frame(row_f, bg=bg)
        cells.pack(fill="x")
        for i, (val, w) in enumerate(zip(values, widths)):
            if i == 6:
                badge_bg = GREEN_LIGHT if ok else "#FEE2E2"
                badge_fg = GREEN       if ok else RED_ALERT
                tk.Label(cells, text=val, font=F_BADGE,
                         bg=badge_bg, fg=badge_fg,
                         padx=6, pady=1).pack(side="left", padx=6, pady=4)
            else:
                font = F_MONO if i in (0, 1, 5) else F_TABLE
                tk.Label(cells, text=val, font=font,
                         bg=bg, fg=GRAY_600,
                         width=w//8, anchor="w", padx=6).pack(side="left")
        return row_f

    def _insert_row_top(self, record):
        if self.filtered_variant and record["variant"] != self.filtered_variant:
            self._update_count_label(); return
        for idx, row_f in enumerate(self._row_widgets):
            new_bg = GRAY_50 if idx % 2 == 0 else WHITE
            row_f.config(bg=new_bg)
            children = row_f.winfo_children()
            if len(children) >= 2:
                cells_frame = children[1]
                cells_frame.config(bg=new_bg)
                for widget in cells_frame.winfo_children():
                    try:
                        if widget.cget("bg") not in (GREEN_LIGHT, "#FEE2E2"):
                            widget.config(bg=new_bg)
                    except Exception: pass
        ok        = record["status"] == "OK"
        t_short   = record["timestamp"][11:]
        var_short = record["variant"][:16] + ("…" if len(record["variant"]) > 16 else "")
        row_num   = len(self.saved_data)
        filler    = record.get("filler", "–")
        values    = [str(row_num), t_short, var_short, record["machine"],
                     filler, f"{record['weight']:.2f}g", record["status"]]
        new_row   = self._make_row_widget(self.table_inner, values, ok=ok, row_index=0)
        if self._row_widgets:
            new_row.pack(fill="x", before=self._row_widgets[0])
        else:
            new_row.pack(fill="x")
        self._row_widgets.insert(0, new_row)
        self._table_row_count += 1
        self._table_canvas.yview_moveto(0)
        self._update_count_label()

    def _update_count_label(self):
        if self.filtered_variant:
            shown = sum(1 for d in self.saved_data if d["variant"] == self.filtered_variant)
            total = len(self.saved_data)
            short = (self.filtered_variant[:18] + "…"
                     if len(self.filtered_variant) > 18 else self.filtered_variant)
            self.filter_lbl.config(text=f"▶ {short}", fg=RED)
            self.reset_filter_btn.pack(side="left", padx=(6, 0))
            self.count_lbl.config(
                text=f"{shown} dari {total} data" if shown != total else f"{total} data",
                fg=RED if shown != total else GRAY_400)
        else:
            total = len(self.saved_data)
            self.filter_lbl.config(text="", fg=GRAY_400)
            self.reset_filter_btn.pack_forget()
            self.count_lbl.config(text=f"{total} data" if total else "0 data", fg=GRAY_400)

    def _rebuild_table(self):
        for w in self.table_inner.winfo_children(): w.destroy()
        self._table_row_count = 0; self._row_widgets = []
        data = ([d for d in self.saved_data if d["variant"] == self.filtered_variant]
                if self.filtered_variant else self.saved_data)
        for i, d in enumerate(reversed(data), 1):
            ok        = d["status"] == "OK"
            t_short   = d["timestamp"][11:]
            var_short = d["variant"][:16] + ("…" if len(d["variant"]) > 16 else "")
            filler    = d.get("filler", "–")
            row_f = self._make_row_widget(
                self.table_inner,
                [str(i), t_short, var_short, d["machine"],
                 filler, f"{d['weight']:.2f}g", d["status"]],
                ok=ok, row_index=i-1)
            row_f.pack(fill="x")
            self._row_widgets.append(row_f)
            self._table_row_count += 1
        self.root.after(30, lambda: self._table_canvas.yview_moveto(0))
        self._update_count_label()

    def _refresh_table(self): self._rebuild_table()
    def _reset_filter(self): self.filtered_variant = None; self._refresh_table()

    # ─────────────────────────────────────────────────────────────
    # FOOTER
    # ─────────────────────────────────────────────────────────────
    def _build_footer(self):
        ft = tk.Frame(self.root, bg=GRAY_50,
                      highlightbackground=GRAY_100, highlightthickness=1)
        ft.grid(row=3, column=0, sticky="ew")
        self.footer_lbl = tk.Label(
            ft, text=f"Port: Mendeteksi CH340... · API: {API_URL}",
            font=F_SMALL, bg=GRAY_50, fg=GRAY_400)
        self.footer_lbl.pack(side="left", padx=16, pady=6)
        tk.Button(ft, text="↓ Export Excel",
                  font=F_BTN_SM, bg=WHITE, fg=GRAY_600, relief="flat",
                  highlightbackground=GRAY_200, highlightthickness=1,
                  cursor="hand2", padx=10, pady=4,
                  command=self._export_excel).pack(side="right", padx=12, pady=6)
        tk.Label(ft, text="[SPACE] = Konfirmasi & Simpan",
                 font=F_SMALL, bg=GRAY_50, fg=GRAY_200).pack(side="right", padx=0, pady=6)

    # ─────────────────────────────────────────────────────────────
    # SERIAL  ← v11: HANYA CH340
    # ─────────────────────────────────────────────────────────────
    def _auto_detect_port(self):
        def _detect():
            all_ports = serial.tools.list_ports.comports()

            # ── v11: filter keras — hanya CH340 ──────────────────
            ch340_ports = [p for p in all_ports if _is_ch340_port(p)]

            print(f"[PORT SCAN] {len(all_ports)} port total — "
                  f"{len(ch340_ports)} CH340 ditemukan:")
            for p in all_ports:
                tag = "✓ CH340" if _is_ch340_port(p) else "✗ skip"
                print(f"  [{tag}]  {p.device:10s}  desc={p.description!r}")

            if not ch340_ports:
                self.root.after(0, self._set_conn_status, False,
                                "USB-SERIAL CH340 tidak ditemukan")
                self.root.after(5000, self._auto_detect_port)
                return

            cfgs = [
                {"bytesize": DATABITS, "parity": serial.PARITY_EVEN, "label": "7E1"},
                {"bytesize": 8,        "parity": serial.PARITY_NONE, "label": "8N1"},
            ]

            for p in ch340_ports:
                for cfg in cfgs:
                    conn = self._try_open_serial(p.device, cfg)
                    if conn:
                        # Probe: tunggu 2 detik, cek data timbangan valid
                        print(f"[PROBE] {p.device} ({cfg['label']}) — menunggu data...")
                        deadline = time.time() + 2.0
                        valid = False
                        while time.time() < deadline:
                            try:
                                if conn.in_waiting > 0:
                                    raw = conn.readline().decode("ascii", errors="ignore").strip()
                                    if raw and re.match(
                                            r"^[A-Z]{2},[+\-]?\d+\.?\d*\s*[a-zA-Z]+$", raw):
                                        print(f"[PROBE] ✓ Data valid: {raw!r}")
                                        valid = True
                                        break
                            except Exception:
                                break
                            time.sleep(0.05)

                        if not valid:
                            print(f"[PROBE] ✗ Tidak ada data valid dari {p.device}, skip.")
                            try:
                                conn.close()
                            except Exception:
                                pass
                            continue

                        self.serial_conn    = conn
                        self.auto_port      = p.device
                        self.is_reading     = True
                        self._active_config = cfg["label"]
                        self.root.after(0, self._set_conn_status, True,
                                        f"{p.device} [CH340] [{cfg['label']}]")
                        self.thread = threading.Thread(
                            target=self._read_thread, daemon=True)
                        self.thread.start()
                        return

            # CH340 ditemukan tapi gagal konek / tidak ada data valid
            self.root.after(0, self._set_conn_status, False,
                            f"CH340 ditemukan ({len(ch340_ports)}x) — gagal konek")
            self.root.after(5000, self._auto_detect_port)

        threading.Thread(target=_detect, daemon=True).start()

    def _try_open_serial(self, port, cfg, retries=3, delay=1.0):
        for attempt in range(retries):
            try:
                conn = serial.Serial(port=port, baudrate=BAUDRATE,
                                    bytesize=cfg["bytesize"], parity=cfg["parity"],
                                    stopbits=STOPBITS, timeout=1)
                conn.reset_input_buffer()
                return conn
            except serial.SerialException as e:
                print(f"[SERIAL] {port} ({cfg['label']}) attempt {attempt+1}/{retries}: {e}")
                if attempt < retries - 1: time.sleep(delay)
        return None

    def _manual_reconnect(self):
        if self._reconnecting: return
        self._reconnecting = True
        self._set_conn_status(False, "Reconnecting...")

        def _do():
            self.is_reading = False
            if self.thread and self.thread.is_alive():
                self.thread.join(timeout=2.5)
            self.thread = None
            if self.serial_conn:
                try:
                    if self.serial_conn.is_open:
                        self.serial_conn.reset_input_buffer()
                        self.serial_conn.reset_output_buffer()
                        self.serial_conn.close()
                except Exception as e:
                    print(f"[SERIAL] close: {e}")
            self.serial_conn = None
            time.sleep(0.8)
            self._reconnecting = False
            self.root.after(0, self._auto_detect_port)

        threading.Thread(target=_do, daemon=True).start()

    def _set_conn_status(self, connected, detail):
        if connected:
            self.conn_dot.config(fg="#4ADE80")
            self.conn_lbl.config(text=f"{detail} · Terhubung")
            self.footer_lbl.config(text=f"Port: {detail} (auto) · 9600bps · API: {API_URL}")
        else:
            self.conn_dot.config(fg=GRAY_400)
            self.conn_lbl.config(text=detail)
            self.footer_lbl.config(text=f"Port: {detail} · API: {API_URL}")

    def _read_thread(self):
        while self.is_reading:
            try:
                if self.serial_conn and self.serial_conn.in_waiting > 0:
                    raw = self.serial_conn.readline().decode("ascii", errors="ignore").strip()
                    if raw:
                        parsed = self._parse(raw)
                        if parsed: self.data_queue.put(parsed)
            except Exception: break
            time.sleep(0.01)

    def _parse(self, raw):
        try:
            m = re.match(r"^([A-Z]{2}),([+\-]?\d+\.?\d*)\s*([a-zA-Z]+)$", raw)
            if m:
                return {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "scale_status": m.group(1),
                        "weight": float(m.group(2)),
                        "unit":   m.group(3),
                        "raw":    raw}
        except Exception: pass
        return None

    def _check_queue(self):
        try:
            while True:
                data = self.data_queue.get_nowait()
                self._update_weight(data)
        except queue.Empty: pass
        self.root.after(50, self._check_queue)

    def _update_weight(self, data):
        self.live_weight = data["weight"]
        self.live_unit   = data["unit"]
        if self._confirm_pending:
            self.current_data = data
            return
        self.current_data = data
        self.weight_lbl.config(text=f"{data['weight']:.2f}")
        unit_map = {"g": "gram", "kg": "kilogram"}
        self.unit_lbl.config(text=unit_map.get(data["unit"], data["unit"]))
        self._validate_display(data["weight"])

    # ─────────────────────────────────────────────────────────────
    # VALIDATE DISPLAY
    # ─────────────────────────────────────────────────────────────
    def _validate_display(self, weight):
        if not self.sel_variant:
            self.weight_lbl.config(fg=RED)
            self.status_dot.config(fg=GRAY_400)
            self.status_txt.config(text="Pilih variant untuk validasi", fg=GRAY_600)
            self.valid_lbl.config(text="")
            self._hide_abnormal_banner()
            return

        std = VARIANT_STANDARDS[self.sel_variant]
        rev_lo = std.get("rev_lo", std["std"] / 2.0)
        rev_hi = std.get("rev_hi", std["std"] * 1.5)

        if weight < rev_lo or weight > rev_hi:
            if weight <= 0:
                label = "ABNORMAL — Timbangan kosong / berat 0 atau minus"
            elif weight < rev_lo:
                label = f"ABNORMAL — Di luar batas sistem (< {rev_lo:.2f}g)"
            else:
                label = f"ABNORMAL — Di luar batas sistem (> {rev_hi:.2f}g)"
            self.weight_lbl.config(fg=ABNORMAL_FG)
            self.status_dot.config(fg=ABNORMAL_FG)
            self.status_txt.config(text=f"⛔ {label}", fg=ABNORMAL_FG)
            self.valid_lbl.config(
                text=f"Batas sistem: {rev_lo:.2f}–{rev_hi:.2f}g  |  Std={std['std']:.2f}g",
                fg=ABNORMAL_FG)
            self._show_abnormal_banner(weight, std)
            self._check_save_ready(force_disable_abnormal=True)
            return

        self._hide_abnormal_banner()
        tu2 = std.get("tu2"); tu1 = std.get("tu1")
        if weight < (tu2 if tu2 else std["min"]):
            color = RED_ALERT; label = "NOT OK — Berat Jauh di Bawah (< TU2)"
        elif tu2 and tu1 and weight < tu1:
            color = AMBER;     label = "WASPADA — TU2–TU1"
        elif tu1 and weight < std["min"]:
            color = AMBER;     label = "WASPADA — TU1–Min Standar"
        elif weight > std["max"]:
            color = RED_ALERT; label = "NOT OK — Berat Melebihi Maks"
        else:
            color = GREEN;     label = "OK — Sesuai Standar"
        self.weight_lbl.config(fg=color)
        self.status_dot.config(fg=color)
        self.status_txt.config(text=label, fg=color)
        self.valid_lbl.config(
            text=f"TU2: {tu2}g  TU1: {tu1}g  Range OK: {std['min']}–{std['max']}g",
            fg=GRAY_400)
        self._check_save_ready()

    def _show_abnormal_banner(self, weight, std_data):
        rev_lo = std_data.get("rev_lo", std_data["std"] / 2.0)
        rev_hi = std_data.get("rev_hi", std_data["std"] * 1.5)
        self.abnormal_lbl.config(
            text=f"⛔ DATA ABNORMAL — tidak dapat disimpan\n"
                 f"Batas penerimaan sistem: {rev_lo:.2f}g – {rev_hi:.2f}g  |  Berat: {weight:.2f}g")
        self.abnormal_banner.pack(fill="x", padx=8, pady=(0, 8))

    def _hide_abnormal_banner(self):
        self.abnormal_banner.pack_forget()

    # ─────────────────────────────────────────────────────────────
    # NIK
    # ─────────────────────────────────────────────────────────────
    def _confirm_nik(self):
        nik = self.nik_entry.get().strip()
        if not nik: return
        if nik not in VALID_NIKS:
            self.nik_entry.config(highlightbackground=RED_ALERT, highlightthickness=2)
            self.nik_ok_lbl.config(text=f"✕  NIK {nik} tidak terdaftar — akses ditolak",
                                   fg=RED_ALERT)
            self.nik_confirmed = False
            return
        self.nik_entry.config(highlightbackground=GREEN, highlightthickness=2)
        self.nik_confirmed = True
        self.nik_ok_lbl.config(text=f"✓  NIK {nik} — Operator Terverifikasi", fg=GREEN)
        self._set_step(1)
        self.root.focus_set()

    def _nik_reset_style(self):
        self.nik_entry.config(highlightbackground=GRAY_200, highlightthickness=1)
        if not self.nik_confirmed: self.nik_ok_lbl.config(text="", fg=GREEN)

    # ─────────────────────────────────────────────────────────────
    # FILLER
    # ─────────────────────────────────────────────────────────────
    def _pick_filler(self, val):
        for v, chip in self._filler_chips.items():
            chip.deselect()
        self._filler_chips[val].select()
        self.sel_filler_val = val
        self.filler_selected_lbl.config(text=f"✓ Filler #{val}", fg=GREEN)
        self.root.focus_set()
        self._check_save_ready()

    def _update_filler_chips(self, variant_name):
        self.sel_filler_val = ""
        self.filler_selected_lbl.config(text="")
        if variant_name is None:
            for chip in self._filler_chips.values():
                chip.deselect(); chip.disable()
            self.filler_hint_lbl.config(text="Pilih variant dulu", fg=GRAY_400)
            return
        max_f = VARIANT_FILLER_MAX.get(variant_name, 8)
        self.filler_hint_lbl.config(text=f"Nomor filler (1–{max_f}):", fg=GRAY_600)
        for val, chip in self._filler_chips.items():
            chip.deselect()
            if int(val) <= max_f: chip.enable()
            else:                 chip.disable()

    # ─────────────────────────────────────────────────────────────
    # VARIANT
    # ─────────────────────────────────────────────────────────────
    def _pick_variant(self, name):
        if not self.nik_confirmed:
            messagebox.showwarning("NIK Belum Dikonfirmasi",
                                   "Masukkan NIK operator terlebih dahulu.")
            return
        if self._confirm_pending:
            self._confirm_pending = False
            self._hide_confirm_banner()

        for n, btn in self.variant_btns.items():
            btn.deselect() if n != name else btn.select()
        self.sel_variant = name
        self.filtered_variant = name
        self.sel_machine = None

        std = VARIANT_STANDARDS[name]
        self.std_tu2.config(text=f"{std.get('tu2','–')}")
        self.std_tu1.config(text=f"{std.get('tu1','–')}")
        self.std_min.config(text=f"{std['min']:.2f}")
        self.std_std.config(text=f"{std['std']:.2f}")
        self.std_max.config(text=f"{std['max']:.2f}")

        allowed = VARIANT_MESIN.get(name)
        for letter, btn in self.mesin_btns.items():
            btn.deselect()
            if allowed is None or letter in allowed: btn.enable()
            else: btn.disable()
        self.mesin_hint_lbl.config(
            text=(f"{len(allowed)} mesin tersedia untuk variant ini"
                  if allowed else "Semua mesin tersedia untuk variant ini"),
            fg=GRAY_400)

        self._update_filler_chips(name)
        if self.live_weight: self._validate_display(self.live_weight)
        self._check_save_ready()
        self._refresh_table()

    # ─────────────────────────────────────────────────────────────
    # MESIN
    # ─────────────────────────────────────────────────────────────
    def _pick_mesin(self, letter):
        if not self.nik_confirmed:
            messagebox.showwarning("NIK Belum Dikonfirmasi",
                                   "Masukkan NIK operator terlebih dahulu.")
            return
        if self._confirm_pending:
            self._confirm_pending = False
            self._hide_confirm_banner()
        for l, btn in self.mesin_btns.items():
            if l != letter:
                if not self.mesin_btns[l]._disabled: btn.deselect()
            else: btn.select()
        self.sel_machine = letter
        self._check_save_ready()
        if self.sel_variant and self.sel_machine: self._set_step(2)

    # ─────────────────────────────────────────────────────────────
    # SAVE READINESS
    # ─────────────────────────────────────────────────────────────
    def _check_save_ready(self, force_disable_abnormal=False):
        ready = self._is_ready()
        if force_disable_abnormal:
            self.save_btn.config(state="disabled", bg=GRAY_200, fg=GRAY_400, cursor="arrow")
            self.space_hint.config(text="⛔ Berat abnormal — tidak dapat disimpan",
                                   fg=ABNORMAL_FG)
            return
        if ready and self.live_weight and self.sel_variant:
            if _is_abnormal(self.live_weight, VARIANT_STANDARDS.get(self.sel_variant)):
                self.save_btn.config(state="disabled", bg=GRAY_200, fg=GRAY_400, cursor="arrow")
                self.space_hint.config(text="⛔ Berat abnormal — tidak dapat disimpan",
                                       fg=ABNORMAL_FG)
                return
        if ready:
            self.save_btn.config(state="normal", bg=RED, fg=WHITE, cursor="hand2")
            self.space_hint.config(
                text="Tekan [SPACE] → konfirmasi → [SPACE] lagi untuk simpan", fg=GRAY_400)
            self._set_step(2)
        else:
            self.save_btn.config(state="disabled", bg=GRAY_200, fg=GRAY_400, cursor="arrow")
            self.space_hint.config(text="")

    # ─────────────────────────────────────────────────────────────
    # CONFIRM BANNER
    # ─────────────────────────────────────────────────────────────
    def _show_confirm_banner(self):
        frozen_w = self._frozen_data["weight"]
        self.weight_lbl.config(text=f"{frozen_w:.2f}", fg=AMBER)
        unit_map = {"g": "gram", "kg": "kilogram"}
        self.unit_lbl.config(text=unit_map.get(self._frozen_data["unit"],
                                                self._frozen_data["unit"]))
        self.status_dot.config(fg=AMBER)
        self.status_txt.config(text="⚠  Berat di-freeze — tekan SPACE untuk simpan", fg=AMBER)
        self.confirm_banner.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        self.save_btn.grid(row=1, column=0, sticky="ew")
        self.space_hint.grid(row=2, column=0, sticky="e", pady=(2, 0))

    def _hide_confirm_banner(self):
        self.confirm_banner.grid_forget()
        self.save_btn.grid(row=0, column=0, sticky="ew")
        self.space_hint.grid(row=1, column=0, sticky="e", pady=(2, 0))
        self.weight_lbl.config(text=f"{self.live_weight:.2f}")
        unit_map = {"g": "gram", "kg": "kilogram"}
        self.unit_lbl.config(text=unit_map.get(self.live_unit, self.live_unit))
        if self.live_weight and self.sel_variant:
            self._validate_display(self.live_weight)
        else:
            self.weight_lbl.config(fg=RED)

    def _confirm_and_save(self):
        frozen = self._frozen_data
        if frozen is None:
            return
        if _is_abnormal(frozen["weight"], VARIANT_STANDARDS.get(self.sel_variant)):
            self._alert_abnormal(frozen["weight"], VARIANT_STANDARDS[self.sel_variant])
            self._confirm_pending = False
            self._hide_confirm_banner()
            self._frozen_data = None
            return
        self._confirm_pending = False
        self._hide_confirm_banner()
        self._save_data(source_override=frozen)

    def _show_loading(self, msg="Mengirim data ke server..."):
        self._loading = tk.Toplevel(self.root)
        self._loading.overrideredirect(True)
        self._loading.attributes("-alpha", 0.82)
        self._loading.configure(bg=GRAY_800)
        self.root.update_idletasks()
        w = self.root.winfo_width(); h = self.root.winfo_height()
        x = self.root.winfo_rootx(); y = self.root.winfo_rooty()
        self._loading.geometry(f"{w}x{h}+{x}+{y}")
        self._loading.lift()
        frm = tk.Frame(self._loading, bg=WHITE,
                       highlightbackground=GRAY_200, highlightthickness=1)
        frm.place(relx=0.5, rely=0.5, anchor="center")
        tk.Label(frm, text="⏳  Mengirim data...", font=F_BODY_B,
                 bg=WHITE, fg=GRAY_800, padx=30, pady=10).pack()
        self._loading_lbl = tk.Label(frm, text=msg, font=F_SMALL,
                                     bg=WHITE, fg=GRAY_600, padx=30, pady=6)
        self._loading_lbl.pack()
        self._loading.update()

    def _update_loading(self, msg):
        if hasattr(self, "_loading_lbl") and self._loading_lbl.winfo_exists():
            self._loading_lbl.config(text=msg)
            self._loading.update()

    def _hide_loading(self):
        if hasattr(self, "_loading") and self._loading.winfo_exists():
            self._loading.destroy()

    def _save_data(self, source_override=None):
        source = source_override if source_override is not None else self.current_data
        if not source:
            messagebox.showwarning("Tidak Ada Data",
                                "Timbangan belum terbaca. Pastikan kabel terhubung.")
            return
        if not (self.sel_variant and self.sel_machine and self.nik_confirmed):
            messagebox.showerror("Lengkapi Data", "NIK, Variant, dan Mesin harus dipilih.")
            return

        w = source["weight"]
        std_data = VARIANT_STANDARDS.get(self.sel_variant)
        if std_data and _is_abnormal(w, std_data):
            self._alert_abnormal(w, std_data)
            return

        std    = VARIANT_STANDARDS[self.sel_variant]
        ok     = std["min"] <= w <= std["max"]
        status = "OK" if ok else "NOT OK"

        if not ok:
            if w < std["min"]:
                selisih = std["min"] - w
                keterangan = (
                    f"Berat terbaca  : {w:.2f} g\n"
                    f"Min standar    : {std['min']:.2f} g\n"
                    f"Selisih kurang : {selisih:.2f} g\n\n"
                    f"Berat DI BAWAH batas minimum standar."
                )
            else:
                selisih = w - std["max"]
                keterangan = (
                    f"Berat terbaca  : {w:.2f} g\n"
                    f"Max standar    : {std['max']:.2f} g\n"
                    f"Selisih lebih  : {selisih:.2f} g\n\n"
                    f"Berat MELEBIHI batas maksimum standar."
                )
            lanjut = messagebox.askyesno(
                "⚠ Berat NOT OK — Konfirmasi Pengiriman",
                f"Data ini berstatus NOT OK!\n\n"
                f"{keterangan}\n"
                f"Tetap kirim data NOT OK ini ke database?\n\n"
                f"  ▸ Klik YA   → data dikirim dengan status NOT OK\n"
                f"  ▸ Klik TIDAK → batal, data tidak dikirim",
                icon="warning")
            if not lanjut:
                return

        nik    = self.nik_entry.get().strip()
        filler = self.sel_filler_val
        form   = {"nik": nik, "mesin": self.sel_machine,
                "variant": self.sel_variant, "waktu": source["timestamp"],
                "berat": str(w), "unit": source["unit"],
                "status": status, "filler": filler}

        MAX_RETRY   = 3
        RETRY_DELAY = 2.0
        result = {"ok": False, "api_status": "Gagal", "error": ""}

        def _post_with_retry():
            for attempt in range(1, MAX_RETRY + 1):
                self.root.after(0, self._update_loading, f"Percobaan {attempt}/{MAX_RETRY}...")
                try:
                    resp = requests.post(API_URL, data=form, timeout=8)
                    if resp.status_code in (200, 201):
                        result["ok"]         = True
                        result["api_status"] = "✓ Terkirim"
                        break
                    else:
                        result["api_status"] = f"Error {resp.status_code}"
                        result["error"]      = f"Server mengembalikan HTTP {resp.status_code}"
                except requests.exceptions.ConnectionError as e:
                    result["api_status"] = "Offline"
                    result["error"]      = f"Tidak dapat terhubung ke server: {e}"
                except requests.exceptions.Timeout:
                    result["api_status"] = "Timeout"
                    result["error"]      = "Server tidak merespons (timeout 8 detik)"
                except Exception as e:
                    result["api_status"] = "Gagal"
                    result["error"]      = str(e)
                if attempt < MAX_RETRY:
                    self.root.after(0, self._update_loading,
                                    f"Gagal, mencoba ulang {attempt + 1}/{MAX_RETRY}...")
                    time.sleep(RETRY_DELAY)

            # Selesai — kembalikan ke main thread
            self.root.after(0, self._on_post_done, result, source, form)

        self._show_loading("Menghubungi server...")
        threading.Thread(target=_post_with_retry, daemon=True).start()
        # ← fungsi selesai di sini, TIDAK ada kode lagi setelah ini

    def _on_post_done(self, result, source, form):
        """Dipanggil dari main thread setelah POST selesai."""
        self._hide_loading()

        if not result["ok"]:
            messagebox.showerror(
                "❌ Gagal Mengirim Data",
                f"Data TIDAK tersimpan ke database!\n\n"
                f"Status  : {result['api_status']}\n"
                f"Detail  : {result['error']}\n\n"
                f"Sudah dicoba 3x.\n"
                f"Periksa koneksi jaringan lalu tekan SPACE lagi untuk mencoba ulang.")
            return

        nik    = form["nik"]
        filler = form["filler"]
        status = form["status"]

        record = {**source, "nik": nik, "machine": self.sel_machine,
                "variant": self.sel_variant, "status": status,
                "filler": filler, "api_status": result["api_status"]}
        self.saved_data.append(record)
        self._frozen_data = None
        self._insert_row_top(record)
        self._set_step(2)
        self.weight_lbl.config(fg=GREEN)
        self.root.after(400, lambda: (
            self._validate_display(self.live_weight)
            if self.live_weight and self.sel_variant
            else self.weight_lbl.config(fg=RED)))

    def _export_excel(self):
        if not self.saved_data:
            messagebox.showinfo("Kosong", "Belum ada data untuk di-export.")
            return
        fn = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"timbangan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not fn: return
        try:
            wb = Workbook(); ws = wb.active; ws.title = "Data Timbangan"
            thin = Side(style="thin"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.merge_cells("A1:K1")
            c = ws["A1"]
            c.value = "DATA TIMBANGAN — AND GX-4000 | Quality Control"
            c.font  = Font(name="Calibri", size=13, bold=True, color="E53E3E")
            c.fill  = PatternFill("solid", fgColor="FFF5F5")
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28
            headers = ["No","NIK","Mesin","Variant","Filler","Tanggal","Waktu","Berat","Unit","Status","API"]
            hfill = PatternFill("solid", fgColor="E53E3E")
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(2, ci, h)
                cell.font      = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                cell.fill      = hfill
                cell.alignment = Alignment(horizontal="center")
                cell.border    = border
            ws.row_dimensions[2].height = 20
            fill_a   = PatternFill("solid", fgColor="FFFFFF")
            fill_b   = PatternFill("solid", fgColor="FFF5F5")
            fill_ok  = PatternFill("solid", fgColor="DCFCE7")
            fill_nok = PatternFill("solid", fgColor="FEE2E2")
            count_ok = count_nok = 0
            for i, d in enumerate(self.saved_data, 1):
                row = i + 2; fill = fill_a if i % 2 else fill_b
                dt = datetime.strptime(d["timestamp"], "%Y-%m-%d %H:%M:%S")
                is_ok = d["status"] == "OK"
                if is_ok: count_ok += 1
                else:     count_nok += 1
                vals = [i, d.get("nik",""), d["machine"], d["variant"],
                        d.get("filler","–"), dt.strftime("%Y-%m-%d"),
                        dt.strftime("%H:%M:%S"), d["weight"], d["unit"],
                        d["status"], d.get("api_status","–")]
                for ci, val in enumerate(vals, 1):
                    cell = ws.cell(row, ci, val)
                    cell.font      = Font(name="Calibri", size=10)
                    cell.fill      = (fill_ok  if (ci == 10 and is_ok) else
                                      fill_nok if (ci == 10 and not is_ok) else fill)
                    cell.border    = border
                    cell.alignment = Alignment(horizontal="center")
            from openpyxl.utils import get_column_letter
            for ci, w in enumerate([6,12,10,24,8,14,10,10,8,10,12], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            sr = len(self.saved_data) + 4
            ws.cell(sr,   1).value = "RINGKASAN"
            ws.cell(sr,   1).font  = Font(bold=True, color="E53E3E")
            ws.cell(sr+1, 1, f"Total  : {len(self.saved_data)}")
            ws.cell(sr+2, 1, f"OK     : {count_ok}")
            ws.cell(sr+3, 1, f"NOT OK : {count_nok}")
            wb.save(fn)
            messagebox.showinfo(
                "Export Berhasil",
                f"Data berhasil disimpan!\n\nFile  : {fn}\n"
                f"Total : {len(self.saved_data)} record\n"
                f"OK    : {count_ok}  |  NOT OK: {count_nok}")
        except Exception as e:
            messagebox.showerror("Export Gagal", str(e))

    def _on_close(self):
        if not messagebox.askokcancel(
            "Tutup Aplikasi",
            "Apakah Anda yakin ingin menutup aplikasi?"
            + ("\n\nKoneksi serial akan ditutup otomatis." if self.is_reading else "")):
            return
        self.is_reading = False
        if self.thread and self.thread.is_alive(): self.thread.join(timeout=2.5)
        self.thread = None
        if self.serial_conn:
            try:
                if self.serial_conn.is_open:
                    self.serial_conn.reset_input_buffer()
                    self.serial_conn.close()
            except Exception as e: print(f"[CLOSE] {e}")
        self.serial_conn = None
        time.sleep(1.5)
        self.root.destroy()


def main():
    root = tk.Tk()
    root.geometry("1280x800")
    root.minsize(1100, 700)
    ScaleApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()