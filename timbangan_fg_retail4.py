"""
Timbangan AND GX-4000 — Quality Control Data Logger
UI: Merah Terang & Putih, 3-column layout
Flow: NIK → Variant → Mesin → Simpan

Revisi v4:
  - Serial: retry 3x saat buka port, thread join sebelum close
  - Tombol Reconnect manual di header
  - Varian & Mesin mengikuti tabel resmi (dengan filter mesin per varian)
  - Nilai TU1, TU2, Min BPS, Standard, Max BPS sesuai tabel gambar
  - Tambah variant baru: Sachet YB 12,5gr PCS & RENCENG
  - Hapus fitur filler (radio button & kolom filler)
  - Data logger urutan descending (terbaru di atas)
  - Filter variant tetap dipertahankan
  - Shortcut SPACE untuk simpan (aktif hanya jika NIK + variant + mesin lengkap)

Requirements:
    pip install pyserial openpyxl requests
"""

import tkinter as tk
from tkinter import messagebox, filedialog
import serial
import serial.tools.list_ports
import threading
import queue
import time
import re
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ── PALETTE ────────────────────────────────────────────────────
RED          = "#E53E3E"
RED_DARK     = "#C53030"
RED_LIGHT    = "#FC8181"
RED_PALE     = "#FFF5F5"
RED_SOFT     = "#FED7D7"
WHITE        = "#FFFFFF"
GRAY_50      = "#FAFAFA"
GRAY_100     = "#F4F4F5"
GRAY_200     = "#E4E4E7"
GRAY_400     = "#A1A1AA"
GRAY_600     = "#52525B"
GRAY_800     = "#27272A"
GREEN        = "#16A34A"
GREEN_LIGHT  = "#DCFCE7"
RED_ALERT    = "#DC2626"
BG_ROOT      = "#F8ECEC"

# ── FONTS ──────────────────────────────────────────────────────
F_TITLE   = ("Segoe UI", 13, "bold")
F_SUB     = ("Segoe UI", 9)
F_LABEL   = ("Segoe UI", 8, "bold")
F_BODY    = ("Segoe UI", 10)
F_BODY_B  = ("Segoe UI", 10, "bold")
F_SMALL   = ("Segoe UI", 8)
F_CHIP    = ("Segoe UI", 9, "bold")
F_CHIP_S  = ("Segoe UI", 8)
F_MESIN   = ("Segoe UI", 12, "bold")
F_MESIN_S = ("Segoe UI", 8)
F_WEIGHT  = ("Courier New", 52, "bold")
F_UNIT    = ("Segoe UI", 13)
F_STATUS  = ("Segoe UI", 10, "bold")
F_NIK     = ("Courier New", 13, "bold")
F_BTN     = ("Segoe UI", 11, "bold")
F_BTN_SM  = ("Segoe UI", 9, "bold")
F_TABLE   = ("Segoe UI", 9)
F_TABLE_H = ("Segoe UI", 8, "bold")
F_MONO    = ("Courier New", 9)
F_BADGE   = ("Segoe UI", 8, "bold")

# ── SERIAL CONFIG ───────────────────────────────────────────────
BAUDRATE = 9600
DATABITS = 8
PARITY   = "N"
STOPBITS = 1

# ── API ─────────────────────────────────────────────────────────
API_URL = "http://10.11.10.130:8081/api/mesin"

# ── DATA VARIAN + STANDAR ───────────────────────────────────────
# Sumber: tabel gambar (TU2, TU1, Min BPS, Standard, Max BPS)
VARIANT_STANDARDS = {
    # ── Sachet ──────────────────────────────────────────────────
    "Sachet YB 12,5gr PCS":    {"min":  12.05, "std":  13.05, "max":  14.05, "tu1":  11.93, "tu2":  10.80, "code": "S12.5G-P"},
    "Sachet YB 12,5gr RENCENG":{"min": 154.60, "std": 156.60, "max": 168.60, "tu1": 143.10, "tu2": 129.60, "code": "S12.5G-R"},
    "Sachet YB 20gr PCS":      {"min":  19.14, "std":  20.64, "max":  21.64, "tu1":  18.84, "tu2":  17.04, "code": "S20G-P"},
    "Sachet YB 20gr RENCENG":  {"min": 244.68, "std": 247.68, "max": 259.68, "tu1": 226.08, "tu2": 204.48, "code": "S20G-R"},
    "Sachet BB 40gr PCS":      {"min":  39.10, "std":  41.10, "max":  42.10, "tu1":  37.50, "tu2":  33.90, "code": "S40G-P"},
    "Sachet BB 40gr RENCENG":  {"min": 489.20, "std": 493.20, "max": 505.20, "tu1": 450.00, "tu2": 406.80, "code": "S40G-R"},
    # ── Pouch ───────────────────────────────────────────────────
    "Pouch YB 77gr":           {"min":  78.70, "std":  79.20, "max":  82.70, "tu1":  74.70, "tu2":  70.20, "code": "P77G-YB"},
    "Pouch BB 77gr":           {"min":  78.70, "std":  79.20, "max":  82.70, "tu1":  74.70, "tu2":  70.20, "code": "P77G-BB"},
    "Pouch YB 250gr":          {"min": 253.00, "std": 255.00, "max": 257.00, "tu1": 246.00, "tu2": 237.00, "code": "P250G"},
    "Pouch BB 270gr":          {"min": 273.00, "std": 275.00, "max": 277.00, "tu1": 266.00, "tu2": 257.00, "code": "P270G"},
    "Pouch YB 550gr":          {"min": 556.00, "std": 561.00, "max": 566.00, "tu1": 545.80, "tu2": 530.80, "code": "P550G"},
    "Pouch YB 700gr":          {"min": 706.00, "std": 711.00, "max": 716.00, "tu1": 696.00, "tu2": 681.00, "code": "P700G"},
    "Pouch BB 725gr":          {"min": 730.00, "std": 735.00, "max": 740.00, "tu1": 720.00, "tu2": 705.00, "code": "P725G"},
    "Pouch YB 1000gr":         {"min":1007.50, "std":1012.50, "max":1017.50, "tu1": 997.50, "tu2": 982.50, "code": "P1000G"},
}

# ── RELASI VARIAN → MESIN (filter tombol mesin saat variant dipilih) ──
# Sesuai tabel gambar. Variant yang tidak ada → tampilkan SEMUA mesin.
VARIANT_MESIN = {
    "Sachet YB 12,5gr PCS":    {"Y", "Z"},
    "Sachet YB 12,5gr RENCENG":{"Y", "Z"},
    "Sachet YB 20gr PCS":      {"O", "P", "W", "X"},
    "Sachet YB 20gr RENCENG":  {"O", "P", "W", "X"},
    "Sachet BB 40gr PCS":      {"Q", "R"},
    "Sachet BB 40gr RENCENG":  {"Q", "R"},
    "Pouch YB 77gr":           {"F","G","H","I","D","E","J","K","C","L","AE","AG","B","AF","AI","AJ"},
    "Pouch BB 77gr":           {"C","L","AE","AG","B","AF","AI","AJ"},
    "Pouch YB 250gr":          {"AH"},
    "Pouch BB 270gr":          {"AH"},
    "Pouch YB 550gr":          {"A", "U", "V"},
    "Pouch YB 700gr":          {"A", "U", "V"},
    "Pouch BB 725gr":          {"A", "U", "V"},
    "Pouch YB 1000gr":         {"A", "U", "V"},
}

# ── SEMUA MESIN ─────────────────────────────────────────────────
MACHINES = [
    ("A","F2"),("AE","D12"),("AF","D13"),("AG","D14"),("AH","D15"),
    ("AI","D16"),("AJ","D17"),("AK","D18"),("B","D11/E5"),("C","D9"),
    ("D","D1"),("E","D2"),("F","D3"),("G","D4"),("H","D5"),("I","D6"),
    ("J","D7"),("K","D8"),("L","LD10"),("O","C1"),("P","C2"),("Q","A2"),
    ("R","C3"),("U","F3"),("V","F1"),("W","C7"),("X","C8"),("Y","B6"),("Z","B3"),
]
MACHINE_POS = {letter: pos for letter, pos in MACHINES}


# ════════════════════════════════════════════════════════════════
class ChipButton(tk.Frame):
    """Selectable chip/tag button."""
    def __init__(self, parent, text, subtext="", command=None,
                 normal_bg=WHITE, sel_bg=RED,
                 normal_fg=GRAY_800, sel_fg=WHITE,
                 normal_border=GRAY_200, sel_border=RED, **kwargs):
        super().__init__(parent, bg=parent["bg"], **kwargs)
        self._cmd = command
        self._sel = False
        self._nbg = normal_bg
        self._sbg = sel_bg
        self._row_widgets = []
        self._nfg = normal_fg
        self._sfg = sel_fg
        self._nb  = normal_border
        self._sb  = sel_border

        self._card = tk.Frame(self, bg=normal_bg,
                              highlightbackground=normal_border,
                              highlightthickness=1,
                              cursor="hand2")
        self._card.pack(fill="both", expand=True)

        self._lbl = tk.Label(self._card, text=text, font=F_CHIP,
                             bg=normal_bg, fg=normal_fg, anchor="w",
                             padx=7, pady=3)
        self._lbl.pack(side="top", fill="x")

        if subtext:
            self._sub = tk.Label(self._card, text=subtext, font=F_CHIP_S,
                                 bg=normal_bg, fg=GRAY_400, anchor="w",
                                 padx=7, pady=0)
            self._sub.pack(side="top", fill="x")
        else:
            self._sub = None

        for w in (self._card, self._lbl) + ((self._sub,) if self._sub else ()):
            w.bind("<Button-1>", self._click)
            w.bind("<Enter>",    self._hover)
            w.bind("<Leave>",    self._leave)

    def _click(self, e=None): self._cmd and self._cmd()

    def _hover(self, e=None):
        if not self._sel:
            self._card.config(bg=RED_PALE, highlightbackground=RED_LIGHT)
            self._lbl.config(bg=RED_PALE)
            if self._sub: self._sub.config(bg=RED_PALE)

    def _leave(self, e=None):
        if not self._sel:
            self._card.config(bg=self._nbg, highlightbackground=self._nb)
            self._lbl.config(bg=self._nbg)
            if self._sub: self._sub.config(bg=self._nbg)

    def select(self):
        self._sel = True
        self._card.config(bg=self._sbg, highlightbackground=self._sb)
        self._lbl.config(bg=self._sbg, fg=self._sfg)
        if self._sub: self._sub.config(bg=self._sbg, fg=WHITE)

    def deselect(self):
        self._sel = False
        self._card.config(bg=self._nbg, highlightbackground=self._nb)
        self._lbl.config(bg=self._nbg, fg=self._nfg)
        if self._sub: self._sub.config(bg=self._nbg, fg=GRAY_400)


# ════════════════════════════════════════════════════════════════
class MesinButton(tk.Frame):
    """Square card for machine selection."""
    def __init__(self, parent, letter, pos, command=None, **kwargs):
        super().__init__(parent, bg=parent["bg"], **kwargs)
        self._cmd = command
        self._sel = False
        self._disabled = False

        self._card = tk.Frame(self, bg=WHITE,
                              highlightbackground=GRAY_200,
                              highlightthickness=1,
                              cursor="hand2")
        self._card.pack(fill="both", expand=True)

        self._let = tk.Label(self._card, text=letter, font=F_MESIN,
                             bg=WHITE, fg=RED, pady=4)
        self._let.pack()
        self._pos = tk.Label(self._card, text=pos, font=F_MESIN_S,
                             bg=WHITE, fg=GRAY_400, pady=0)
        self._pos.pack()

        for w in (self._card, self._let, self._pos):
            w.bind("<Button-1>", self._click)
            w.bind("<Enter>",    self._hover)
            w.bind("<Leave>",    self._leave)

    def _click(self, e=None):
        if not self._disabled and self._cmd:
            self._cmd()

    def _hover(self, e=None):
        if not self._sel and not self._disabled:
            for w in (self._card, self._let, self._pos):
                w.config(bg=RED_PALE)
            self._card.config(highlightbackground=RED_LIGHT)

    def _leave(self, e=None):
        if not self._sel and not self._disabled:
            for w in (self._card, self._let, self._pos):
                w.config(bg=WHITE)
            self._card.config(highlightbackground=GRAY_200)

    def select(self):
        self._sel = True
        self._disabled = False
        self._card.config(cursor="hand2")
        for w in (self._card, self._let, self._pos):
            w.config(bg=RED)
        self._let.config(fg=WHITE)
        self._pos.config(fg=WHITE)
        self._card.config(highlightbackground=RED_DARK)

    def deselect(self):
        self._sel = False
        for w in (self._card, self._let, self._pos):
            w.config(bg=WHITE)
        self._let.config(fg=RED)
        self._pos.config(fg=GRAY_400)
        self._card.config(highlightbackground=GRAY_200)

    def disable(self):
        """Tampilkan sebagai mesin tidak tersedia untuk variant ini."""
        self._disabled = True
        self._sel = False
        self._card.config(cursor="arrow")
        for w in (self._card, self._let, self._pos):
            w.config(bg=GRAY_100)
        self._let.config(fg=GRAY_400)
        self._pos.config(fg=GRAY_200)
        self._card.config(highlightbackground=GRAY_200)

    def enable(self):
        self._disabled = False
        self._card.config(cursor="hand2")
        for w in (self._card, self._let, self._pos):
            w.config(bg=WHITE)
        self._let.config(fg=RED)
        self._pos.config(fg=GRAY_400)
        self._card.config(highlightbackground=GRAY_200)


# ════════════════════════════════════════════════════════════════
class ScaleApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Timbangan AND GX-4000 — Quality Control")
        self.root.configure(bg=BG_ROOT)
        self.root.resizable(True, True)

        # State
        self.serial_conn   = None
        self.thread        = None
        self.is_reading    = False
        self.data_queue    = queue.Queue()
        self.current_data  = None
        self.saved_data    = []
        self.sel_variant   = None
        self.sel_machine   = None
        self.nik_confirmed = False
        self.live_weight   = 0.0
        self.live_unit     = "g"
        self.auto_port     = None
        self.filtered_variant = None
        self._active_config   = "7E1"
        self._reconnecting    = False

        self.variant_btns  = {}
        self.mesin_btns    = {}

        self._build_ui()
        self._bind_space()
        self._auto_detect_port()
        self._check_queue()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── BUILD UI ────────────────────────────────────────────────
    def _build_ui(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(2, weight=1)
        self._build_header()
        self._build_stepbar()
        self._build_body()
        self._build_footer()

    # ── SPACE SHORTCUT ──────────────────────────────────────────
    def _bind_space(self):
        self.root.bind("<space>", self._space_save)

    def _space_save(self, event=None):
        # Jangan trigger jika fokus di Entry (user sedang ketik NIK)
        focused = self.root.focus_get()
        if isinstance(focused, tk.Entry):
            return
        # Hanya aktif jika NIK + variant + mesin sudah lengkap
        if self.nik_confirmed and self.sel_variant and self.sel_machine:
            self._save_data()

    # ── HEADER ──────────────────────────────────────────────────
    def _build_header(self):
        hdr = tk.Frame(self.root, bg=RED, height=60)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.columnconfigure(1, weight=1)
        hdr.grid_propagate(False)

        tk.Label(hdr, text="⚖", font=("Segoe UI", 18),
                 bg=RED, fg=WHITE, padx=16).grid(row=0, column=0,
                                                  sticky="ns", pady=10)

        title_f = tk.Frame(hdr, bg=RED)
        title_f.grid(row=0, column=1, sticky="w", pady=10)
        tk.Label(title_f, text="Timbangan AND GX-4000",
                 font=F_TITLE, bg=RED, fg=WHITE).pack(anchor="w")
        tk.Label(title_f, text="Quality Control · Data Logger",
                 font=F_SUB, bg=RED, fg=WHITE).pack(anchor="w")

        # Badge koneksi + tombol reconnect
        badge = tk.Frame(hdr, bg=RED,
                         highlightbackground=WHITE, highlightthickness=1)
        badge.grid(row=0, column=2, sticky="e", padx=(0, 8), pady=15)

        self.conn_dot = tk.Label(badge, text="●", font=("Segoe UI", 10),
                                 bg=RED, fg=GRAY_400)
        self.conn_dot.pack(side="left", padx=(10, 4))
        self.conn_lbl = tk.Label(badge, text="Mendeteksi port...",
                                 font=F_SMALL, bg=RED, fg=WHITE)
        self.conn_lbl.pack(side="left", padx=(0, 6))

        # Tombol Reconnect
        tk.Button(hdr, text="↺ Reconnect",
                  font=F_BTN_SM, bg=RED_DARK, fg=WHITE,
                  relief="flat", cursor="hand2",
                  activebackground=RED_DARK, activeforeground=WHITE,
                  padx=10, pady=4,
                  command=self._manual_reconnect).grid(
            row=0, column=3, sticky="e", padx=12, pady=15)

    # ── STEP BAR ────────────────────────────────────────────────
    def _build_stepbar(self):
        bar = tk.Frame(self.root, bg=RED_PALE,
                       highlightbackground=RED_SOFT, highlightthickness=1)
        bar.grid(row=1, column=0, sticky="ew")

        inner = tk.Frame(bar, bg=RED_PALE)
        inner.pack(side="left", padx=24, pady=10)

        steps = [("1", "NIK Operator"), ("2", "Variant & Mesin"), ("3", "Catat Timbangan")]
        self.step_nums = []
        self.step_lbls = []

        for i, (num, lbl) in enumerate(steps):
            if i > 0:
                tk.Label(inner, text="──", font=("Segoe UI", 9),
                         bg=RED_PALE, fg=RED_SOFT).pack(side="left", padx=4)
            sf = tk.Frame(inner, bg=RED_PALE)
            sf.pack(side="left")
            nf = tk.Label(sf, text=num, font=("Segoe UI", 9, "bold"),
                          bg=RED_SOFT, fg=RED, width=2, relief="flat")
            nf.pack(side="left", padx=(0, 6))
            ll = tk.Label(sf, text=lbl, font=("Segoe UI", 9, "bold"),
                          bg=RED_PALE, fg=GRAY_400)
            ll.pack(side="left")
            self.step_nums.append(nf)
            self.step_lbls.append(ll)

        self._set_step(0)

    def _set_step(self, active):
        for i, (nf, ll) in enumerate(zip(self.step_nums, self.step_lbls)):
            if i < active:
                nf.config(bg=GREEN, fg=WHITE, text="✓")
                ll.config(fg=GREEN)
            elif i == active:
                nf.config(bg=RED, fg=WHITE, text=str(i+1))
                ll.config(fg=RED)
            else:
                nf.config(bg=RED_SOFT, fg=RED, text=str(i+1))
                ll.config(fg=GRAY_400)

    # ── BODY ────────────────────────────────────────────────────
    def _build_body(self):
        body = tk.Frame(self.root, bg=WHITE)
        body.grid(row=2, column=0, sticky="nsew")
        body.columnconfigure(2, weight=1)
        body.rowconfigure(0, weight=1)
        self._build_col1(body)
        self._build_col2(body)
        self._build_col3(body)

    # ── COL 1 — NIK + Variant ───────────────────────────────────
    def _build_col1(self, parent):
        col = tk.Frame(parent, bg=WHITE,
                       highlightbackground=GRAY_100, highlightthickness=1)
        col.grid(row=0, column=0, sticky="nsew")
        col.configure(width=320)
        parent.columnconfigure(0, minsize=320)

        inner = tk.Frame(col, bg=WHITE)
        inner.pack(fill="both", expand=True, padx=18, pady=16)

        # NIK
        tk.Label(inner, text="NIK OPERATOR", font=F_LABEL,
                 bg=WHITE, fg=RED).pack(anchor="w", pady=(0, 6))

        nik_row = tk.Frame(inner, bg=WHITE)
        nik_row.pack(fill="x")

        self.nik_entry = tk.Entry(nik_row, font=F_NIK,
                                  bg=GRAY_50, fg=GRAY_800,
                                  insertbackground=RED,
                                  relief="flat",
                                  highlightbackground=GRAY_200,
                                  highlightthickness=1, width=12)
        self.nik_entry.pack(side="left", fill="y", ipady=6, padx=(0, 6))
        self.nik_entry.bind("<Return>", lambda e: self._confirm_nik())

        tk.Button(nik_row, text="Masuk", font=F_BTN_SM,
                  bg=RED, fg=WHITE, relief="flat",
                  activebackground=RED_DARK, activeforeground=WHITE,
                  cursor="hand2", padx=10,
                  command=self._confirm_nik).pack(side="left", fill="y")

        self.nik_ok_lbl = tk.Label(inner, text="", font=("Segoe UI", 9),
                                   bg=WHITE, fg=GREEN)
        self.nik_ok_lbl.pack(anchor="w", pady=(4, 0))

        tk.Frame(inner, bg=GRAY_100, height=1).pack(fill="x", pady=12)

        # Variant
        tk.Label(inner, text="VARIANT PRODUK", font=F_LABEL,
                 bg=WHITE, fg=RED).pack(anchor="w", pady=(0, 8))

        vf = tk.Frame(inner, bg=WHITE)
        vf.pack(fill="x")

        for i, (name, data) in enumerate(VARIANT_STANDARDS.items()):
            row     = i // 2
            col_idx = i % 2
            chip = ChipButton(
                vf, text=name, subtext=data["code"],
                command=lambda n=name: self._pick_variant(n)
            )
            chip.grid(row=row, column=col_idx, sticky="ew",
                      padx=(0 if col_idx == 0 else 3, 3 if col_idx == 0 else 0),
                      pady=2)
            self.variant_btns[name] = chip

        vf.columnconfigure(0, weight=1)
        vf.columnconfigure(1, weight=1)

    # ── COL 2 — Mesin ───────────────────────────────────────────
    def _build_col2(self, parent):
        col = tk.Frame(parent, bg=WHITE,
                       highlightbackground=GRAY_100, highlightthickness=1)
        col.grid(row=0, column=1, sticky="nsew")
        col.configure(width=300)
        parent.columnconfigure(1, minsize=300)

        inner = tk.Frame(col, bg=WHITE)
        inner.pack(fill="both", expand=True, padx=18, pady=16)

        tk.Label(inner, text="PILIH MESIN", font=F_LABEL,
                 bg=WHITE, fg=RED).pack(anchor="w", pady=(0, 4))

        # Info hint — tampil saat belum pilih variant
        self.mesin_hint_lbl = tk.Label(
            inner,
            text="Pilih variant terlebih dahulu untuk melihat mesin yang tersedia",
            font=F_SMALL, bg=WHITE, fg=GRAY_400, wraplength=260, justify="left")
        self.mesin_hint_lbl.pack(anchor="w", pady=(0, 8))

        gf = tk.Frame(inner, bg=WHITE)
        gf.pack(fill="x")

        COLS = 4
        for i, (letter, pos) in enumerate(MACHINES):
            row = i // COLS
            ci  = i % COLS
            btn = MesinButton(
                gf, letter=letter, pos=pos,
                command=lambda l=letter: self._pick_mesin(l)
            )
            btn.grid(row=row, column=ci, sticky="ew",
                     padx=2, pady=2, ipadx=0, ipady=2)
            self.mesin_btns[letter] = btn
            btn.disable()   # semua disabled sampai variant dipilih

        for c in range(COLS):
            gf.columnconfigure(c, weight=1)

    # ── COL 3 — Weight + Logger Table ───────────────────────────
    def _build_col3(self, parent):
        col = tk.Frame(parent, bg=WHITE)
        col.grid(row=0, column=2, sticky="nsew")

        inner = tk.Frame(col, bg=WHITE)
        inner.pack(fill="both", expand=True, padx=18, pady=16)
        inner.rowconfigure(5, weight=1)
        inner.columnconfigure(0, weight=1)

        # Weight card
        wcard = tk.Frame(inner, bg=RED_PALE,
                         highlightbackground=RED_SOFT, highlightthickness=1)
        wcard.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        tk.Label(wcard, text="BERAT TERBACA", font=F_LABEL,
                 bg=RED_PALE, fg=RED, pady=10).pack()

        self.weight_lbl = tk.Label(wcard, text="–––.––",
                                   font=F_WEIGHT, bg=RED_PALE, fg=RED)
        self.weight_lbl.pack()

        self.unit_lbl = tk.Label(wcard, text="gram",
                                 font=F_UNIT, bg=RED_PALE, fg=GRAY_400)
        self.unit_lbl.pack(pady=(0, 4))

        status_row = tk.Frame(wcard, bg=RED_PALE)
        status_row.pack(pady=(0, 12))

        self.status_dot = tk.Label(status_row, text="●",
                                   font=("Segoe UI", 10),
                                   bg=RED_PALE, fg=GRAY_400)
        self.status_dot.pack(side="left", padx=(0, 4))
        self.status_txt = tk.Label(status_row, text="Pilih variant untuk validasi",
                                   font=F_STATUS, bg=RED_PALE, fg=GRAY_600)
        self.status_txt.pack(side="left")

        # STD pills (min, std, max + TU1, TU2)
        std_row = tk.Frame(inner, bg=WHITE)
        std_row.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        for c in range(5):
            std_row.columnconfigure(c, weight=1)

        self.std_tu2 = self._std_pill(std_row, "TU2",     "–", GRAY_400,   0)
        self.std_tu1 = self._std_pill(std_row, "TU1",     "–", GRAY_600,   1)
        self.std_min = self._std_pill(std_row, "Min",     "–", RED_ALERT,  2)
        self.std_std = self._std_pill(std_row, "Standar", "–", GRAY_800,   3)
        self.std_max = self._std_pill(std_row, "Max",     "–", RED,        4)

        # Validation message
        self.valid_lbl = tk.Label(inner, text="", font=("Segoe UI", 9),
                                  bg=WHITE, fg=GRAY_400)
        self.valid_lbl.grid(row=2, column=0, sticky="w", pady=(0, 6))

        # Save button + SPACE hint
        btn_row = tk.Frame(inner, bg=WHITE)
        btn_row.grid(row=3, column=0, sticky="ew", pady=(0, 14))
        btn_row.columnconfigure(0, weight=1)

        self.save_btn = tk.Button(
            btn_row, text="Simpan Data Timbangan",
            font=F_BTN, bg=GRAY_200, fg=GRAY_400,
            relief="flat", cursor="arrow",
            activebackground=RED_DARK, activeforeground=WHITE,
            command=self._save_data, state="disabled",
            pady=10
        )
        self.save_btn.grid(row=0, column=0, sticky="ew")

        self.space_hint = tk.Label(btn_row, text="",
                                   font=F_SMALL, bg=WHITE, fg=GRAY_400)
        self.space_hint.grid(row=1, column=0, sticky="e", pady=(2, 0))

        # Table header
        tbl_hdr = tk.Frame(inner, bg=WHITE)
        tbl_hdr.grid(row=4, column=0, sticky="ew", pady=(0, 6))
        tbl_hdr.columnconfigure(0, weight=1)

        tk.Label(tbl_hdr, text="DATA TERSIMPAN", font=F_LABEL,
                 bg=WHITE, fg=RED).grid(row=0, column=0, sticky="w")

        right_info = tk.Frame(tbl_hdr, bg=WHITE)
        right_info.grid(row=0, column=1, sticky="e")

        self.filter_lbl = tk.Label(right_info, text="",
                                   font=F_SMALL, bg=WHITE, fg=RED)
        self.filter_lbl.pack(side="left", padx=(0, 6))

        self.count_lbl = tk.Label(right_info, text="0 data",
                                  font=F_SMALL, bg=WHITE, fg=GRAY_400)
        self.count_lbl.pack(side="left")

        self.reset_filter_btn = tk.Button(
            right_info, text="Tampilkan Semua",
            font=F_SMALL, bg=WHITE, fg=GRAY_600,
            relief="flat",
            highlightbackground=GRAY_200, highlightthickness=1,
            cursor="hand2", padx=6, pady=1,
            command=self._reset_filter
        )
        self.reset_filter_btn.pack(side="left", padx=(6, 0))
        self.reset_filter_btn.pack_forget()

        self._build_table(inner, row=5)

    def _std_pill(self, parent, label, val, color, col):
        f = tk.Frame(parent, bg=WHITE,
                     highlightbackground=GRAY_200, highlightthickness=1)
        f.grid(row=0, column=col, sticky="ew",
               padx=(0, 3) if col < 4 else 0)
        tk.Label(f, text=label, font=F_SMALL, bg=WHITE,
                 fg=GRAY_400, pady=4).pack()
        lbl = tk.Label(f, text=val, font=F_BODY_B, bg=WHITE,
                       fg=color, pady=4)
        lbl.pack()
        return lbl

    def _build_table(self, parent, row):
        tbl_frame = tk.Frame(parent, bg=WHITE)
        tbl_frame.grid(row=row, column=0, sticky="nsew")
        parent.rowconfigure(row, weight=1)
        tbl_frame.columnconfigure(0, weight=1)
        tbl_frame.rowconfigure(1, weight=1)

        # Kolom tabel — tanpa Filler
        cols   = ["#", "Waktu", "Variant", "Mesin", "Berat", "Status"]
        widths = [30,   110,     180,       60,      90,      70]

        hdr = tk.Frame(tbl_frame, bg=GRAY_50)
        hdr.grid(row=0, column=0, sticky="ew")
        for c, w in zip(cols, widths):
            tk.Label(hdr, text=c, font=F_TABLE_H, bg=GRAY_50,
                     fg=GRAY_400, width=w//8, anchor="w",
                     padx=6, pady=5).pack(side="left")

        canvas = tk.Canvas(tbl_frame, bg=WHITE, highlightthickness=0)
        sb = tk.Scrollbar(tbl_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        canvas.grid(row=1, column=0, sticky="nsew")
        sb.grid(row=1, column=1, sticky="ns")

        self.table_inner  = tk.Frame(canvas, bg=WHITE)
        self.table_window = canvas.create_window(
            0, 0, window=self.table_inner, anchor="nw")
        self._table_canvas = canvas

        def _on_frame_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_configure(e):
            canvas.itemconfig(self.table_window, width=e.width)

        self.table_inner.bind("<Configure>", _on_frame_configure)
        canvas.bind("<Configure>", _on_canvas_configure)
        self._table_row_count = 0

    def _make_row_widget(self, parent, values, ok=True, row_index=0):
        """Buat satu frame row tabel dan return-kan (tidak di-pack di sini)."""
        bg     = WHITE if row_index % 2 == 0 else GRAY_50
        widths = [30, 110, 180, 60, 90, 70]

        row_f = tk.Frame(parent, bg=bg)
        tk.Frame(row_f, bg=GRAY_100, height=1).pack(fill="x")
        cells = tk.Frame(row_f, bg=bg)
        cells.pack(fill="x")

        for i, (val, w) in enumerate(zip(values, widths)):
            if i == 5:  # Status badge
                badge_bg = GREEN_LIGHT if ok else "#FEE2E2"
                badge_fg = GREEN       if ok else RED_ALERT
                tk.Label(cells, text=val, font=F_BADGE,
                         bg=badge_bg, fg=badge_fg,
                         padx=6, pady=1).pack(side="left", padx=6, pady=4)
            else:
                font = F_MONO if i in (0, 1, 4) else F_TABLE
                tk.Label(cells, text=val, font=font,
                         bg=bg, fg=GRAY_600,
                         width=w//8, anchor="w", padx=6).pack(side="left")
        return row_f

    def _insert_row_top(self, record):
        if self.filtered_variant and record["variant"] != self.filtered_variant:
            self._update_count_label()
            return

        for idx, row_f in enumerate(self._row_widgets):
            new_bg = GRAY_50 if idx % 2 == 0 else WHITE
            row_f.config(bg=new_bg)
            children = row_f.winfo_children()
            if len(children) >= 2:
                cells_frame = children[1]
                cells_frame.config(bg=new_bg)
                for widget in cells_frame.winfo_children():
                    try:
                        curr_bg = widget.cget("bg")
                        if curr_bg not in (GREEN_LIGHT, "#FEE2E2"):
                            widget.config(bg=new_bg)
                    except Exception:
                        pass

        ok        = record["status"] == "OK"
        t_short   = record["timestamp"][11:]
        var_short = record["variant"][:18] + ("…" if len(record["variant"]) > 18 else "")
        row_num   = len(self.saved_data)

        values  = [str(row_num), t_short, var_short, record["machine"],
                f"{record['weight']:.2f}g", record["status"]]
        new_row = self._make_row_widget(self.table_inner, values, ok=ok, row_index=0)

        if self._row_widgets:
            new_row.pack(fill="x", before=self._row_widgets[0])
        else:
            new_row.pack(fill="x")

        self._row_widgets.insert(0, new_row)
        self._table_row_count += 1
        self._table_canvas.yview_moveto(0)
        self._update_count_label()

    def _update_count_label(self):
        if self.filtered_variant:
            shown = sum(1 for d in self.saved_data
                        if d["variant"] == self.filtered_variant)
            total = len(self.saved_data)
            short_name = (self.filtered_variant[:18] + "…"
                          if len(self.filtered_variant) > 18
                          else self.filtered_variant)
            self.filter_lbl.config(text=f"▶ {short_name}", fg=RED)
            self.reset_filter_btn.pack(side="left", padx=(6, 0))
            if shown == total:
                self.count_lbl.config(text=f"{total} data", fg=GRAY_400)
            else:
                self.count_lbl.config(text=f"{shown} dari {total} data", fg=RED)
        else:
            total = len(self.saved_data)
            self.filter_lbl.config(text="", fg=GRAY_400)
            self.reset_filter_btn.pack_forget()
            self.count_lbl.config(
                text=f"{total} data" if total else "0 data", fg=GRAY_400)

    def _rebuild_table(self):
        for widget in self.table_inner.winfo_children():
            widget.destroy()
        self._table_row_count = 0
        self._row_widgets = []   # ← reset

        if self.filtered_variant:
            data_to_show = [d for d in self.saved_data
                            if d["variant"] == self.filtered_variant]
        else:
            data_to_show = self.saved_data

        for i, d in enumerate(reversed(data_to_show), 1):
            ok        = d["status"] == "OK"
            t_short   = d["timestamp"][11:]
            var_short = d["variant"][:18] + ("…" if len(d["variant"]) > 18 else "")
            row_f = self._make_row_widget(
                self.table_inner,
                [str(i), t_short, var_short, d["machine"],
                f"{d['weight']:.2f}g", d["status"]],
                ok=ok, row_index=i - 1
            )
            row_f.pack(fill="x")
            self._row_widgets.append(row_f)   # ← track
            self._table_row_count += 1

        self.root.after(30, lambda: self._table_canvas.yview_moveto(0))
        self._update_count_label()

    def _refresh_table(self):
        """Alias rebuild — dipakai saat filter/variant berubah."""
        self._rebuild_table()

    def _reset_filter(self):
        self.filtered_variant = None
        self._refresh_table()

    # ── FOOTER ──────────────────────────────────────────────────
    def _build_footer(self):
        ft = tk.Frame(self.root, bg=GRAY_50,
                      highlightbackground=GRAY_100, highlightthickness=1)
        ft.grid(row=3, column=0, sticky="ew")

        self.footer_lbl = tk.Label(
            ft, text=f"Port: Mendeteksi... · API: {API_URL}",
            font=F_SMALL, bg=GRAY_50, fg=GRAY_400)
        self.footer_lbl.pack(side="left", padx=16, pady=6)

        tk.Button(ft, text="↓ Export Excel",
                  font=F_BTN_SM, bg=WHITE, fg=GRAY_600,
                  relief="flat",
                  highlightbackground=GRAY_200, highlightthickness=1,
                  cursor="hand2", padx=10, pady=4,
                  command=self._export_excel).pack(
            side="right", padx=12, pady=6)

        # SPACE shortcut info
        tk.Label(ft, text="[SPACE] = Simpan cepat",
                 font=F_SMALL, bg=GRAY_50, fg=GRAY_200).pack(
            side="right", padx=0, pady=6)

    # ── AUTO PORT DETECT ────────────────────────────────────────
    def _auto_detect_port(self):
        def _detect():
            ports = serial.tools.list_ports.comports()
            if not ports:
                self.root.after(0, self._set_conn_status, False,
                                "Tidak ada port ditemukan")
                self.root.after(3000, self._auto_detect_port)
                return

            serial_configs = [
                {"bytesize": DATABITS, "parity": serial.PARITY_EVEN,  "label": "7E1"},
                {"bytesize": 8,        "parity": serial.PARITY_NONE,  "label": "8N1"},
            ]

            for p in ports:
                for cfg in serial_configs:
                    conn = self._try_open_serial(p.device, cfg)
                    if conn:
                        self.serial_conn    = conn
                        self.auto_port      = p.device
                        self.is_reading     = True
                        self._active_config = cfg["label"]

                        self.root.after(0, self._set_conn_status, True,
                                        f"{p.device} [{cfg['label']}]")
                        self.thread = threading.Thread(
                            target=self._read_thread, daemon=True)
                        self.thread.start()
                        return

            self.root.after(0, self._set_conn_status, False,
                            f"{len(ports)} port ditemukan, gagal konek")
            self.root.after(5000, self._auto_detect_port)

        threading.Thread(target=_detect, daemon=True).start()

    def _try_open_serial(self, port, cfg, retries=3, delay=1.0):
        """Coba buka port dengan retry — menghindari PermissionError(13)."""
        for attempt in range(retries):
            try:
                conn = serial.Serial(
                    port     = port,
                    baudrate = BAUDRATE,
                    bytesize = cfg["bytesize"],
                    parity   = cfg["parity"],
                    stopbits = STOPBITS,
                    timeout  = 1
                )
                conn.reset_input_buffer()
                return conn
            except serial.SerialException as e:
                print(f"[SERIAL] {port} ({cfg['label']}) attempt {attempt+1}/{retries}: {e}")
                if attempt < retries - 1:
                    time.sleep(delay)
        return None

    def _manual_reconnect(self):
        """Tutup koneksi lama → lepas port → deteksi ulang."""
        if self._reconnecting:
            return
        self._reconnecting = True
        self._set_conn_status(False, "Reconnecting...")

        def _do():
            # Stop thread dulu
            self.is_reading = False
            if self.thread and self.thread.is_alive():
                self.thread.join(timeout=2.5)
            self.thread = None

            # Close port
            if self.serial_conn:
                try:
                    if self.serial_conn.is_open:
                        self.serial_conn.reset_input_buffer()
                        self.serial_conn.close()
                except Exception as e:
                    print(f"[SERIAL] close error: {e}")
            self.serial_conn = None

            # Tunggu Windows lepas handle
            time.sleep(0.8)

            self._reconnecting = False
            self.root.after(0, self._auto_detect_port)

        threading.Thread(target=_do, daemon=True).start()

    def _set_conn_status(self, connected, detail):
        if connected:
            self.conn_dot.config(fg="#4ADE80")
            self.conn_lbl.config(text=f"{detail} · Terhubung")
            self.footer_lbl.config(
                text=f"Port: {detail} (auto) · 9600bps · API: {API_URL}")
        else:
            self.conn_dot.config(fg=GRAY_400)
            self.conn_lbl.config(text=detail)
            self.footer_lbl.config(
                text=f"Port: {detail} · API: {API_URL}")

    # ── SERIAL READ ─────────────────────────────────────────────
    def _read_thread(self):
        while self.is_reading:
            try:
                if self.serial_conn and self.serial_conn.in_waiting > 0:
                    raw = self.serial_conn.readline().decode(
                        "ascii", errors="ignore").strip()
                    if raw:
                        parsed = self._parse(raw)
                        if parsed:
                            self.data_queue.put(parsed)
            except Exception:
                break
            time.sleep(0.01)

    def _parse(self, raw):
        try:
            m = re.match(r"^([A-Z]{2}),([+\-]?\d+\.?\d*)\s*([a-zA-Z]+)$", raw)
            if m:
                return {
                    "timestamp":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "scale_status": m.group(1),
                    "weight":       float(m.group(2)),
                    "unit":         m.group(3),
                    "raw":          raw,
                }
        except Exception:
            pass
        return None

    def _check_queue(self):
        try:
            while True:
                data = self.data_queue.get_nowait()
                self._update_weight(data)
        except queue.Empty:
            pass
        self.root.after(50, self._check_queue)

    def _update_weight(self, data):
        self.current_data = data
        self.live_weight  = data["weight"]
        self.live_unit    = data["unit"]
        self.weight_lbl.config(text=f"{data['weight']:.2f}")
        unit_map = {"g": "gram", "kg": "kilogram"}
        self.unit_lbl.config(text=unit_map.get(data["unit"], data["unit"]))
        self._validate_display(data["weight"])

    def _validate_display(self, weight):
        if not self.sel_variant:
            self.weight_lbl.config(fg=RED)
            self.status_dot.config(fg=GRAY_400)
            self.status_txt.config(text="Pilih variant untuk validasi", fg=GRAY_600)
            self.valid_lbl.config(text="")
            return

        std = VARIANT_STANDARDS[self.sel_variant]
        tu2 = std.get("tu2", None)
        tu1 = std.get("tu1", None)

        if weight < (tu2 if tu2 else std["min"]):
            color = RED_ALERT
            label = "NOT OK — Berat Jauh di Bawah (< TU2)"
        elif tu2 and tu1 and weight < tu1:
            color = "#D97706"   # amber
            label = "WASPADA — TU2–TU1"
        elif tu1 and weight < std["min"]:
            color = "#D97706"
            label = "WASPADA — TU1–Min Standar"
        elif weight > std["max"]:
            color = RED_ALERT
            label = "NOT OK — Berat Melebihi Maks"
        else:
            color = GREEN
            label = "OK — Sesuai Standar"

        self.weight_lbl.config(fg=color)
        self.status_dot.config(fg=color)
        self.status_txt.config(text=label, fg=color)

        detail = f"TU2: {tu2}g  TU1: {tu1}g  Range OK: {std['min']}–{std['max']}g"
        self.valid_lbl.config(text=detail, fg=GRAY_400)

    # ── NIK ─────────────────────────────────────────────────────
    def _confirm_nik(self):
        nik = self.nik_entry.get().strip()
        if not nik:
            return
        self.nik_confirmed = True
        self.nik_ok_lbl.config(
            text=f"✓  NIK {nik} — Operator Terverifikasi", fg=GREEN)
        self._set_step(1)
        self.root.focus_set()

    # ── VARIANT PICK ────────────────────────────────────────────
    def _pick_variant(self, name):
        if not self.nik_confirmed:
            messagebox.showwarning("NIK Belum Dikonfirmasi",
                                   "Masukkan NIK operator terlebih dahulu.")
            return

        for n, btn in self.variant_btns.items():
            btn.deselect() if n != name else btn.select()

        self.sel_variant      = name
        self.filtered_variant = name
        self.sel_machine      = None   # reset mesin saat ganti variant

        std = VARIANT_STANDARDS[name]
        self.std_tu2.config(text=f"{std.get('tu2', '–')}")
        self.std_tu1.config(text=f"{std.get('tu1', '–')}")
        self.std_min.config(text=f"{std['min']:.2f}")
        self.std_std.config(text=f"{std['std']:.2f}")
        self.std_max.config(text=f"{std['max']:.2f}")

        # Filter mesin: enable hanya mesin yang sesuai variant
        allowed = VARIANT_MESIN.get(name, None)
        for letter, btn in self.mesin_btns.items():
            btn.deselect()
            if allowed is None or letter in allowed:
                btn.enable()
            else:
                btn.disable()

        # Update hint label
        if allowed:
            self.mesin_hint_lbl.config(
                text=f"{len(allowed)} mesin tersedia untuk variant ini",
                fg=GRAY_400)
        else:
            self.mesin_hint_lbl.config(
                text="Semua mesin tersedia untuk variant ini",
                fg=GRAY_400)

        if self.live_weight:
            self._validate_display(self.live_weight)

        self._check_save_ready()
        self._refresh_table()

        if self.sel_variant and self.sel_machine:
            self._set_step(2)

    # ── MESIN PICK ───────────────────────────────────────────────
    def _pick_mesin(self, letter):
        if not self.nik_confirmed:
            messagebox.showwarning("NIK Belum Dikonfirmasi",
                                   "Masukkan NIK operator terlebih dahulu.")
            return

        for l, btn in self.mesin_btns.items():
            if l != letter:
                # Jangan reset mesin yang di-disabled
                if not self.mesin_btns[l]._disabled:
                    btn.deselect()
            else:
                btn.select()

        self.sel_machine = letter
        self._check_save_ready()

        if self.sel_variant and self.sel_machine:
            self._set_step(2)

    def _check_save_ready(self):
        ready = self.sel_variant and self.sel_machine and self.nik_confirmed
        if ready:
            self.save_btn.config(state="normal", bg=RED, fg=WHITE,
                                 cursor="hand2")
            self.space_hint.config(text="Tekan [SPACE] untuk simpan cepat", fg=GRAY_400)
            self._set_step(2)
        else:
            self.save_btn.config(state="disabled", bg=GRAY_200, fg=GRAY_400,
                                 cursor="arrow")
            self.space_hint.config(text="")

    # ── SAVE DATA ───────────────────────────────────────────────
    def _save_data(self):
        if not self.current_data:
            messagebox.showwarning(
                "Tidak Ada Data",
                "Timbangan belum terbaca. Pastikan kabel terhubung.")
            return

        if not (self.sel_variant and self.sel_machine and self.nik_confirmed):
            messagebox.showerror("Lengkapi Data",
                                 "NIK, Variant, dan Mesin harus dipilih.")
            return

        w   = self.current_data["weight"]
        std = VARIANT_STANDARDS[self.sel_variant]
        ok  = std["min"] <= w <= std["max"]
        status = "OK" if ok else "NOT OK"
        nik    = self.nik_entry.get().strip()

        form = {
            "nik":     nik,
            "mesin":   self.sel_machine,
            "variant": self.sel_variant,
            "waktu":   self.current_data["timestamp"],
            "berat":   str(w),
            "unit":    self.current_data["unit"],
            "status":  status,
        }

        def _post():
            try:
                resp = requests.post(API_URL, data=form, timeout=8)
                api_status = ("✓ Terkirim"
                              if resp.status_code in (200, 201)
                              else f"Error {resp.status_code}")
            except requests.exceptions.ConnectionError:
                api_status = "Offline"
            except Exception:
                api_status = "Gagal"

            for d in reversed(self.saved_data):
                if d.get("api_status") == "Pending":
                    d["api_status"] = api_status
                    break

        record = {
            **self.current_data,
            "nik":        nik,
            "machine":    self.sel_machine,
            "variant":    self.sel_variant,
            "status":     status,
            "api_status": "Pending",
        }
        self.saved_data.append(record)
        threading.Thread(target=_post, daemon=True).start()

        self._insert_row_top(record)
        self._set_step(2)

        # Flash hijau sebentar
        orig = self.weight_lbl.cget("fg")
        self.weight_lbl.config(fg=GREEN)
        self.root.after(300, lambda: self.weight_lbl.config(fg=orig))

    # ── EXPORT EXCEL ────────────────────────────────────────────
    def _export_excel(self):
        if not self.saved_data:
            messagebox.showinfo("Kosong", "Belum ada data untuk di-export.")
            return

        fn = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"timbangan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not fn:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Timbangan"

            thin   = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            ws.merge_cells("A1:J1")
            c = ws["A1"]
            c.value = "DATA TIMBANGAN — AND GX-4000 | Quality Control"
            c.font  = Font(name="Calibri", size=13, bold=True, color="E53E3E")
            c.fill  = PatternFill("solid", fgColor="FFF5F5")
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            # Header tanpa kolom Filler
            headers = ["No", "NIK", "Mesin", "Variant",
                       "Tanggal", "Waktu", "Berat", "Unit", "Status", "API"]
            hfill = PatternFill("solid", fgColor="E53E3E")
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(2, ci, h)
                cell.font      = Font(name="Calibri", size=10,
                                      bold=True, color="FFFFFF")
                cell.fill      = hfill
                cell.alignment = Alignment(horizontal="center")
                cell.border    = border
            ws.row_dimensions[2].height = 20

            fill_a   = PatternFill("solid", fgColor="FFFFFF")
            fill_b   = PatternFill("solid", fgColor="FFF5F5")
            fill_ok  = PatternFill("solid", fgColor="DCFCE7")
            fill_nok = PatternFill("solid", fgColor="FEE2E2")

            count_ok = count_nok = 0
            for i, d in enumerate(self.saved_data, 1):
                row   = i + 2
                fill  = fill_a if i % 2 else fill_b
                dt    = datetime.strptime(d["timestamp"], "%Y-%m-%d %H:%M:%S")
                is_ok = d["status"] == "OK"
                if is_ok: count_ok  += 1
                else:     count_nok += 1

                vals = [
                    i,
                    d.get("nik", ""),
                    d["machine"],
                    d["variant"],
                    dt.strftime("%Y-%m-%d"),
                    dt.strftime("%H:%M:%S"),
                    d["weight"],
                    d["unit"],
                    d["status"],
                    d.get("api_status", "–"),
                ]
                for ci, val in enumerate(vals, 1):
                    cell = ws.cell(row, ci, val)
                    cell.font      = Font(name="Calibri", size=10)
                    cell.fill      = (fill_ok  if (ci == 9 and is_ok)  else
                                      fill_nok if (ci == 9 and not is_ok) else fill)
                    cell.border    = border
                    cell.alignment = Alignment(horizontal="center")

            for ci, w in enumerate([6, 12, 10, 26, 14, 10, 10, 8, 10, 12], 1):
                ws.column_dimensions[
                    ws.cell(1, ci).column_letter].width = w

            sr = len(self.saved_data) + 4
            ws.cell(sr,   1).value = "RINGKASAN"
            ws.cell(sr,   1).font  = Font(bold=True, color="E53E3E")
            ws.cell(sr+1, 1, f"Total  : {len(self.saved_data)}")
            ws.cell(sr+2, 1, f"OK     : {count_ok}")
            ws.cell(sr+3, 1, f"NOT OK : {count_nok}")

            wb.save(fn)
            messagebox.showinfo(
                "Export Berhasil",
                f"Data berhasil disimpan!\n\n"
                f"File  : {fn}\n"
                f"Total : {len(self.saved_data)} record\n"
                f"OK    : {count_ok}  |  NOT OK: {count_nok}"
            )
        except Exception as e:
            messagebox.showerror("Export Gagal", str(e))

    # ── CLOSE ───────────────────────────────────────────────────
    def _on_close(self):
        if not messagebox.askokcancel(
            "Tutup Aplikasi",
            "Apakah Anda yakin ingin menutup aplikasi?"
            + ("\n\nKoneksi serial akan ditutup otomatis."
               if self.is_reading else "")
        ):
            return

        # Stop thread DULU sebelum close port
        self.is_reading = False
        if self.thread and self.thread.is_alive():
            self.thread.join(timeout=2.5)
        self.thread = None

        # Baru close port
        if self.serial_conn:
            try:
                if self.serial_conn.is_open:
                    self.serial_conn.reset_input_buffer()
                    self.serial_conn.close()
            except Exception as e:
                print(f"[CLOSE] serial close error: {e}")
        self.serial_conn = None

        time.sleep(1.5)
        self.root.destroy()


# ════════════════════════════════════════════════════════════════
def main():
    root = tk.Tk()
    root.geometry("1200x780")
    root.minsize(1050, 700)
    app = ScaleApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()