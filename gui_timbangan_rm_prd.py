"""
GUI Timbangan AD-4329A — A&D Weighing Indicator
================================================
RS-232: 2400 bps | 7 bit | None | 1 stop
Format: *mT*+0000000kg

- Display berat realtime
- Klik SIMPAN untuk simpan data saat itu
- Export ke Excel
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import serial
import serial.tools.list_ports
import threading
import time
import re
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ── RS-232 CONFIG ──────────────────────────────────────────────
DEFAULT_BAUD     = 2400
DEFAULT_BYTESIZE = serial.SEVENBITS
DEFAULT_PARITY   = serial.PARITY_NONE
DEFAULT_STOPBITS = serial.STOPBITS_ONE
DEFAULT_TIMEOUT  = 2

# ── PALET WARNA — Refined Dark Industrial ──────────────────────
BG          = "#111318"      # latar utama — hampir hitam
SURFACE     = "#1c1f2b"      # kartu / panel
SURFACE2    = "#222636"      # panel lebih terang
BORDER      = "#2e3347"      # garis pemisah halus
ACCENT      = "#3ecfb2"      # teal mint — satu aksen utama
ACCENT_DIM  = "#2a8f7c"      # aksen lebih gelap (hover)
DANGER      = "#e05c6e"      # merah soft (stop/hapus)
DANGER_DIM  = "#a5394a"
SUCCESS     = "#3dba7e"      # hijau soft (simpan)
SUCCESS_DIM = "#2a8a5a"
INFO        = "#4a90d9"      # biru info (export)
INFO_DIM    = "#2e6bad"
TXT         = "#dce1ef"      # teks utama
TXT_MED     = "#8b94b0"      # teks sekunder
TXT_DIM     = "#4e5775"      # teks sangat redup
ROW_A       = "#1c1f2b"
ROW_B       = "#191c28"

# ── FONT ──────────────────────────────────────────────────────
F_UI        = ("Segoe UI", 9)
F_UI_B      = ("Segoe UI", 9, "bold")
F_UI_SM     = ("Segoe UI", 8)
F_DISPLAY   = ("Courier New", 58, "bold")
F_UNIT      = ("Courier New", 20, "bold")
F_LABEL     = ("Segoe UI", 10)
F_BTN       = ("Segoe UI", 10, "bold")
F_BTN_LG    = ("Segoe UI", 13, "bold")


def parse_berat(raw: str):
    raw = raw.strip()
    m = re.search(r'\*m[A-Z]\*([+\-])\s*(\d+[\.,]?\d*)\s*([a-zA-Z]*)', raw)
    if m:
        sign   = -1 if m.group(1) == '-' else 1
        nilai  = float(m.group(2).replace(',', '.'))
        satuan = m.group(3) or 'kg'
        return sign * nilai, satuan
    m2 = re.search(r'([+\-])\s*(\d+[\.,]?\d*)\s*([a-zA-Z]+)', raw)
    if m2:
        return (-1 if m2.group(1) == '-' else 1) * float(m2.group(2)), m2.group(3)
    return None, None


class TimbanganApp:
    def __init__(self, root):
        self.root = root
        self.root.title("AD-4329A  Weighing Indicator")
        self.root.geometry("620x680")
        self.root.minsize(540, 580)
        self.root.resizable(True, True)
        self.root.configure(bg=BG)

        self.ser         = None
        self.thread      = None
        self.running     = False
        self.berat_live  = None
        self.satuan_live = "kg"
        self.data        = []

        self.berat_var  = tk.StringVar(value="––––.–––")
        self.satuan_var = tk.StringVar(value="kg")
        self.status_var = tk.StringVar(value="Tidak Terhubung")
        self.port_var   = tk.StringVar()
        self.count_var  = tk.StringVar(value="0 data tersimpan")

        self._apply_styles()
        self._build()
        self._refresh_ports()

    # ─────────────────────────────────────────────────────────
    #  STYLES
    # ─────────────────────────────────────────────────────────

    def _apply_styles(self):
        style = ttk.Style()
        style.theme_use("clam")

        # Treeview
        style.configure("Treeview",
                        background=ROW_A,
                        foreground=TXT,
                        fieldbackground=ROW_A,
                        rowheight=26,
                        font=("Consolas", 10),
                        borderwidth=0,
                        relief="flat")
        style.configure("Treeview.Heading",
                        background=SURFACE2,
                        foreground=TXT_MED,
                        font=("Segoe UI", 8, "bold"),
                        relief="flat",
                        borderwidth=0)
        style.map("Treeview",
                  background=[("selected", ACCENT_DIM)],
                  foreground=[("selected", TXT)])
        style.map("Treeview.Heading",
                  background=[("active", SURFACE2)],
                  relief=[("active", "flat")])

        # Scrollbar
        style.configure("Vertical.TScrollbar",
                        background=SURFACE2,
                        troughcolor=SURFACE,
                        bordercolor=BORDER,
                        arrowcolor=TXT_DIM,
                        relief="flat",
                        width=8)
        style.map("Vertical.TScrollbar",
                  background=[("active", BORDER)])

        # Combobox — sepenuhnya custom via Option menu
        style.configure("Elegant.TCombobox",
                        fieldbackground=SURFACE2,
                        background=SURFACE2,
                        foreground=TXT,
                        arrowcolor=ACCENT,
                        bordercolor=BORDER,
                        lightcolor=BORDER,
                        darkcolor=BORDER,
                        insertcolor=TXT,
                        padding=(8, 6))
        style.map("Elegant.TCombobox",
                  fieldbackground=[("readonly", SURFACE2),
                                   ("disabled", SURFACE)],
                  foreground=[("readonly", TXT),
                               ("disabled", TXT_DIM)],
                  bordercolor=[("focus", ACCENT),
                                ("!focus", BORDER)],
                  arrowcolor=[("disabled", TXT_DIM),
                               ("!disabled", ACCENT)])

    # ─────────────────────────────────────────────────────────
    #  BUILD UI
    # ─────────────────────────────────────────────────────────

    def _build(self):
        # ── ROOT GRID ──
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(3, weight=1)   # tabel bisa expand

        # ─────────────────────────────────────
        # 0) HEADER
        # ─────────────────────────────────────
        hdr = tk.Frame(self.root, bg=SURFACE, height=52)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.columnconfigure(1, weight=1)
        hdr.grid_propagate(False)

        # Garis aksen kiri
        tk.Frame(hdr, bg=ACCENT, width=4).grid(
            row=0, column=0, sticky="ns", padx=(0, 14))

        tk.Label(hdr,
                 text="⚖  A&D Weighing  —  AD-4329A",
                 font=("Segoe UI", 11, "bold"),
                 bg=SURFACE, fg=TXT).grid(row=0, column=1, sticky="w")

        tk.Label(hdr,
                 text="RS-232  ·  2400 bps  ·  7-N-1",
                 font=F_UI_SM,
                 bg=SURFACE, fg=TXT_DIM).grid(row=0, column=2,
                                              sticky="e", padx=16)

        # Garis pemisah bawah header
        tk.Frame(self.root, bg=BORDER, height=1).grid(
            row=0, column=0, sticky="ew", pady=(52, 0))

        # ─────────────────────────────────────
        # 1) PANEL KONEKSI
        # ─────────────────────────────────────
        conn_wrap = tk.Frame(self.root, bg=BG)
        conn_wrap.grid(row=1, column=0, sticky="ew", padx=16, pady=(14, 0))
        conn_wrap.columnconfigure(0, weight=1)

        conn = tk.Frame(conn_wrap, bg=SURFACE,
                        highlightbackground=BORDER, highlightthickness=1)
        conn.grid(row=0, column=0, sticky="ew")
        conn.columnconfigure(3, weight=1)

        inner = tk.Frame(conn, bg=SURFACE)
        inner.pack(fill="x", padx=14, pady=10)

        # Label COM
        tk.Label(inner, text="COM PORT",
                 font=("Segoe UI", 8, "bold"),
                 bg=SURFACE, fg=TXT_DIM).pack(side="left", padx=(0, 8))

        # Combobox
        self.cb_port = ttk.Combobox(inner,
                                    textvariable=self.port_var,
                                    width=11,
                                    state="readonly",
                                    style="Elegant.TCombobox")
        self.cb_port.pack(side="left", padx=(0, 6))

        # Tombol refresh port
        self._flat_btn(inner, "↺", self._refresh_ports,
                       bg=SURFACE2, fg=TXT_MED,
                       hover_bg=BORDER, w=3, padx=4, pady=4).pack(
                       side="left", padx=(0, 12))

        # Separator vertikal
        tk.Frame(inner, bg=BORDER, width=1, height=24).pack(
            side="left", padx=(0, 12))

        # Tombol koneksi
        self.btn_conn = self._flat_btn(
            inner, "▶  MULAI", self._toggle_conn,
            bg=ACCENT, fg=BG,
            hover_bg=ACCENT_DIM, w=13, pady=6, bold=True)
        self.btn_conn.pack(side="left")

        # Status indicator (kanan)
        status_frame = tk.Frame(inner, bg=SURFACE)
        status_frame.pack(side="right")
        self.dot = tk.Label(status_frame, text="●",
                            font=("Segoe UI", 10),
                            bg=SURFACE, fg=TXT_DIM)
        self.dot.pack(side="left", padx=(0, 4))
        tk.Label(status_frame, textvariable=self.status_var,
                 font=F_UI_SM, bg=SURFACE, fg=TXT_DIM).pack(side="left")

        # ─────────────────────────────────────
        # 2) DISPLAY BERAT
        # ─────────────────────────────────────
        disp_wrap = tk.Frame(self.root, bg=BG)
        disp_wrap.grid(row=2, column=0, sticky="ew", padx=16, pady=12)
        disp_wrap.columnconfigure(0, weight=1)

        disp = tk.Frame(disp_wrap, bg=SURFACE,
                        highlightbackground=BORDER, highlightthickness=1)
        disp.grid(row=0, column=0, sticky="ew")

        # Label sub-judul
        sub_row = tk.Frame(disp, bg=SURFACE)
        sub_row.pack(fill="x", padx=20, pady=(16, 0))
        tk.Label(sub_row, text="BERAT  REALTIME",
                 font=("Segoe UI", 8, "bold"),
                 bg=SURFACE, fg=TXT_DIM).pack(side="left")

        # Garis dekoratif
        tk.Frame(sub_row, bg=BORDER, height=1).pack(
            side="left", fill="x", expand=True, padx=(12, 0), pady=6)

        # Nilai berat
        berat_row = tk.Frame(disp, bg=SURFACE)
        berat_row.pack(pady=(8, 4))

        self.lbl_berat = tk.Label(
            berat_row, textvariable=self.berat_var,
            font=F_DISPLAY,
            bg=SURFACE, fg=TXT_DIM,
            width=9, anchor="e")
        self.lbl_berat.pack(side="left")

        self.lbl_satuan = tk.Label(
            berat_row, textvariable=self.satuan_var,
            font=F_UNIT,
            bg=SURFACE, fg=TXT_DIM)
        self.lbl_satuan.pack(side="left",
                             anchor="s", pady=(0, 18), padx=(10, 0))

        # Tombol SIMPAN
        self.btn_save = tk.Button(
            disp,
            text="SIMPAN  DATA",
            command=self._simpan,
            font=F_BTN_LG,
            bg=SUCCESS, fg="white",
            activebackground=SUCCESS_DIM, activeforeground="white",
            relief="flat", cursor="hand2",
            pady=12, state="disabled",
            bd=0
        )
        self.btn_save.pack(fill="x", padx=20, pady=(4, 16))
        self._add_hover(self.btn_save, SUCCESS, SUCCESS_DIM)

        # ─────────────────────────────────────
        # 3) TABEL DATA
        # ─────────────────────────────────────
        tbl_wrap = tk.Frame(self.root, bg=BG)
        tbl_wrap.grid(row=3, column=0, sticky="nsew", padx=16, pady=(0, 14))
        tbl_wrap.columnconfigure(0, weight=1)
        tbl_wrap.rowconfigure(1, weight=1)

        # Sub-header tabel
        sub = tk.Frame(tbl_wrap, bg=BG)
        sub.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        sub.columnconfigure(0, weight=1)

        tk.Label(sub, textvariable=self.count_var,
                 font=("Segoe UI", 9, "bold"),
                 bg=BG, fg=TXT_MED).grid(row=0, column=0, sticky="w")

        btn_row = tk.Frame(sub, bg=BG)
        btn_row.grid(row=0, column=1, sticky="e")

        self._flat_btn(btn_row, "🗑  Hapus", self._hapus,
                       bg=SURFACE2, fg=DANGER,
                       hover_bg=BORDER, w=10).pack(side="right", padx=(6, 0))
        self._flat_btn(btn_row, "📊  Export", self._export,
                       bg=INFO, fg="white",
                       hover_bg=INFO_DIM, w=10).pack(side="right")

        # Card tabel
        card = tk.Frame(tbl_wrap, bg=SURFACE,
                        highlightbackground=BORDER, highlightthickness=1)
        card.grid(row=1, column=0, sticky="nsew")
        card.columnconfigure(0, weight=1)
        card.rowconfigure(0, weight=1)

        # Treeview
        cols = ("no", "waktu", "berat", "satuan")
        self.tree = ttk.Treeview(card, columns=cols,
                                 show="headings", selectmode="browse")
        self.tree.heading("no",     text="No",     anchor="center")
        self.tree.heading("waktu",  text="Waktu",  anchor="w")
        self.tree.heading("berat",  text="Berat",  anchor="e")
        self.tree.heading("satuan", text="Sat.",   anchor="center")
        self.tree.column("no",     width=44,  minwidth=40,  anchor="center",  stretch=False)
        self.tree.column("waktu",  width=180, minwidth=140, anchor="w")
        self.tree.column("berat",  width=130, minwidth=100, anchor="e")
        self.tree.column("satuan", width=56,  minwidth=48,  anchor="center",  stretch=False)

        sb = ttk.Scrollbar(card, orient="vertical",
                           command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")

        self.tree.tag_configure("odd",  background=ROW_A)
        self.tree.tag_configure("even", background=ROW_B)

        # ─────────────────────────────────────
        # 4) STATUS BAR
        # ─────────────────────────────────────
        sbar = tk.Frame(self.root, bg=SURFACE, height=24)
        sbar.grid(row=4, column=0, sticky="ew")
        sbar.grid_propagate(False)
        sbar.columnconfigure(0, weight=1)

        tk.Frame(sbar, bg=BORDER, height=1).place(x=0, y=0, relwidth=1)

        tk.Label(sbar, text="A&D Weighing  ·  AD-4329A  ·  RS-232",
                 font=("Segoe UI", 7), bg=SURFACE, fg=TXT_DIM).pack(
                 side="right", padx=12, pady=4)

    # ─────────────────────────────────────────────────────────
    #  HELPER: TOMBOL FLAT + HOVER
    # ─────────────────────────────────────────────────────────

    def _flat_btn(self, parent, text, cmd, bg, fg,
                  hover_bg=None, w=10, padx=8, pady=5, bold=False):
        font = ("Segoe UI", 9, "bold") if bold else F_UI_B
        btn  = tk.Button(parent, text=text, command=cmd,
                         bg=bg, fg=fg,
                         activebackground=hover_bg or bg,
                         activeforeground=fg,
                         relief="flat", font=font,
                         width=w, cursor="hand2",
                         padx=padx, pady=pady, bd=0,
                         highlightthickness=0)
        if hover_bg:
            self._add_hover(btn, bg, hover_bg)
        return btn

    def _add_hover(self, widget, normal_bg, hover_bg):
        widget.bind("<Enter>",
                    lambda e: widget.config(bg=hover_bg,
                                            activebackground=hover_bg))
        widget.bind("<Leave>",
                    lambda e: widget.config(bg=normal_bg,
                                            activebackground=hover_bg))

    # ─────────────────────────────────────────────────────────
    #  PORT
    # ─────────────────────────────────────────────────────────

    def _refresh_ports(self):
        ports = [p.device for p in serial.tools.list_ports.comports()]
        self.cb_port["values"] = ports
        if ports:
            self.cb_port.current(0)

    # ─────────────────────────────────────────────────────────
    #  KONEKSI
    # ─────────────────────────────────────────────────────────

    def _toggle_conn(self):
        if self.running:
            self._stop()
        else:
            self._start()

    def _start(self):
        port = self.port_var.get()
        if not port:
            messagebox.showwarning("Port Kosong", "Pilih COM port terlebih dahulu.")
            return
        try:
            try:
                _t = serial.Serial(port=port); _t.close()
            except Exception:
                pass
            time.sleep(0.2)

            self.ser = serial.Serial(
                port=port,
                baudrate=DEFAULT_BAUD,
                bytesize=DEFAULT_BYTESIZE,
                parity=DEFAULT_PARITY,
                stopbits=DEFAULT_STOPBITS,
                timeout=DEFAULT_TIMEOUT
            )
            self.ser.reset_input_buffer()
            self.running = True

            self.btn_conn.config(text="■  STOP",
                                 bg=DANGER, fg="white",
                                 activebackground=DANGER_DIM,
                                 activeforeground="white")
            self._add_hover(self.btn_conn, DANGER, DANGER_DIM)

            self.btn_save.config(state="normal")
            self.status_var.set(f"Terhubung — {port}")
            self.dot.config(fg=ACCENT)
            self.lbl_berat.config(fg=ACCENT)
            self.lbl_satuan.config(fg=ACCENT)

            self.thread = threading.Thread(
                target=self._read_loop, daemon=True)
            self.thread.start()

        except serial.SerialException as e:
            messagebox.showerror("Gagal Terhubung", str(e))

    def _stop(self):
        self.running = False
        if self.thread and self.thread.is_alive():
            self.thread.join(timeout=2.0)
        self.thread = None
        try:
            if self.ser and self.ser.is_open:
                self.ser.reset_input_buffer()
                self.ser.close()
        except Exception:
            pass
        self.ser = None
        try:
            self.btn_conn.config(text="▶  MULAI",
                                 bg=ACCENT, fg=BG,
                                 activebackground=ACCENT_DIM,
                                 activeforeground=BG)
            self._add_hover(self.btn_conn, ACCENT, ACCENT_DIM)
            self.btn_save.config(state="disabled")
            self.status_var.set("Tidak Terhubung")
            self.dot.config(fg=TXT_DIM)
            self.berat_var.set("––––.–––")
            self.lbl_berat.config(fg=TXT_DIM)
            self.lbl_satuan.config(fg=TXT_DIM)
        except Exception:
            pass

    # ─────────────────────────────────────────────────────────
    #  READ LOOP (thread)
    # ─────────────────────────────────────────────────────────

    def _read_loop(self):
        while self.running:
            try:
                if self.ser and self.ser.in_waiting > 0:
                    raw = self.ser.readline().decode(
                        'ascii', errors='replace').strip()
                    if raw:
                        berat, satuan = parse_berat(raw)
                        if berat is not None:
                            self.berat_live  = berat
                            self.satuan_live = satuan
                            self.root.after(0, self._update_display,
                                            berat, satuan)
            except Exception:
                break
            time.sleep(0.01)

    def _update_display(self, berat, satuan):
        self.berat_var.set(f"{int(round(berat)):,}".replace(",", "."))
        self.satuan_var.set(satuan)

    # ─────────────────────────────────────────────────────────
    #  SIMPAN DATA
    # ─────────────────────────────────────────────────────────

    def _simpan(self):
        if self.berat_live is None:
            messagebox.showwarning("Belum Ada Data",
                                   "Belum ada pembacaan berat dari timbangan.")
            return

        no     = len(self.data) + 1
        waktu  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        berat  = self.berat_live
        satuan = self.satuan_live

        self.data.append((no, waktu, berat, satuan))

        tag = "odd" if no % 2 == 1 else "even"
        berat_fmt = f"{int(round(berat)):,}".replace(",", ".")
        self.tree.insert("", "end",
                         values=(no, waktu, berat_fmt, satuan),
                         tags=(tag,))
        self.tree.yview_moveto(1.0)
        self.count_var.set(f"{no} data tersimpan")

        # Flash konfirmasi
        orig = self.lbl_berat.cget("fg")
        self.lbl_berat.config(fg=SUCCESS)
        self.root.after(300, lambda: self.lbl_berat.config(fg=orig))

    # ─────────────────────────────────────────────────────────
    #  HAPUS
    # ─────────────────────────────────────────────────────────

    def _hapus(self):
        if not self.data:
            messagebox.showinfo("Kosong", "Tidak ada data untuk dihapus.")
            return
        if messagebox.askyesno("Hapus Semua Data",
                               f"Hapus semua {len(self.data)} data tersimpan?"):
            self.data.clear()
            for item in self.tree.get_children():
                self.tree.delete(item)
            self.count_var.set("0 data tersimpan")

    # ─────────────────────────────────────────────────────────
    #  EXPORT
    # ─────────────────────────────────────────────────────────

    def _export(self):
        if not self.data:
            messagebox.showinfo("Kosong", "Belum ada data untuk diekspor.")
            return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        if EXCEL_OK:
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Workbook", "*.xlsx"), ("Semua File", "*.*")],
                initialfile=f"timbangan_{ts}.xlsx",
                title="Export ke Excel"
            )
            if path:
                self._write_xlsx(path)
        else:
            path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV", "*.csv"), ("Semua File", "*.*")],
                initialfile=f"timbangan_{ts}.csv",
                title="Export ke CSV"
            )
            if path:
                self._write_csv(path)

    def _write_xlsx(self, path):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Timbangan"

            thin   = Side(style='thin', color="2E3347")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Judul
            ws.merge_cells("A1:D1")
            c = ws["A1"]
            c.value     = "DATA TIMBANGAN  —  AD-4329A  |  A&D Weighing"
            c.font      = Font(name="Calibri", size=13,
                               bold=True, color="3ECFB2")
            c.fill      = PatternFill("solid", fgColor="111318")
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            # Info
            berats = [d[2] for d in self.data]
            info = [
                f"Tanggal Export : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"Total Data     : {len(self.data)} entri",
                f"Min / Maks     : {min(berats):.3f} / {max(berats):.3f}",
                f"Rata-rata      : {sum(berats)/len(berats):.3f}",
            ]
            for i, txt in enumerate(info, 2):
                cell = ws.cell(i, 1)
                cell.value = txt
                cell.font  = Font(name="Calibri", size=9,
                                  color="8B94B0", italic=True)
            ws.row_dimensions[6].height = 6

            # Header tabel
            headers = ["No", "Waktu", "Berat", "Satuan"]
            hfill = PatternFill("solid", fgColor="1C1F2B")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(7, col, h)
                cell.font      = Font(name="Calibri", size=10,
                                      bold=True, color="3ECFB2")
                cell.fill      = hfill
                cell.alignment = Alignment(horizontal="center")
                cell.border    = border
            ws.row_dimensions[7].height = 22

            # Data rows
            fill_a = PatternFill("solid", fgColor="1C1F2B")
            fill_b = PatternFill("solid", fgColor="191C28")
            for i, (no, waktu, berat, satuan) in enumerate(self.data, 1):
                row  = 7 + i
                fill = fill_a if i % 2 == 1 else fill_b
                fnt  = Font(name="Calibri", size=10, color="DCE1EF")
                berat_int = int(round(berat))
                vals   = [no, waktu, berat_int, satuan]
                aligns = ["center", "left", "right", "center"]
                for col, (val, aln) in enumerate(zip(vals, aligns), 1):
                    cell = ws.cell(row, col, val)
                    cell.font      = fnt
                    cell.fill      = fill
                    cell.border    = border
                    cell.alignment = Alignment(horizontal=aln)
                    if col == 3:
                        cell.number_format = '#,##0'

            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 24
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 9

            wb.save(path)
            messagebox.showinfo("Berhasil",
                                f"✅  Data berhasil diekspor ke Excel!\n\n{path}")

        except Exception as e:
            messagebox.showerror("Gagal Export", str(e))

    def _write_csv(self, path):
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write("No,Waktu,Berat,Satuan\n")
                for no, waktu, berat, satuan in self.data:
                    f.write(f"{no},{waktu},{int(round(berat))},{satuan}\n")
            messagebox.showinfo("Berhasil",
                                f"✅  Tersimpan sebagai CSV\n"
                                f"(Install openpyxl untuk format .xlsx)\n\n{path}")
        except Exception as e:
            messagebox.showerror("Gagal Export", str(e))

    # ─────────────────────────────────────────────────────────
    #  CLOSE
    # ─────────────────────────────────────────────────────────

    def on_close(self):
        self.running = False
        if self.thread and self.thread.is_alive():
            self.thread.join(timeout=2.0)
        self.thread = None
        try:
            if self.ser and self.ser.is_open:
                self.ser.reset_input_buffer()
                self.ser.close()
        except Exception:
            pass
        self.ser = None
        time.sleep(0.3)
        self.root.destroy()


# ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app  = TimbanganApp(root)
    root.protocol("WM_DELETE_WINDOW", app.on_close)
    root.mainloop()