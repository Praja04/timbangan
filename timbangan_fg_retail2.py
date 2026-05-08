"""
Timbangan AND GX-4000 — Quality Control Data Logger
UI: Maroon & White, 3-column layout
Flow: NIK → Variant → Mesin → Simpan

Perubahan:
  - Tabel logger difilter otomatis sesuai variant yang dipilih
  - NIK ditambahkan ke payload API
  - _refresh_table() method untuk re-render tabel
  - Filter info label (x/total data)

Requirements:
    pip install pyserial openpyxl requests
"""

import tkinter as tk
from tkinter import messagebox, filedialog
import serial
import serial.tools.list_ports
import threading
import queue
import time
import re
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ── PALETTE ────────────────────────────────────────────────────
MAROON       = "#7B1D1D"
MAROON_DARK  = "#5C1414"
MAROON_LIGHT = "#9B2C2C"
MAROON_PALE  = "#FDF2F2"
MAROON_SOFT  = "#F5D5D5"
WHITE        = "#FFFFFF"
GRAY_50      = "#FAFAFA"
GRAY_100     = "#F4F4F5"
GRAY_200     = "#E4E4E7"
GRAY_400     = "#A1A1AA"
GRAY_600     = "#52525B"
GRAY_800     = "#27272A"
GREEN        = "#16A34A"
GREEN_LIGHT  = "#DCFCE7"
RED_ALERT    = "#DC2626"
BG_ROOT      = "#EDE8E8"

# ── FONTS ──────────────────────────────────────────────────────
F_TITLE   = ("Segoe UI", 13, "bold")
F_SUB     = ("Segoe UI", 9)
F_LABEL   = ("Segoe UI", 8, "bold")
F_BODY    = ("Segoe UI", 10)
F_BODY_B  = ("Segoe UI", 10, "bold")
F_SMALL   = ("Segoe UI", 8)
F_CHIP    = ("Segoe UI", 9, "bold")
F_CHIP_S  = ("Segoe UI", 8)
F_MESIN   = ("Segoe UI", 12, "bold")
F_MESIN_S = ("Segoe UI", 8)
F_WEIGHT  = ("Courier New", 52, "bold")
F_UNIT    = ("Segoe UI", 13)
F_STATUS  = ("Segoe UI", 10, "bold")
F_NIK     = ("Courier New", 13, "bold")
F_BTN     = ("Segoe UI", 11, "bold")
F_BTN_SM  = ("Segoe UI", 9, "bold")
F_TABLE   = ("Segoe UI", 9)
F_TABLE_H = ("Segoe UI", 8, "bold")
F_MONO    = ("Courier New", 9)
F_BADGE   = ("Segoe UI", 8, "bold")

# ── SERIAL CONFIG ───────────────────────────────────────────────
BAUDRATE = 9600
DATABITS = 7
PARITY   = "E"
STOPBITS = 1

# ── API ─────────────────────────────────────────────────────────
API_URL = "http://10.11.10.130:8081/api/mesin"

# ── DATA ────────────────────────────────────────────────────────
VARIANT_STANDARDS = {
    "YB P1000G":       {"min": 1007.50, "std": 1012.50, "max": 1017.50, "code": "P1000G"},
    "YB P700G Piring": {"min":  706.00, "std":  711.00, "max":  716.00, "code": "P700G-PIR"},
    "BB P700G NP":     {"min":  706.00, "std":  711.00, "max":  716.00, "code": "P700G-NP"},
    "BB P270G":        {"min":  273.00, "std":  275.00, "max":  277.00, "code": "P270G"},
    "BB P725G":        {"min":  730.00, "std":  735.00, "max":  740.00, "code": "P725G"},
    "YB P700G":        {"min":  706.00, "std":  711.00, "max":  716.00, "code": "P700G"},
    "BB P77G Harga":   {"min":   78.70, "std":   79.20, "max":   82.70, "code": "P77G-H"},
    "BB P77G BDKT":    {"min":   78.70, "std":   79.20, "max":   82.70, "code": "P77G-B"},
    "YB P77G B5G1":    {"min":   78.70, "std":   79.20, "max":   82.70, "code": "P77G-B5"},
    "YB P77G BDKT":    {"min":   78.70, "std":   79.20, "max":   82.70, "code": "P77G-BD"},
    "YB P250G":        {"min":  253.00, "std":  255.00, "max":  257.00, "code": "P250G"},
    "YB S20G":         {"min":  244.68, "std":  247.68, "max":  259.68, "code": "S20G"},
    "YB S20G PCS":     {"min":   19.14, "std":   20.64, "max":   21.64, "code": "S20G-P"},
    "BB S40G":         {"min":  489.20, "std":  493.20, "max":  505.20, "code": "S40G"},
    "BB S40G PCS":     {"min":   39.10, "std":   41.10, "max":   42.10, "code": "S40G-P"},
    "YB P550G":        {"min":  556.00, "std":  561.00, "max":  566.00, "code": "P550G"},
}

MACHINES = [
    ("A","F2"),("AE","D12"),("AF","D13"),("AG","D14"),("AH","D15"),
    ("AI","D16"),("AJ","D17"),("AK","D18"),("B","D11/E5"),("C","D9"),
    ("D","D1"),("E","D2"),("F","D3"),("G","D4"),("H","D5"),("I","D6"),
    ("J","D7"),("K","D8"),("L","LD10"),("O","C1"),("P","C2"),("Q","A2"),
    ("R","C3"),("U","F3"),("V","F1"),("W","C7"),("X","C8"),("Y","B6"),("Z","B3"),
]


# ════════════════════════════════════════════════════════════════
class RoundedFrame(tk.Canvas):
    """Canvas-based rounded rectangle container."""
    def __init__(self, parent, radius=12, bg=WHITE, border=None,
                 border_width=1, **kwargs):
        super().__init__(parent, highlightthickness=0,
                         bg=parent["bg"] if hasattr(parent, "__getitem__") else BG_ROOT,
                         **kwargs)
        self._radius = radius
        self._bg     = bg
        self._border = border or GRAY_200
        self._bw     = border_width
        self.bind("<Configure>", self._draw)
        self._frame  = tk.Frame(self, bg=bg)
        self.create_window(0, 0, window=self._frame, anchor="nw", tags="inner")

    def _draw(self, event=None):
        self.delete("bg")
        w, h, r = self.winfo_width(), self.winfo_height(), self._radius
        self._rounded_rect(2, 2, w-2, h-2, r)
        self.itemconfig("inner", width=w-4, height=h-4)
        self.coords("inner", 2, 2)

    def _rounded_rect(self, x1, y1, x2, y2, r):
        pts = [
            x1+r, y1,  x2-r, y1,
            x2,   y1,  x2,   y1+r,
            x2,   y2-r,x2,   y2,
            x2-r, y2,  x1+r, y2,
            x1,   y2,  x1,   y2-r,
            x1,   y1+r,x1,   y1,
        ]
        self.create_polygon(pts, smooth=True,
                            fill=self._bg, outline=self._border,
                            width=self._bw, tags="bg")

    @property
    def inner(self):
        return self._frame


# ════════════════════════════════════════════════════════════════
class ChipButton(tk.Frame):
    """Selectable chip/tag button."""
    def __init__(self, parent, text, subtext="", command=None,
                 normal_bg=WHITE, sel_bg=MAROON,
                 normal_fg=GRAY_800, sel_fg=WHITE,
                 normal_border=GRAY_200, sel_border=MAROON, **kwargs):
        super().__init__(parent, bg=parent["bg"], **kwargs)
        self._cmd = command
        self._sel = False
        self._nbg = normal_bg
        self._sbg = sel_bg
        self._nfg = normal_fg
        self._sfg = sel_fg
        self._nb  = normal_border
        self._sb  = sel_border

        self._card = tk.Frame(self, bg=normal_bg,
                              highlightbackground=normal_border,
                              highlightthickness=1,
                              cursor="hand2")
        self._card.pack(fill="both", expand=True)

        self._lbl = tk.Label(self._card, text=text, font=F_CHIP,
                             bg=normal_bg, fg=normal_fg, anchor="w",
                             padx=7, pady=3)
        self._lbl.pack(side="top", fill="x")

        if subtext:
            self._sub = tk.Label(self._card, text=subtext, font=F_CHIP_S,
                                 bg=normal_bg, fg=GRAY_400, anchor="w",
                                 padx=7, pady=0)
            self._sub.pack(side="top", fill="x")
        else:
            self._sub = None

        for w in (self._card, self._lbl) + ((self._sub,) if self._sub else ()):
            w.bind("<Button-1>", self._click)
            w.bind("<Enter>",    self._hover)
            w.bind("<Leave>",    self._leave)

    def _click(self, e=None):
        if self._cmd:
            self._cmd()

    def _hover(self, e=None):
        if not self._sel:
            self._card.config(bg=MAROON_PALE, highlightbackground=MAROON_LIGHT)
            self._lbl.config(bg=MAROON_PALE)
            if self._sub:
                self._sub.config(bg=MAROON_PALE)

    def _leave(self, e=None):
        if not self._sel:
            self._card.config(bg=self._nbg, highlightbackground=self._nb)
            self._lbl.config(bg=self._nbg)
            if self._sub:
                self._sub.config(bg=self._nbg)

    def select(self):
        self._sel = True
        self._card.config(bg=self._sbg, highlightbackground=self._sb)
        self._lbl.config(bg=self._sbg, fg=self._sfg)
        if self._sub:
            self._sub.config(bg=self._sbg, fg=WHITE)

    def deselect(self):
        self._sel = False
        self._card.config(bg=self._nbg, highlightbackground=self._nb)
        self._lbl.config(bg=self._nbg, fg=self._nfg)
        if self._sub:
            self._sub.config(bg=self._nbg, fg=GRAY_400)


# ════════════════════════════════════════════════════════════════
class MesinButton(tk.Frame):
    """Square card for machine selection."""
    def __init__(self, parent, letter, pos, command=None, **kwargs):
        super().__init__(parent, bg=parent["bg"], **kwargs)
        self._cmd = command
        self._sel = False

        self._card = tk.Frame(self, bg=WHITE,
                              highlightbackground=GRAY_200,
                              highlightthickness=1,
                              cursor="hand2")
        self._card.pack(fill="both", expand=True)

        self._let = tk.Label(self._card, text=letter, font=F_MESIN,
                             bg=WHITE, fg=MAROON, pady=4)
        self._let.pack()
        self._pos = tk.Label(self._card, text=pos, font=F_MESIN_S,
                             bg=WHITE, fg=GRAY_400, pady=0)
        self._pos.pack()

        for w in (self._card, self._let, self._pos):
            w.bind("<Button-1>", self._click)
            w.bind("<Enter>",    self._hover)
            w.bind("<Leave>",    self._leave)

    def _click(self, e=None):
        if self._cmd:
            self._cmd()

    def _hover(self, e=None):
        if not self._sel:
            for w in (self._card, self._let, self._pos):
                w.config(bg=MAROON_PALE)
            self._card.config(highlightbackground=MAROON_LIGHT)

    def _leave(self, e=None):
        if not self._sel:
            for w in (self._card, self._let, self._pos):
                w.config(bg=WHITE)
            self._card.config(highlightbackground=GRAY_200)

    def select(self):
        self._sel = True
        for w in (self._card, self._let, self._pos):
            w.config(bg=MAROON)
        self._let.config(fg=WHITE)
        self._pos.config(fg=WHITE)
        self._card.config(highlightbackground=MAROON_DARK)

    def deselect(self):
        self._sel = False
        for w in (self._card, self._let, self._pos):
            w.config(bg=WHITE)
        self._let.config(fg=MAROON)
        self._pos.config(fg=GRAY_400)
        self._card.config(highlightbackground=GRAY_200)


# ════════════════════════════════════════════════════════════════
class ScaleApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Timbangan AND GX-4000 — Quality Control")
        self.root.configure(bg=BG_ROOT)
        self.root.resizable(True, True)

        # State
        self.serial_conn   = None
        self.thread        = None
        self.is_reading    = False
        self.data_queue    = queue.Queue()
        self.current_data  = None
        self.saved_data    = []          # semua record tersimpan (di RAM)
        self.sel_variant   = None
        self.sel_machine   = None
        self.nik_confirmed = False
        self.live_weight   = 0.0
        self.live_unit     = "g"
        self.auto_port     = None
        self.filtered_variant = None   # variant aktif untuk filter tabel
        self._active_config   = "2013"    # config serial berhasil (7E1/8N1)

        # Chip/button registries
        self.variant_btns = {}
        self.mesin_btns   = {}

        self._build_ui()
        self._auto_detect_port()
        self._check_queue()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── BUILD UI ────────────────────────────────────────────────
    def _build_ui(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)
        self._build_header()
        self._build_stepbar()
        self._build_body()
        self._build_footer()

    # ── HEADER ──────────────────────────────────────────────────
    def _build_header(self):
        hdr = tk.Frame(self.root, bg=MAROON, height=60)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.columnconfigure(1, weight=1)
        hdr.grid_propagate(False)

        tk.Label(hdr, text="⚖", font=("Segoe UI", 18),
                 bg=MAROON, fg=WHITE, padx=16).grid(row=0, column=0,
                                                    sticky="ns", pady=10)

        title_f = tk.Frame(hdr, bg=MAROON)
        title_f.grid(row=0, column=1, sticky="w", pady=10)
        tk.Label(title_f, text="Timbangan AND GX-4000",
                 font=F_TITLE, bg=MAROON, fg=WHITE).pack(anchor="w")
        tk.Label(title_f, text="Quality Control · Data Logger",
                 font=F_SUB, bg=MAROON, fg=WHITE).pack(anchor="w")

        badge = tk.Frame(hdr, bg=MAROON,
                         highlightbackground=WHITE, highlightthickness=1)
        badge.grid(row=0, column=2, sticky="e", padx=20, pady=15)

        self.conn_dot = tk.Label(badge, text="●", font=("Segoe UI", 10),
                                 bg=MAROON, fg=GRAY_400)
        self.conn_dot.pack(side="left", padx=(10, 4))
        self.conn_lbl = tk.Label(badge, text="Mendeteksi port...",
                                 font=F_SMALL, bg=MAROON, fg=WHITE)
        self.conn_lbl.pack(side="left", padx=(0, 10))

    # ── STEP BAR ────────────────────────────────────────────────
    def _build_stepbar(self):
        bar = tk.Frame(self.root, bg=MAROON_PALE,
                       highlightbackground=MAROON_SOFT, highlightthickness=1)
        bar.grid(row=1, column=0, sticky="ew")

        inner = tk.Frame(bar, bg=MAROON_PALE)
        inner.pack(side="left", padx=24, pady=10)

        steps = [("1", "NIK Operator"), ("2", "Variant & Mesin"), ("3", "Catat Timbangan")]
        self.step_nums = []
        self.step_lbls = []

        for i, (num, lbl) in enumerate(steps):
            if i > 0:
                tk.Label(inner, text="──", font=("Segoe UI", 9),
                         bg=MAROON_PALE, fg=MAROON_SOFT).pack(side="left", padx=4)

            sf = tk.Frame(inner, bg=MAROON_PALE)
            sf.pack(side="left")

            nf = tk.Label(sf, text=num, font=("Segoe UI", 9, "bold"),
                          bg=MAROON_SOFT, fg=MAROON, width=2, relief="flat")
            nf.pack(side="left", padx=(0, 6))

            ll = tk.Label(sf, text=lbl, font=("Segoe UI", 9, "bold"),
                          bg=MAROON_PALE, fg=GRAY_400)
            ll.pack(side="left")

            self.step_nums.append(nf)
            self.step_lbls.append(ll)

        self._set_step(0)

    def _set_step(self, active):
        for i, (nf, ll) in enumerate(zip(self.step_nums, self.step_lbls)):
            if i < active:
                nf.config(bg=GREEN, fg=WHITE, text="✓")
                ll.config(fg=GREEN)
            elif i == active:
                nf.config(bg=MAROON, fg=WHITE, text=str(i+1))
                ll.config(fg=MAROON)
            else:
                nf.config(bg=MAROON_SOFT, fg=MAROON, text=str(i+1))
                ll.config(fg=GRAY_400)

    # ── BODY ────────────────────────────────────────────────────
    def _build_body(self):
        body = tk.Frame(self.root, bg=WHITE)
        body.grid(row=2, column=0, sticky="nsew")
        self.root.rowconfigure(2, weight=1)
        body.columnconfigure(2, weight=1)
        body.rowconfigure(0, weight=1)
        self._build_col1(body)
        self._build_col2(body)
        self._build_col3(body)

    # ── COL 1 — NIK + Variant ───────────────────────────────────
    def _build_col1(self, parent):
        col = tk.Frame(parent, bg=WHITE,
                       highlightbackground=GRAY_100, highlightthickness=1)
        col.grid(row=0, column=0, sticky="nsew")
        col.configure(width=320)
        parent.columnconfigure(0, minsize=320)

        inner = tk.Frame(col, bg=WHITE)
        inner.pack(fill="both", expand=True, padx=18, pady=16)

        # NIK section
        tk.Label(inner, text="NIK OPERATOR", font=F_LABEL,
                 bg=WHITE, fg=MAROON).pack(anchor="w", pady=(0, 6))

        nik_row = tk.Frame(inner, bg=WHITE)
        nik_row.pack(fill="x")

        self.nik_entry = tk.Entry(nik_row, font=F_NIK,
                                  bg=GRAY_50, fg=GRAY_800,
                                  insertbackground=MAROON,
                                  relief="flat",
                                  highlightbackground=GRAY_200,
                                  highlightthickness=1, width=12)
        self.nik_entry.pack(side="left", fill="y", ipady=6, padx=(0, 6))
        self.nik_entry.bind("<Return>", lambda e: self._confirm_nik())

        tk.Button(nik_row, text="Masuk", font=F_BTN_SM,
                  bg=MAROON, fg=WHITE, relief="flat",
                  activebackground=MAROON_DARK, activeforeground=WHITE,
                  cursor="hand2", padx=10,
                  command=self._confirm_nik).pack(side="left", fill="y")

        self.nik_ok_lbl = tk.Label(inner, text="", font=("Segoe UI", 9),
                                   bg=WHITE, fg=GREEN)
        self.nik_ok_lbl.pack(anchor="w", pady=(4, 0))

        tk.Frame(inner, bg=GRAY_100, height=1).pack(fill="x", pady=12)

        # Variant section
        tk.Label(inner, text="VARIANT PRODUK", font=F_LABEL,
                 bg=WHITE, fg=MAROON).pack(anchor="w", pady=(0, 8))

        vf = tk.Frame(inner, bg=WHITE)
        vf.pack(fill="x")

        for i, (name, data) in enumerate(VARIANT_STANDARDS.items()):
            row     = i // 2
            col_idx = i % 2
            chip = ChipButton(
                vf, text=name, subtext=data["code"],
                command=lambda n=name: self._pick_variant(n)
            )
            chip.grid(row=row, column=col_idx, sticky="ew",
                      padx=(0 if col_idx == 0 else 3, 3 if col_idx == 0 else 0),
                      pady=2)
            self.variant_btns[name] = chip

        vf.columnconfigure(0, weight=1)
        vf.columnconfigure(1, weight=1)

    # ── COL 2 — Mesin ───────────────────────────────────────────
    def _build_col2(self, parent):
        col = tk.Frame(parent, bg=WHITE,
                       highlightbackground=GRAY_100, highlightthickness=1)
        col.grid(row=0, column=1, sticky="nsew")
        col.configure(width=300)
        parent.columnconfigure(1, minsize=300)

        inner = tk.Frame(col, bg=WHITE)
        inner.pack(fill="both", expand=True, padx=18, pady=16)

        tk.Label(inner, text="PILIH MESIN", font=F_LABEL,
                 bg=WHITE, fg=MAROON).pack(anchor="w", pady=(0, 8))

        gf = tk.Frame(inner, bg=WHITE)
        gf.pack(fill="x")

        COLS = 4
        for i, (letter, pos) in enumerate(MACHINES):
            row = i // COLS
            ci  = i % COLS
            btn = MesinButton(
                gf, letter=letter, pos=pos,
                command=lambda l=letter: self._pick_mesin(l)
            )
            btn.grid(row=row, column=ci, sticky="ew",
                     padx=2, pady=2, ipadx=0, ipady=2)
            self.mesin_btns[letter] = btn

        for c in range(COLS):
            gf.columnconfigure(c, weight=1)

    # ── COL 3 — Weight + Logger Table ───────────────────────────
    def _build_col3(self, parent):
        col = tk.Frame(parent, bg=WHITE)
        col.grid(row=0, column=2, sticky="nsew")

        inner = tk.Frame(col, bg=WHITE)
        inner.pack(fill="both", expand=True, padx=18, pady=16)
        inner.rowconfigure(5, weight=1)
        inner.columnconfigure(0, weight=1)

        # ── Weight card ──────────────────────────────────────────
        wcard = tk.Frame(inner, bg=MAROON_PALE,
                         highlightbackground=MAROON_SOFT, highlightthickness=1)
        wcard.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        tk.Label(wcard, text="BERAT TERBACA", font=F_LABEL,
                 bg=MAROON_PALE, fg=MAROON, pady=10).pack()

        self.weight_lbl = tk.Label(wcard, text="–––.––",
                                   font=F_WEIGHT, bg=MAROON_PALE, fg=MAROON)
        self.weight_lbl.pack()

        self.unit_lbl = tk.Label(wcard, text="gram",
                                 font=F_UNIT, bg=MAROON_PALE, fg=GRAY_400)
        self.unit_lbl.pack(pady=(0, 4))

        status_row = tk.Frame(wcard, bg=MAROON_PALE)
        status_row.pack(pady=(0, 12))

        self.status_dot = tk.Label(status_row, text="●",
                                   font=("Segoe UI", 10),
                                   bg=MAROON_PALE, fg=GRAY_400)
        self.status_dot.pack(side="left", padx=(0, 4))
        self.status_txt = tk.Label(status_row, text="Pilih variant untuk validasi",
                                   font=F_STATUS, bg=MAROON_PALE, fg=GRAY_600)
        self.status_txt.pack(side="left")

        # ── STD pills ────────────────────────────────────────────
        std_row = tk.Frame(inner, bg=WHITE)
        std_row.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        std_row.columnconfigure((0, 1, 2), weight=1)

        self.std_min = self._std_pill(std_row, "Min",     "–", RED_ALERT, 0)
        self.std_std = self._std_pill(std_row, "Standar", "–", GRAY_800,  1)
        self.std_max = self._std_pill(std_row, "Max",     "–", MAROON,    2)

        # ── Validation message ───────────────────────────────────
        self.valid_lbl = tk.Label(inner, text="", font=("Segoe UI", 9),
                                  bg=WHITE, fg=GRAY_400)
        self.valid_lbl.grid(row=2, column=0, sticky="w", pady=(0, 6))

        # ── Save button ──────────────────────────────────────────
        self.save_btn = tk.Button(
            inner, text="Simpan Data Timbangan",
            font=F_BTN, bg=GRAY_200, fg=GRAY_400,
            relief="flat", cursor="arrow",
            activebackground=MAROON_DARK, activeforeground=WHITE,
            command=self._save_data, state="disabled",
            pady=10
        )
        self.save_btn.grid(row=3, column=0, sticky="ew", pady=(0, 14))

        # ── Table header row ─────────────────────────────────────
        tbl_hdr = tk.Frame(inner, bg=WHITE)
        tbl_hdr.grid(row=4, column=0, sticky="ew", pady=(0, 6))
        tbl_hdr.columnconfigure(0, weight=1)

        tk.Label(tbl_hdr, text="DATA TERSIMPAN", font=F_LABEL,
                 bg=WHITE, fg=MAROON).grid(row=0, column=0, sticky="w")

        # Filter info + count pada satu frame kanan
        right_info = tk.Frame(tbl_hdr, bg=WHITE)
        right_info.grid(row=0, column=1, sticky="e")

        self.filter_lbl = tk.Label(right_info, text="",
                                   font=F_SMALL, bg=WHITE, fg=MAROON)
        self.filter_lbl.pack(side="left", padx=(0, 6))

        self.count_lbl = tk.Label(right_info, text="0 data",
                                  font=F_SMALL, bg=WHITE, fg=GRAY_400)
        self.count_lbl.pack(side="left")

        # Tombol reset filter
        self.reset_filter_btn = tk.Button(
            right_info, text="Tampilkan Semua",
            font=F_SMALL, bg=WHITE, fg=GRAY_600,
            relief="flat",
            highlightbackground=GRAY_200, highlightthickness=1,
            cursor="hand2", padx=6, pady=1,
            command=self._reset_filter
        )
        self.reset_filter_btn.pack(side="left", padx=(6, 0))
        self.reset_filter_btn.pack_forget()   # sembunyikan dulu

        # ── Table body ───────────────────────────────────────────
        self._build_table(inner, row=5)

    def _std_pill(self, parent, label, val, color, col):
        f = tk.Frame(parent, bg=WHITE,
                     highlightbackground=GRAY_200, highlightthickness=1)
        f.grid(row=0, column=col, sticky="ew",
               padx=(0, 4) if col < 2 else 0)
        tk.Label(f, text=label, font=F_SMALL, bg=WHITE,
                 fg=GRAY_400, pady=4).pack()
        lbl = tk.Label(f, text=val, font=F_BODY_B, bg=WHITE,
                       fg=color, pady=4)
        lbl.pack()
        return lbl

    def _build_table(self, parent, row):
        tbl_frame = tk.Frame(parent, bg=WHITE)
        tbl_frame.grid(row=row, column=0, sticky="nsew")
        parent.rowconfigure(row, weight=1)
        tbl_frame.columnconfigure(0, weight=1)
        tbl_frame.rowconfigure(1, weight=1)

        # Header
        cols   = ["#", "Waktu", "Variant", "Mesin", "Berat", "Status"]
        widths = [30,   120,     160,       60,      80,      70]

        hdr = tk.Frame(tbl_frame, bg=GRAY_50)
        hdr.grid(row=0, column=0, sticky="ew")
        for c, w in zip(cols, widths):
            tk.Label(hdr, text=c, font=F_TABLE_H, bg=GRAY_50,
                     fg=GRAY_400, width=w//8, anchor="w",
                     padx=6, pady=5).pack(side="left")

        # Scrollable body
        canvas = tk.Canvas(tbl_frame, bg=WHITE, highlightthickness=0)
        sb = tk.Scrollbar(tbl_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        canvas.grid(row=1, column=0, sticky="nsew")
        sb.grid(row=1, column=1, sticky="ns")

        self.table_inner  = tk.Frame(canvas, bg=WHITE)
        self.table_window = canvas.create_window(
            0, 0, window=self.table_inner, anchor="nw")

        def _on_frame_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_configure(e):
            canvas.itemconfig(self.table_window, width=e.width)

        self.table_inner.bind("<Configure>", _on_frame_configure)
        canvas.bind("<Configure>", _on_canvas_configure)
        self._table_row_count = 0

    # ── Render satu baris tabel ──────────────────────────────────
    def _add_table_row(self, values, ok=True):
        r  = self._table_row_count
        bg = WHITE if r % 2 == 0 else GRAY_50
        widths = [30, 120, 160, 60, 80, 70]

        row_f = tk.Frame(self.table_inner, bg=bg,
                         highlightbackground=GRAY_100, highlightthickness=0)
        row_f.pack(fill="x")
        tk.Frame(row_f, bg=GRAY_100, height=1).pack(fill="x")
        cells = tk.Frame(row_f, bg=bg)
        cells.pack(fill="x")

        for i, (val, w) in enumerate(zip(values, widths)):
            if i == 5:  # Status badge
                badge_bg = GREEN_LIGHT if ok else "#FEE2E2"
                badge_fg = GREEN       if ok else RED_ALERT
                tk.Label(cells, text=val, font=F_BADGE,
                         bg=badge_bg, fg=badge_fg,
                         padx=6, pady=1).pack(side="left", padx=6, pady=4)
            else:
                font = F_MONO if i in (0, 1, 4) else F_TABLE
                tk.Label(cells, text=val, font=font,
                         bg=bg, fg=GRAY_600,
                         width=w//8, anchor="w", padx=6).pack(side="left")

        self._table_row_count += 1

    # ── Refresh tabel dengan filter variant aktif ────────────────
    def _refresh_table(self):
        # Hapus semua baris lama
        for widget in self.table_inner.winfo_children():
            widget.destroy()
        self._table_row_count = 0

        # Tentukan data yang ditampilkan
        if self.filtered_variant:
            data_to_show = [d for d in self.saved_data
                            if d["variant"] == self.filtered_variant]
        else:
            data_to_show = self.saved_data

        # Render ulang baris
        for i, d in enumerate(data_to_show, 1):
            ok        = d["status"] == "OK"
            t_short   = d["timestamp"][11:]   # HH:MM:SS
            var_short = d["variant"][:16] + ("…" if len(d["variant"]) > 16 else "")
            self._add_table_row(
                [str(i), t_short, var_short, d["machine"],
                 f"{d['weight']:.2f}g", d["status"]],
                ok=ok
            )

        # Update label filter & count
        shown = len(data_to_show)
        total = len(self.saved_data)

        if self.filtered_variant:
            short_name = (self.filtered_variant[:18] + "…"
                          if len(self.filtered_variant) > 18
                          else self.filtered_variant)
            self.filter_lbl.config(
                text=f"▶ {short_name}", fg=MAROON)
            self.reset_filter_btn.pack(side="left", padx=(6, 0))
        else:
            self.filter_lbl.config(text="", fg=GRAY_400)
            self.reset_filter_btn.pack_forget()

        if total == 0:
            self.count_lbl.config(text="0 data", fg=GRAY_400)
        elif shown == total:
            self.count_lbl.config(text=f"{total} data", fg=GRAY_400)
        else:
            self.count_lbl.config(
                text=f"{shown} dari {total} data", fg=MAROON)

    # ── Reset filter → tampilkan semua ──────────────────────────
    def _reset_filter(self):
        self.filtered_variant = None
        self._refresh_table()

    # ── FOOTER ──────────────────────────────────────────────────
    def _build_footer(self):
        ft = tk.Frame(self.root, bg=GRAY_50,
                      highlightbackground=GRAY_100, highlightthickness=1)
        ft.grid(row=3, column=0, sticky="ew")

        self.footer_lbl = tk.Label(
            ft, text=f"Port: Mendeteksi... · API: {API_URL}",
            font=F_SMALL, bg=GRAY_50, fg=GRAY_400)
        self.footer_lbl.pack(side="left", padx=16, pady=6)

        tk.Button(ft, text="↓ Export Excel",
                  font=F_BTN_SM, bg=WHITE, fg=GRAY_600,
                  relief="flat",
                  highlightbackground=GRAY_200, highlightthickness=1,
                  cursor="hand2", padx=10, pady=4,
                  command=self._export_excel).pack(
            side="right", padx=12, pady=6)

    # ── AUTO PORT DETECT ────────────────────────────────────────
    def _auto_detect_port(self):
        def _detect():
            ports = serial.tools.list_ports.comports()
            if not ports:
                self.root.after(0, self._set_conn_status, False,
                                "Tidak ada port ditemukan")
                self.root.after(3000, self._auto_detect_port)
                return

            parity_map = {"N": serial.PARITY_NONE,
                          "E": serial.PARITY_EVEN,
                          "O": serial.PARITY_ODD}

            # Urutan setting yang dicoba:
            # 1. 7E1 (setting resmi AND GX-4000)
            # 2. 8N1 (fallback untuk driver CH340 bermasalah)
            serial_configs = [
                {"bytesize": DATABITS,  "parity": parity_map[PARITY],  "label": "7E1"},
                {"bytesize": 8,         "parity": serial.PARITY_NONE,  "label": "8N1"},
            ]

            for p in ports:
                for cfg in serial_configs:
                    try:
                        conn = serial.Serial(
                            port     = p.device,
                            baudrate = BAUDRATE,
                            bytesize = cfg["bytesize"],
                            parity   = cfg["parity"],
                            stopbits = STOPBITS,
                            timeout  = 1
                        )
                        conn.reset_input_buffer()

                        self.serial_conn    = conn
                        self.auto_port      = p.device
                        self.is_reading     = True
                        self._active_config = cfg["label"]

                        self.root.after(0, self._set_conn_status, True,
                                        f"{p.device} [{cfg['label']}]")
                        self.thread = threading.Thread(
                            target=self._read_thread, daemon=True)
                        self.thread.start()
                        return   # sukses, berhenti

                    except Exception as e:
                        print(f"[GAGAL] {p.device} ({cfg['label']}) → {e}")
                        continue   # coba config berikutnya

            # Semua port + semua config gagal
            self.root.after(0, self._set_conn_status, False,
                            f"{len(ports)} port ditemukan, gagal konek")
            self.root.after(5000, self._auto_detect_port)

        threading.Thread(target=_detect, daemon=True).start()

    def _set_conn_status(self, connected, detail):
        if connected:
            self.conn_dot.config(fg="#4ADE80")
            self.conn_lbl.config(text=f"{detail} · Terhubung")
            self.footer_lbl.config(
                text=f"Port: {detail} (auto) · 9600bps · API: {API_URL}")
        else:
            self.conn_dot.config(fg=GRAY_400)
            self.conn_lbl.config(text=detail)
            self.footer_lbl.config(
                text=f"Port: {detail} · API: {API_URL}")

    # ── SERIAL READ ─────────────────────────────────────────────
    def _read_thread(self):
        while self.is_reading:
            try:
                if self.serial_conn and self.serial_conn.in_waiting > 0:
                    raw = self.serial_conn.readline().decode(
                        "ascii", errors="ignore").strip()
                    if raw:
                        parsed = self._parse(raw)
                        if parsed:
                            self.data_queue.put(parsed)
            except Exception:
                break
            time.sleep(0.01)

    def _parse(self, raw):
        """Parse AND GX-4000 format: ST,+00711.50  g"""
        try:
            m = re.match(
                r"^([A-Z]{2}),([+\-]?\d+\.?\d*)\s*([a-zA-Z]+)$", raw)
            if m:
                return {
                    "timestamp":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "scale_status": m.group(1),
                    "weight":       float(m.group(2)),
                    "unit":         m.group(3),
                    "raw":          raw,
                }
        except Exception:
            pass
        return None

    def _check_queue(self):
        try:
            while True:
                data = self.data_queue.get_nowait()
                self._update_weight(data)
        except queue.Empty:
            pass
        self.root.after(50, self._check_queue)

    def _update_weight(self, data):
        self.current_data = data
        self.live_weight  = data["weight"]
        self.live_unit    = data["unit"]
        self.weight_lbl.config(text=f"{data['weight']:.2f}")
        unit_map = {"g": "gram", "kg": "kilogram"}
        self.unit_lbl.config(text=unit_map.get(data["unit"], data["unit"]))
        self._validate_display(data["weight"])

    def _validate_display(self, weight):
        if not self.sel_variant:
            self.weight_lbl.config(fg=MAROON)
            self.status_dot.config(fg=GRAY_400)
            self.status_txt.config(text="Pilih variant untuk validasi",
                                   fg=GRAY_600)
            self.valid_lbl.config(text="")
            return

        std = VARIANT_STANDARDS[self.sel_variant]
        if weight < std["min"]:
            self.weight_lbl.config(fg=RED_ALERT)
            self.status_dot.config(fg=RED_ALERT)
            self.status_txt.config(text="NOT OK — Berat Kurang", fg=RED_ALERT)
            self.valid_lbl.config(
                text=f"Min: {std['min']}g  |  Selisih: {std['min']-weight:.2f}g",
                fg=RED_ALERT)
        elif weight > std["max"]:
            self.weight_lbl.config(fg=RED_ALERT)
            self.status_dot.config(fg=RED_ALERT)
            self.status_txt.config(text="NOT OK — Berat Lebih", fg=RED_ALERT)
            self.valid_lbl.config(
                text=f"Max: {std['max']}g  |  Selisih: {weight-std['max']:.2f}g",
                fg=RED_ALERT)
        else:
            self.weight_lbl.config(fg=GREEN)
            self.status_dot.config(fg=GREEN)
            self.status_txt.config(text="OK — Sesuai Standar", fg=GREEN)
            self.valid_lbl.config(
                text=f"Range: {std['min']}g – {std['max']}g", fg=GREEN)

    # ── NIK ─────────────────────────────────────────────────────
    def _confirm_nik(self):
        nik = self.nik_entry.get().strip()
        if not nik:
            return
        self.nik_confirmed = True
        self.nik_ok_lbl.config(
            text=f"✓  NIK {nik} — Operator Terverifikasi", fg=GREEN)
        self._set_step(1)

    # ── VARIANT PICK ────────────────────────────────────────────
    def _pick_variant(self, name):
        if not self.nik_confirmed:
            messagebox.showwarning("NIK Belum Dikonfirmasi",
                                   "Masukkan NIK operator terlebih dahulu.")
            return

        for n, btn in self.variant_btns.items():
            btn.deselect() if n != name else btn.select()

        self.sel_variant      = name
        self.filtered_variant = name   # ← set filter tabel ke variant ini

        std = VARIANT_STANDARDS[name]
        self.std_min.config(text=f"{std['min']:.2f}")
        self.std_std.config(text=f"{std['std']:.2f}")
        self.std_max.config(text=f"{std['max']:.2f}")

        if self.live_weight:
            self._validate_display(self.live_weight)

        self._check_save_ready()
        self._refresh_table()   # ← refresh tabel sesuai variant baru

        if self.sel_variant and self.sel_machine:
            self._set_step(2)

    # ── MESIN PICK ───────────────────────────────────────────────
    def _pick_mesin(self, letter):
        if not self.nik_confirmed:
            messagebox.showwarning("NIK Belum Dikonfirmasi",
                                   "Masukkan NIK operator terlebih dahulu.")
            return

        for l, btn in self.mesin_btns.items():
            btn.deselect() if l != letter else btn.select()

        self.sel_machine = letter
        self._check_save_ready()

        if self.sel_variant and self.sel_machine:
            self._set_step(2)

    def _check_save_ready(self):
        if self.sel_variant and self.sel_machine and self.nik_confirmed:
            self.save_btn.config(state="normal", bg=MAROON, fg=WHITE,
                                 cursor="hand2")
            self._set_step(2)
        else:
            self.save_btn.config(state="disabled", bg=GRAY_200, fg=GRAY_400,
                                 cursor="arrow")

    # ── SAVE DATA ───────────────────────────────────────────────
    def _save_data(self):
        if not self.current_data:
            messagebox.showwarning(
                "Tidak Ada Data",
                "Timbangan belum terbaca. Pastikan kabel terhubung.")
            return

        if not (self.sel_variant and self.sel_machine and self.nik_confirmed):
            messagebox.showerror("Lengkapi Data",
                                 "NIK, Variant, dan Mesin harus dipilih.")
            return

        w   = self.current_data["weight"]
        std = VARIANT_STANDARDS[self.sel_variant]
        ok  = std["min"] <= w <= std["max"]
        status = "OK" if ok else "NOT OK"
        nik    = self.nik_entry.get().strip()

        # Payload ke API — sudah termasuk NIK
        form = {
            "nik":     nik,
            "mesin":   self.sel_machine,
            "variant": self.sel_variant,
            "waktu":   self.current_data["timestamp"],
            "berat":   str(w),
            "unit":    self.current_data["unit"],
            "status":  status,
        }

        # Kirim API di background thread agar UI tidak freeze
        def _post():
            try:
                resp = requests.post(API_URL, data=form, timeout=8)
                api_status = ("✓ Terkirim"
                              if resp.status_code in (200, 201)
                              else f"Error {resp.status_code}")
            except requests.exceptions.ConnectionError:
                api_status = "Offline"
            except Exception:
                api_status = "Gagal"

            # Update record setelah response (opsional)
            for d in reversed(self.saved_data):
                if d.get("api_status") == "Pending":
                    d["api_status"] = api_status
                    break

        # Simpan ke list lokal dulu dengan status Pending
        record = {
            **self.current_data,
            "nik":        nik,
            "machine":    self.sel_machine,
            "variant":    self.sel_variant,
            "status":     status,
            "api_status": "Pending",
        }
        self.saved_data.append(record)

        # Jalankan POST di thread terpisah (non-blocking)
        threading.Thread(target=_post, daemon=True).start()

        # Refresh tabel (otomatis filter sesuai variant aktif)
        self._refresh_table()

        self._set_step(2)

        # Flash weight sebentar
        orig = self.weight_lbl.cget("fg")
        self.weight_lbl.config(fg=GREEN)
        self.root.after(300, lambda: self.weight_lbl.config(fg=orig))

    # ── EXPORT EXCEL ────────────────────────────────────────────
    def _export_excel(self):
        if not self.saved_data:
            messagebox.showinfo("Kosong", "Belum ada data untuk di-export.")
            return

        fn = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"timbangan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not fn:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Timbangan"

            thin   = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Title
            ws.merge_cells("A1:J1")
            c = ws["A1"]
            c.value = "DATA TIMBANGAN — AND GX-4000 | Quality Control"
            c.font  = Font(name="Calibri", size=13, bold=True, color="7B1D1D")
            c.fill  = PatternFill("solid", fgColor="FDF2F2")
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            # Headers — sekarang ada kolom NIK
            headers = ["No", "NIK", "Mesin", "Variant",
                       "Tanggal", "Waktu", "Berat", "Unit", "Status", "API"]
            hfill = PatternFill("solid", fgColor="7B1D1D")
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(2, ci, h)
                cell.font      = Font(name="Calibri", size=10,
                                      bold=True, color="FFFFFF")
                cell.fill      = hfill
                cell.alignment = Alignment(horizontal="center")
                cell.border    = border
            ws.row_dimensions[2].height = 20

            fill_a   = PatternFill("solid", fgColor="FFFFFF")
            fill_b   = PatternFill("solid", fgColor="FDF2F2")
            fill_ok  = PatternFill("solid", fgColor="DCFCE7")
            fill_nok = PatternFill("solid", fgColor="FEE2E2")

            count_ok = count_nok = 0
            for i, d in enumerate(self.saved_data, 1):
                row   = i + 2
                fill  = fill_a if i % 2 else fill_b
                dt    = datetime.strptime(d["timestamp"], "%Y-%m-%d %H:%M:%S")
                is_ok = d["status"] == "OK"
                if is_ok: count_ok  += 1
                else:     count_nok += 1

                vals = [
                    i,
                    d.get("nik", ""),
                    d["machine"],
                    d["variant"],
                    dt.strftime("%Y-%m-%d"),
                    dt.strftime("%H:%M:%S"),
                    d["weight"],
                    d["unit"],
                    d["status"],
                    d.get("api_status", "–"),
                ]
                for ci, val in enumerate(vals, 1):
                    cell = ws.cell(row, ci, val)
                    cell.font      = Font(name="Calibri", size=10)
                    cell.fill      = (fill_ok  if (ci == 9 and is_ok)  else
                                      fill_nok if (ci == 9 and not is_ok) else fill)
                    cell.border    = border
                    cell.alignment = Alignment(horizontal="center")

            # Column widths
            for ci, w in enumerate([6, 12, 10, 28, 14, 10, 10, 8, 10, 12], 1):
                ws.column_dimensions[
                    ws.cell(1, ci).column_letter].width = w

            # Summary
            sr = len(self.saved_data) + 4
            ws.cell(sr,   1).value = "RINGKASAN"
            ws.cell(sr,   1).font  = Font(bold=True, color="7B1D1D")
            ws.cell(sr+1, 1, f"Total  : {len(self.saved_data)}")
            ws.cell(sr+2, 1, f"OK     : {count_ok}")
            ws.cell(sr+3, 1, f"NOT OK : {count_nok}")

            wb.save(fn)
            messagebox.showinfo(
                "Export Berhasil",
                f"Data berhasil disimpan!\n\n"
                f"File  : {fn}\n"
                f"Total : {len(self.saved_data)} record\n"
                f"OK    : {count_ok}  |  NOT OK: {count_nok}"
            )
        except Exception as e:
            messagebox.showerror("Export Gagal", str(e))

    # ── CLOSE ───────────────────────────────────────────────────
    def _on_close(self):
        if not messagebox.askokcancel(
            "Tutup Aplikasi",
            "Apakah Anda yakin ingin menutup aplikasi?"
            + ("\n\nKoneksi serial akan ditutup otomatis."
               if self.is_reading else "")
        ):
            return

        self.is_reading = False
        if self.thread and self.thread.is_alive():
            self.thread.join(timeout=2.0)
        self.thread = None

        if self.serial_conn and self.serial_conn.is_open:
            try:
                self.serial_conn.reset_input_buffer()
                self.serial_conn.close()
            except Exception:
                pass
        self.serial_conn = None

        time.sleep(0.2)
        self.root.destroy()


# ════════════════════════════════════════════════════════════════
def main():
    root = tk.Tk()
    root.geometry("1140x760")
    root.minsize(1000, 680)
    app = ScaleApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()